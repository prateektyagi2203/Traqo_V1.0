# -*- coding: utf-8 -*-
"""
Traqo — RAG Powered Quantitative Candlestick Intelligence by Prateek Tyagi
================================================
Zero external dependencies. Uses Python's built-in http.server.
All HTML/CSS/JS is server-rendered — no React, no Flask, no build step.

Run:
    python paper_trading_dashboard.py
    → Opens http://localhost:8521
"""

import os
import sys
import json
import sqlite3
import webbrowser
import subprocess
import urllib.parse
import threading
import logging
from datetime import date, datetime, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn

# ---- ENCODING FIX: prevent UnicodeEncodeError on Windows when stdout/stderr is a CP1252 pipe ----
# This is safe on all platforms — reconfigure only runs if the stream supports it.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ---- YFINANCE DYNAMIC IMPORT ----
_HAS_YF = False
yf = None

def _ensure_yfinance():
    """Dynamically import yfinance if available. Called at runtime."""
    global _HAS_YF, yf
    if _HAS_YF and yf is not None:
        return True  # Already loaded
    try:
        import yfinance as yf_module
        yf = yf_module
        _HAS_YF = True
        return True
    except ImportError:
        _HAS_YF = False
        return False

# Try initial import
try:
    import yfinance as yf
    _HAS_YF = True
    print(f"✅ yfinance loaded successfully (v{yf.__version__})")
except ImportError as e:
    _HAS_YF = False
    print(f"⚠️  yfinance not available at startup (will retry at runtime)")
    print(f"   Solution: Run in virtual environment with: pip install yfinance")

# ---- Market Cap classification (based on index membership) ----
_LARGECAP_TICKERS = {
    "RELIANCE", "HDFCBANK", "ICICIBANK", "INFY", "TCS", "BHARTIARTL", "SBIN", "LT",
    "BAJFINANCE", "AXISBANK", "KOTAKBANK", "ITC", "HINDUNILVR", "MARUTI", "TATAMOTORS",
    "HCLTECH", "SUNPHARMA", "TITAN", "ADANIENT", "WIPRO", "TATASTEEL", "M&M", "NTPC",
    "POWERGRID", "ULTRACEMCO", "ASIANPAINT", "BAJAJFINSV", "COALINDIA", "NESTLEIND",
    "JSWSTEEL", "GRASIM", "ONGC", "DIVISLAB", "DRREDDY", "CIPLA", "APOLLOHOSP",
    "HEROMOTOCO", "EICHERMOT", "BPCL", "TECHM", "TATACONSUM", "BRITANNIA", "HINDALCO",
    "INDUSINDBK", "SBILIFE", "HDFCLIFE", "BAJAJ-AUTO", "ADANIPORTS", "SHRIRAMFIN",
    "ETERNAL", "TRENT",
    # Nifty Next 50
    "ABB", "ACC", "ADANIGREEN", "ADANIPOWER", "AMBUJACEM", "ATGL", "AUROPHARMA",
    "BAJAJHLDNG", "BANKBARODA", "BEL", "BERGEPAINT", "BIOCON", "BOSCHLTD", "CANBK",
    "CHOLAFIN", "COLPAL", "DABUR", "DLF", "GAIL", "GODREJCP", "HAL", "HAVELLS",
    "ICICIPRULI", "INDIGO", "IOC", "IRCTC", "IRFC", "JINDALSTEL", "JIOFIN", "LICI",
    "LTIM", "LTTS", "LUPIN", "MAXHEALTH", "MOTHERSON", "NAUKRI", "NHPC", "OBEROIRLTY",
    "OFSS", "PAYTM", "PFC", "PIDILITIND", "PNB", "POLYCAB", "RECLTD", "SBICARD",
    "SIEMENS", "SRF", "TATAPOWER",
}

def _get_cap(ticker: str) -> str:
    """Return LargeCap / MidCap based on index membership."""
    base = ticker.replace(".NS", "").replace(".BO", "").upper()
    return "LargeCap" if base in _LARGECAP_TICKERS else "MidCap"

_SECTOR_DISPLAY = {
    "auto": "Auto", "banking": "Banking", "capital_goods": "Capital Goods",
    "chemicals": "Chemicals", "consumer": "Consumer", "consumer_tech": "Consumer Tech",
    "energy": "Energy", "finance": "Finance", "fmcg": "FMCG", "it": "IT",
    "metals": "Metals", "pharma": "Pharma", "realty": "Realty", "unknown": "Other",
    "": "Other",
}

logger = logging.getLogger(__name__)

DB_PATH = "paper_trades/paper_trades.db"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ============================================================
# DATABASE QUERIES
# ============================================================
def get_db():
    conn = sqlite3.connect(os.path.join(SCRIPT_DIR, DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _regenerate_learned_rules():
    """Rebuild learned_rules.json from current feedback_log.json.

    Must be called after ANY trade removal to keep penalties/boosts consistent.
    """
    try:
        sys.path.insert(0, SCRIPT_DIR)
        from paper_trader import PaperTrader
        PaperTrader.regenerate_learned_rules()
    except Exception as e:
        logger.error(f"Failed to regenerate learned rules: {e}")


def cancel_trade(trade_id: int):
    """Cancel an open trade: mark CANCELLED in DB, remove from RAG feedback, rebuild learned rules."""
    # 1) Mark CANCELLED in SQLite
    db_full = os.path.join(SCRIPT_DIR, DB_PATH)
    conn = sqlite3.connect(db_full)
    conn.execute(
        "UPDATE trades SET status='CANCELLED', exit_date=?, exit_reason='user_cancelled',"
        " updated_at=datetime('now') WHERE id=? AND status='OPEN'",
        (date.today().isoformat(), trade_id)
    )
    conn.commit()
    conn.close()

    # 2) Erase from RAG feedback log
    fb_path = os.path.join(SCRIPT_DIR, "feedback", "feedback_log.json")
    if os.path.exists(fb_path):
        try:
            with open(fb_path, "r", encoding="utf-8") as f:
                feedback = json.load(f)
            pid = f"paper_{trade_id}"
            cleaned = [e for e in feedback if e.get("trade_id") != pid]
            if len(cleaned) < len(feedback):
                with open(fb_path, "w", encoding="utf-8") as f:
                    json.dump(cleaned, f, indent=2, default=str)
        except Exception:
            pass

    # 3) Regenerate learned rules (penalties/boosts) from remaining feedback
    _regenerate_learned_rules()


def cancel_trades_bulk(ids: list):
    """Cancel multiple open trades, erase their RAG feedback imprints, and rebuild learned rules."""
    db_full = os.path.join(SCRIPT_DIR, DB_PATH)
    conn = sqlite3.connect(db_full)
    today = date.today().isoformat()
    for trade_id in ids:
        conn.execute(
            "UPDATE trades SET status='CANCELLED', exit_date=?, exit_reason='user_cancelled',"
            " updated_at=datetime('now') WHERE id=? AND status='OPEN'",
            (today, trade_id)
        )
    conn.commit()
    conn.close()

    fb_path = os.path.join(SCRIPT_DIR, "feedback", "feedback_log.json")
    if os.path.exists(fb_path):
        try:
            with open(fb_path, "r", encoding="utf-8") as f:
                feedback = json.load(f)
            pids = {f"paper_{tid}" for tid in ids}
            cleaned = [e for e in feedback if e.get("trade_id") not in pids]
            if len(cleaned) < len(feedback):
                with open(fb_path, "w", encoding="utf-8") as f:
                    json.dump(cleaned, f, indent=2, default=str)
        except Exception:
            pass

    # Regenerate learned rules from remaining feedback
    _regenerate_learned_rules()


def purge_closed_trades(trade_ids: list):
    """Permanently delete closed/expired trades from DB, feedback, and learned rules.

    Unlike cancel (which only works on OPEN trades), this works on any status.
    Removes all traces: DB rows, position_monitoring, feedback_log, learned_rules.
    """
    try:
        sys.path.insert(0, SCRIPT_DIR)
        from paper_trader import PaperTrader
        return PaperTrader.purge_trades_complete(trade_ids)
    except Exception as e:
        logger.error(f"Failed to purge trades: {e}")
        return {"deleted": 0, "feedback_removed": 0, "error": str(e)}


def purge_trades_by_date(from_date: str, to_date: str = None):
    """Permanently delete all non-OPEN trades closed between from_date and to_date."""
    if not to_date:
        to_date = from_date
    db_full = os.path.join(SCRIPT_DIR, DB_PATH)
    conn = sqlite3.connect(db_full)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id FROM trades WHERE status NOT IN ('OPEN') AND exit_date >= ? AND exit_date <= ?",
        (from_date, to_date + "T23:59:59")
    ).fetchall()
    conn.close()
    if rows:
        trade_ids = [r["id"] for r in rows]
        return purge_closed_trades(trade_ids)
    return {"deleted": 0, "feedback_removed": 0}


def q_stats():
    c = get_db()
    open_n = c.execute("SELECT COUNT(*) FROM trades WHERE status='OPEN'").fetchone()[0]
    closed_n = c.execute("SELECT COUNT(*) FROM trades WHERE status NOT IN ('OPEN','CANCELLED')").fetchone()[0]
    wins = c.execute("SELECT COUNT(*) FROM trades WHERE status IN ('WON','EXPIRED_WIN')").fetchone()[0]
    losses = c.execute("SELECT COUNT(*) FROM trades WHERE status IN ('LOST','EXPIRED_LOSS')").fetchone()[0]
    avg_w = c.execute("SELECT AVG(actual_return_pct) FROM trades WHERE status IN ('WON','EXPIRED_WIN')").fetchone()[0] or 0
    avg_l = c.execute("SELECT AVG(actual_return_pct) FROM trades WHERE status IN ('LOST','EXPIRED_LOSS')").fetchone()[0] or 0
    tot_ret = c.execute("SELECT SUM(actual_return_pct) FROM trades WHERE status NOT IN ('OPEN','CANCELLED')").fetchone()[0] or 0
    wr = (wins / closed_n * 100) if closed_n else 0
    pf = (abs(avg_w * wins) / abs(avg_l * losses)) if (losses and avg_l) else 0
    last_scan = c.execute("SELECT MAX(scan_date) FROM scan_log").fetchone()[0] or "Never"
    today_entered = c.execute("SELECT COUNT(*) FROM trades WHERE entry_date=?", (date.today().isoformat(),)).fetchone()[0]
    # D6: Significance progress — count patterns with >= 30 closed trades (Wilson CI threshold)
    SIG_THRESHOLD = 30
    pat_counts = c.execute(
        "SELECT patterns, COUNT(*) as n FROM trades "
        "WHERE status NOT IN ('OPEN','CANCELLED') AND patterns IS NOT NULL "
        "GROUP BY patterns"
    ).fetchall()
    total_patterns = len(pat_counts)
    significant_patterns = sum(1 for _, n in pat_counts if n >= SIG_THRESHOLD)
    c.close()
    return {
        "open_trades": open_n, "closed_trades": closed_n, "total_trades": open_n + closed_n,
        "wins": wins, "losses": losses, "win_rate": round(wr, 1),
        "avg_win_pct": round(avg_w, 2), "avg_loss_pct": round(avg_l, 2),
        "profit_factor": round(pf, 2), "total_return_pct": round(tot_ret, 2),
        "last_scan": last_scan, "today_entered": today_entered,
        "significance_progress": {
            "significant_patterns": significant_patterns,
            "total_patterns": total_patterns,
            "threshold": SIG_THRESHOLD,
            "pct": round(significant_patterns / total_patterns * 100, 1) if total_patterns else 0,
        },
    }


def q_stats_for_range(from_date: str, to_date: str) -> dict:
    """Return analytics for closed trades whose exit_date falls within [from_date, to_date]."""
    c = get_db()
    base_where = "status NOT IN ('OPEN','CANCELLED') AND exit_date >= ? AND exit_date <= ?"
    params = (from_date, to_date)

    total = c.execute(f"SELECT COUNT(*) FROM trades WHERE {base_where}", params).fetchone()[0]
    wins = c.execute(f"SELECT COUNT(*) FROM trades WHERE {base_where} AND status IN ('WON','EXPIRED_WIN')", params).fetchone()[0]
    losses = c.execute(f"SELECT COUNT(*) FROM trades WHERE {base_where} AND status IN ('LOST','EXPIRED_LOSS')", params).fetchone()[0]
    avg_w = c.execute(f"SELECT AVG(actual_return_pct) FROM trades WHERE {base_where} AND status IN ('WON','EXPIRED_WIN')", params).fetchone()[0] or 0
    avg_l = c.execute(f"SELECT AVG(actual_return_pct) FROM trades WHERE {base_where} AND status IN ('LOST','EXPIRED_LOSS')", params).fetchone()[0] or 0
    tot_ret = c.execute(f"SELECT SUM(actual_return_pct) FROM trades WHERE {base_where}", params).fetchone()[0] or 0

    win_rate = round((wins / total * 100), 1) if total else 0
    pf = round((abs(avg_w * wins) / abs(avg_l * losses)), 2) if (losses and avg_l) else 0

    # Equity curve — cumulative sum of returns sorted by exit_date
    eq_rows = c.execute(
        f"SELECT exit_date, actual_return_pct FROM trades WHERE {base_where} ORDER BY exit_date, id",
        params
    ).fetchall()
    cumulative = 0.0
    equity_curve = []
    for row in eq_rows:
        cumulative += (row[1] or 0)
        equity_curve.append({"date": row[0], "cumulative_return": round(cumulative, 2)})

    # Pattern breakdown (top 10 by trade count)
    pat_rows = c.execute(f"""
        SELECT patterns,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(actual_return_pct) as avg_ret
        FROM trades WHERE {base_where}
        GROUP BY patterns ORDER BY total DESC LIMIT 10
    """, params).fetchall()
    pattern_breakdown = []
    for r in pat_rows:
        t = r[1]
        pattern_breakdown.append({
            "pattern": r[0] or "—",
            "total": t,
            "wins": r[2],
            "losses": r[3],
            "win_rate": round(r[2] / t * 100, 1) if t else 0,
            "avg_ret": round(r[4] or 0, 2),
        })

    # Horizon breakdown
    hz_rows = c.execute(f"""
        SELECT horizon_label, horizon_days,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(actual_return_pct) as avg_ret,
               SUM(actual_return_pct) as total_ret
        FROM trades WHERE {base_where}
        GROUP BY horizon_label ORDER BY horizon_days
    """, params).fetchall()
    horizon_breakdown = []
    for r in hz_rows:
        t = r[2]
        horizon_breakdown.append({
            "horizon": r[0] or "—",
            "total": t,
            "wins": r[3],
            "losses": r[4],
            "win_rate": round(r[3] / t * 100, 1) if t else 0,
            "avg_ret": round(r[5] or 0, 2),
            "total_ret": round(r[6] or 0, 2),
        })

    # Sector breakdown
    sec_rows = c.execute(f"""
        SELECT sector,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(actual_return_pct) as avg_ret
        FROM trades WHERE {base_where}
        GROUP BY sector ORDER BY total DESC
    """, params).fetchall()
    sector_breakdown = []
    for r in sec_rows:
        t = r[1]
        sector_display = _SECTOR_DISPLAY.get(r[0] or "", r[0] or "Other")
        sector_breakdown.append({
            "sector": sector_display,
            "total": t,
            "wins": r[2],
            "losses": r[3],
            "win_rate": round(r[2] / t * 100, 1) if t else 0,
            "avg_ret": round(r[4] or 0, 2),
        })

    # Streak — iterate trades sorted by exit_date ascending, find final consecutive run
    streak_rows = c.execute(
        f"SELECT status FROM trades WHERE {base_where} ORDER BY exit_date ASC, id ASC",
        params
    ).fetchall()
    streak_count = 0
    streak_type = "NONE"
    for row in reversed(streak_rows):
        s = row[0]
        is_win = s in ("WON", "EXPIRED_WIN")
        is_loss = s in ("LOST", "EXPIRED_LOSS")
        if streak_count == 0:
            if is_win:
                streak_type = "WIN"
                streak_count = 1
            elif is_loss:
                streak_type = "LOSS"
                streak_count = 1
            else:
                break
        else:
            if streak_type == "WIN" and is_win:
                streak_count += 1
            elif streak_type == "LOSS" and is_loss:
                streak_count += 1
            else:
                break

    # Trade IDs in range for client-side table highlighting
    id_rows = c.execute(
        f"SELECT id FROM trades WHERE {base_where}",
        params
    ).fetchall()
    trade_ids = [r[0] for r in id_rows]

    c.close()
    return {
        "stats": {
            "total": total, "wins": wins, "losses": losses,
            "win_rate": win_rate, "profit_factor": pf,
            "total_return": round(tot_ret, 2),
            "avg_win": round(avg_w, 2), "avg_loss": round(avg_l, 2),
        },
        "equity_curve": equity_curve,
        "pattern_breakdown": pattern_breakdown,
        "horizon_breakdown": horizon_breakdown,
        "sector_breakdown": sector_breakdown,
        "streak": {"type": streak_type, "count": streak_count},
        "trade_ids": trade_ids,
    }


def q_open_trades():
    c = get_db()
    rows = [dict(r) for r in c.execute("SELECT * FROM trades WHERE status='OPEN' ORDER BY entry_date DESC, ticker").fetchall()]
    c.close()
    return rows


def q_closed_trades():
    c = get_db()
    rows = [dict(r) for r in c.execute("SELECT * FROM trades WHERE status NOT IN ('OPEN','CANCELLED') ORDER BY exit_date DESC").fetchall()]
    c.close()
    return rows


def q_today_trades():
    c = get_db()
    today_str = date.today().isoformat()
    rows = [dict(r) for r in c.execute("SELECT * FROM trades WHERE entry_date=? ORDER BY ticker, horizon_days", (today_str,)).fetchall()]
    if not rows:
        last_date = c.execute("SELECT MAX(entry_date) FROM trades").fetchone()[0]
        if last_date:
            rows = [dict(r) for r in c.execute("SELECT * FROM trades WHERE entry_date=? ORDER BY ticker, horizon_days", (last_date,)).fetchall()]
    c.close()
    return rows


def q_stats_by_horizon():
    c = get_db()
    rows = [dict(r) for r in c.execute("""
        SELECT horizon_days, horizon_label,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN actual_return_pct END) as avg_win,
               AVG(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN actual_return_pct END) as avg_loss,
               SUM(actual_return_pct) as total_ret
        FROM trades WHERE status NOT IN ('OPEN','CANCELLED')
        GROUP BY horizon_days ORDER BY horizon_days
    """).fetchall()]
    for r in rows:
        t = r["wins"] + r["losses"]
        r["win_rate"] = round(r["wins"] / t * 100, 1) if t else 0
    c.close()
    return rows


def q_stats_by_pattern():
    c = get_db()
    rows = [dict(r) for r in c.execute("""
        SELECT patterns,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(actual_return_pct) as avg_ret
        FROM trades WHERE status NOT IN ('OPEN','CANCELLED')
        GROUP BY patterns ORDER BY total DESC LIMIT 20
    """).fetchall()]
    for r in rows:
        t = r["wins"] + r["losses"]
        r["win_rate"] = round(r["wins"] / t * 100, 1) if t else 0
    c.close()
    return rows


def q_stats_by_stock():
    c = get_db()
    rows = [dict(r) for r in c.execute("""
        SELECT ticker,
               MAX(sector) as sector,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(actual_return_pct) as avg_ret,
               SUM(actual_return_pct) as total_ret
        FROM trades WHERE status NOT IN ('OPEN','CANCELLED')
        GROUP BY ticker ORDER BY total DESC
    """).fetchall()]
    for r in rows:
        t = r["wins"] + r["losses"]
        r["win_rate"] = round(r["wins"] / t * 100, 1) if t else 0
    c.close()
    return rows


def q_stats_by_sector():
    c = get_db()
    rows = [dict(r) for r in c.execute("""
        SELECT sector,
               COUNT(DISTINCT ticker) as stocks,
               COUNT(*) as total,
               SUM(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN 1 ELSE 0 END) as losses,
               AVG(CASE WHEN status IN ('WON','EXPIRED_WIN') THEN actual_return_pct END) as avg_win,
               AVG(CASE WHEN status IN ('LOST','EXPIRED_LOSS') THEN actual_return_pct END) as avg_loss,
               AVG(actual_return_pct) as avg_ret,
               SUM(actual_return_pct) as total_ret,
               SUM(CASE WHEN actual_return_pct > 0 THEN actual_return_pct ELSE 0 END) as gross_win,
               SUM(CASE WHEN actual_return_pct < 0 THEN actual_return_pct ELSE 0 END) as gross_loss,
               MAX(actual_return_pct) as best_trade,
               MIN(actual_return_pct) as worst_trade
        FROM trades
        WHERE status NOT IN ('OPEN','CANCELLED')
          AND sector IS NOT NULL AND sector != '' AND sector != 'unknown'
        GROUP BY sector ORDER BY total DESC
    """).fetchall()]
    for r in rows:
        t = r["wins"] + r["losses"]
        r["win_rate"] = round(r["wins"] / t * 100, 1) if t else 0
        gw = r["gross_win"] or 0
        gl = abs(r["gross_loss"] or 0)
        if t == 0:
            r["profit_factor"] = None
        elif gl == 0:
            r["profit_factor"] = "∞"
        else:
            r["profit_factor"] = round(gw / gl, 2)
    # Dominant pattern per sector (split comma-separated patterns, count individually)
    pat_rows = [dict(r) for r in c.execute("""
        SELECT sector, patterns, COUNT(*) as cnt
        FROM trades
        WHERE status NOT IN ('OPEN','CANCELLED')
          AND sector IS NOT NULL AND sector != '' AND sector != 'unknown'
          AND patterns IS NOT NULL AND patterns != ''
        GROUP BY sector, patterns ORDER BY sector, cnt DESC
    """).fetchall()]
    sec_pat_counts = {}
    for row in pat_rows:
        sec = row["sector"]; cnt = row["cnt"]
        for p in [x.strip() for x in (row["patterns"] or "").split(",") if x.strip()]:
            sec_pat_counts.setdefault(sec, {})
            sec_pat_counts[sec][p] = sec_pat_counts[sec].get(p, 0) + cnt
    dom_pattern = {sec: max(counts, key=counts.get) for sec, counts in sec_pat_counts.items() if counts}
    # Dominant horizon per sector
    hz_rows = [dict(r) for r in c.execute("""
        SELECT sector, horizon_label, COUNT(*) as cnt
        FROM trades
        WHERE status NOT IN ('OPEN','CANCELLED')
          AND sector IS NOT NULL AND sector != '' AND sector != 'unknown'
          AND horizon_label IS NOT NULL AND horizon_label != ''
        GROUP BY sector, horizon_label ORDER BY sector, cnt DESC
    """).fetchall()]
    dom_horizon = {}
    for row in hz_rows:
        sec = row["sector"]
        if sec not in dom_horizon:
            dom_horizon[sec] = row["horizon_label"]
    c.close()
    for r in rows:
        sec = r["sector"]
        r["dom_pattern"] = dom_pattern.get(sec, "—")
        r["dom_horizon"] = dom_horizon.get(sec, "—")
    return rows


def q_scan_log():
    c = get_db()
    rows = [dict(r) for r in c.execute("SELECT * FROM scan_log ORDER BY scan_date DESC LIMIT 30").fetchall()]
    c.close()
    return rows


def q_daily_summaries():
    c = get_db()
    rows = [dict(r) for r in c.execute("SELECT * FROM daily_summary ORDER BY report_date DESC LIMIT 60").fetchall()]
    c.close()
    return rows


def get_engine_log():
    log_path = os.path.join(SCRIPT_DIR, "paper_trades/logs/paper_trader.log")
    if os.path.exists(log_path):
        with open(log_path, "r", encoding="utf-8") as f:
            return f.readlines()[-60:]
    return []


# ---- Live Engine Runner (background thread + polling) ----
_engine_state = {
    "running": False,
    "action": "",
    "output_lines": [],
    "done": False,
    "success": None,
    "started_at": None,
}
_engine_lock = threading.Lock()

STREAMLIT_NOISE = (
    "ScriptRunContext", "streamlit run", "Session state does not function",
    "missing ScriptRunContext", "warning can be ignored",
    "If you want to run a streamlit", "streamlit app",
)

def _is_noise(line: str) -> bool:
    return any(n in line for n in STREAMLIT_NOISE) or not line.strip()

def _engine_worker(action, extra_args=None):
    global _engine_state
    try:
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"
        
        # Use absolute path to virtual environment Python
        venv_python = r"C:\Users\tyagipra\Coding\Nifty_Data\.venv\Scripts\python.exe"
        python_executable = venv_python if os.path.exists(venv_python) else sys.executable
        
        # Ensure virtual environment paths are in PATH and PYTHONPATH
        venv_dir = r"C:\Users\tyagipra\Coding\Nifty_Data\.venv"
        venv_scripts = os.path.join(venv_dir, "Scripts")
        venv_lib = os.path.join(venv_dir, "Lib", "site-packages")
        
        # Update PATH to include venv scripts first
        env["PATH"] = venv_scripts + os.pathsep + env.get("PATH", "")
        # Clear and set PYTHONPATH to prioritize venv
        env["PYTHONPATH"] = venv_lib
        # Set VIRTUAL_ENV
        env["VIRTUAL_ENV"] = venv_dir
        # Unset PYTHONHOME to avoid conflicts
        env.pop("PYTHONHOME", None)
        
        # Debug information
        with _engine_lock:
            _engine_state["output_lines"].append(f"DEBUG: Using Python: {python_executable}")
            _engine_state["output_lines"].append(f"DEBUG: Venv exists: {os.path.exists(venv_python)}")
            _engine_state["output_lines"].append(f"DEBUG: VIRTUAL_ENV: {env.get('VIRTUAL_ENV', 'Not set')}")
            _engine_state["output_lines"].append(f"DEBUG: PYTHONPATH: {env.get('PYTHONPATH', 'Not set')}")
            
        # First test yfinance import in subprocess
        test_cmd = [python_executable, "-c", "import yfinance; print('yfinance import successful')"]
        with _engine_lock:
            _engine_state["output_lines"].append(f"DEBUG: Testing yfinance import...")
            
        test_proc = subprocess.Popen(
            test_cmd,
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, cwd=SCRIPT_DIR, env=env
        )
        test_output, _ = test_proc.communicate(timeout=10)
        
        with _engine_lock:
            _engine_state["output_lines"].append(f"DEBUG: yfinance test result: {test_output.strip()}")
            
        # If test failed, don't proceed with main command
        if test_proc.returncode != 0:
            with _engine_lock:
                _engine_state["output_lines"].append("ERROR: yfinance import test failed, aborting engine run")
                _engine_state["success"] = False
                _engine_state["done"] = True
                _engine_state["running"] = False
            return
        
        cmd_list = [python_executable, os.path.join(SCRIPT_DIR, "paper_trader.py"), action]
        if extra_args:
            cmd_list.extend(extra_args)
        proc = subprocess.Popen(
            cmd_list,
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, cwd=SCRIPT_DIR, env=env
        )
        for line in iter(proc.stdout.readline, ""):
            if not _is_noise(line):
                with _engine_lock:
                    _engine_state["output_lines"].append(line.rstrip())
        proc.stdout.close()
        ret = proc.wait(timeout=600)
        with _engine_lock:
            _engine_state["success"] = (ret == 0)
            _engine_state["done"] = True
            _engine_state["running"] = False
    except Exception as e:
        with _engine_lock:
            _engine_state["output_lines"].append(f"ERROR: {e}")
            _engine_state["success"] = False
            _engine_state["done"] = True
            _engine_state["running"] = False

def start_engine(action, extra_args=None):
    global _engine_state
    with _engine_lock:
        if _engine_state["running"]:
            return False  # already running
        _engine_state = {
            "running": True,
            "action": action,
            "output_lines": [f"Starting engine: {action}..."],
            "done": False,
            "success": None,
            "started_at": datetime.now().isoformat(),
        }
        # Add debug info about Python executable immediately
        _engine_state["output_lines"].append("DEBUG: start_engine called")
        venv_python = r"C:\Users\tyagipra\Coding\Nifty_Data\.venv\Scripts\python.exe"
        _engine_state["output_lines"].append(f"DEBUG: venv Python exists: {os.path.exists(venv_python)}")
        _engine_state["output_lines"].append(f"DEBUG: venv Python path: {venv_python}")
    t = threading.Thread(target=_engine_worker, args=(action, extra_args), daemon=True)
    t.start()
    return True

def get_engine_status():
    with _engine_lock:
        return {
            "running": _engine_state["running"],
            "done": _engine_state["done"],
            "success": _engine_state["success"],
            "action": _engine_state["action"],
            "lines": list(_engine_state["output_lines"]),
            "started_at": _engine_state["started_at"],
        }


PENDING_SIGNALS_FILE = os.path.join(SCRIPT_DIR, "paper_trades", "pending_signals.json")

def get_pending_signals():
    """Read the pending signals staging file if it exists."""
    if os.path.exists(PENDING_SIGNALS_FILE):
        try:
            with open(PENDING_SIGNALS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None


# ============================================================
# HTML HELPERS
# ============================================================
def _e(s):
    """Escape HTML."""
    if s is None:
        return "—"
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _price(v):
    if v is None:
        return "—"
    return f"₹{float(v):,.2f}"


def _pct(v, sign=True):
    if v is None:
        return "—"
    v = float(v)
    if sign and v > 0:
        return f"+{v:.2f}%"
    return f"{v:.2f}%"


def _date(d):
    if not d:
        return "—"
    try:
        return datetime.strptime(str(d), "%Y-%m-%d").strftime("%d %b %y")
    except Exception:
        return str(d)


def _significance_progress_bar(sp: dict) -> str:
    """D6: Render a progress bar showing how many patterns have reached statistical significance (n>=30)."""
    if not sp:
        return ""
    sig = sp.get("significant_patterns", 0)
    total = sp.get("total_patterns", 0)
    threshold = sp.get("threshold", 30)
    pct = sp.get("pct", 0)
    bar_w = min(100, int(pct))
    color = "bg-emerald-500" if pct >= 60 else "bg-amber-400" if pct >= 30 else "bg-red-400"
    return f'''<div class="glass rounded-xl border border-gray-200 shadow-sm p-4">
      <div class="flex items-center justify-between mb-1">
        <span class="text-xs font-semibold text-gray-600 uppercase tracking-wider">Statistical Significance Progress</span>
        <span class="text-xs text-gray-400">{sig}/{total} patterns &ge;{threshold} trades</span>
      </div>
      <div class="w-full bg-gray-100 rounded-full h-3">
        <div class="{color} h-3 rounded-full transition-all" style="width:{bar_w}%"></div>
      </div>
      <p class="text-xs text-gray-400 mt-1">{pct:.1f}% of seen patterns have enough data for Wilson CI confidence</p>
    </div>'''


def _ticker(t):
    return str(t).replace(".NS", "").replace(".BO", "") if t else ""


def _days_between(a, b):
    try:
        return (datetime.strptime(b, "%Y-%m-%d") - datetime.strptime(a, "%Y-%m-%d")).days
    except Exception:
        return 0


# ============================================================
# LIVE PRICE FETCH
# ============================================================
def fetch_live_prices(tickers: list) -> dict:
    """Fetch current prices for a list of NSE tickers via yfinance.
    Returns {ticker_raw: price} dict. Non-blocking best-effort.
    
    Attempts to dynamically import yfinance if not available at startup."""
    import pandas as pd
    prices = {}
    
    # Try dynamic import if initial import failed
    if not _HAS_YF:
        if not _ensure_yfinance():
            print(f"❌ [LIVE PRICE] yfinance not available (install via: pip install yfinance)")
            return prices
    
    if not tickers:
        print(f"⚠️ [LIVE PRICE] No tickers provided")
        return prices
    
    # Build unique Yahoo symbols — tickers may already have .NS/.BO suffix
    unique = list(set(tickers))
    yf_syms = []
    for t in unique:
        sym = t.strip()
        if not sym.endswith(".NS") and not sym.endswith(".BO") and not sym.startswith("^"):
            sym = sym + ".NS"
        yf_syms.append(sym)
    print(f"[LIVE PRICE] Fetching {len(yf_syms)} tickers: {yf_syms[:5]}...")
    try:
        data = yf.download(yf_syms, period="5d", interval="1d", progress=False, threads=True)
        if data is None:
            print("[LIVE PRICE] yf.download returned None")
            return prices
        if data.empty:
            print("[LIVE PRICE] yf.download returned empty DataFrame")
            return prices
        print(f"[LIVE PRICE] Got data shape={data.shape}, columns type={type(data.columns).__name__}")
        # yfinance 1.2+ always returns MultiIndex columns: (Price, Ticker)
        if isinstance(data.columns, pd.MultiIndex):
            close_df = data["Close"]
            print(f"[LIVE PRICE] Close columns: {list(close_df.columns)}")
            for raw_t, yf_t in zip(unique, yf_syms):
                try:
                    if yf_t in close_df.columns:
                        series = close_df[yf_t].dropna()
                        if not series.empty:
                            prices[raw_t] = float(series.iloc[-1])
                except Exception as ex:
                    print(f"[LIVE PRICE] Error parsing {yf_t}: {ex}")
        else:
            # Fallback for older yfinance (single ticker, flat columns)
            print(f"[LIVE PRICE] Flat columns: {list(data.columns)}")
            if "Close" in data.columns and not data["Close"].dropna().empty:
                prices[unique[0]] = float(data["Close"].dropna().iloc[-1])
        print(f"[LIVE PRICE] Got prices for {len(prices)}/{len(unique)} tickers")
    except Exception as e:
        print(f"[LIVE PRICE] Exception: {e}")
        import traceback
        traceback.print_exc()
    return prices


def _status_classes(s):
    m = {
        "OPEN": ("bg-blue-50 text-blue-700 border-blue-200", "Open"),
        "WON": ("bg-emerald-50 text-emerald-700 border-emerald-200", "Won"),
        "LOST": ("bg-red-50 text-red-700 border-red-200", "Lost"),
        "EXPIRED_WIN": ("bg-emerald-50 text-emerald-700 border-emerald-200", "Exp Win"),
        "EXPIRED_LOSS": ("bg-red-50 text-red-700 border-red-200", "Exp Loss"),
    }
    return m.get(s, ("bg-gray-100 text-gray-600", s))


# ============================================================
# HTML TEMPLATES
# ============================================================
def page_shell(title, active_tab, body_html):
    tabs = [
        ("dashboard", "Dashboard", "M4 5a1 1 0 011-1h4a1 1 0 011 1v5a1 1 0 01-1 1H5a1 1 0 01-1-1V5zM14 5a1 1 0 011-1h4a1 1 0 011 1v2a1 1 0 01-1 1h-4a1 1 0 01-1-1V5zM4 15a1 1 0 011-1h4a1 1 0 011 1v4a1 1 0 01-1 1H5a1 1 0 01-1-1v-4zM14 12a1 1 0 011-1h4a1 1 0 011 1v7a1 1 0 01-1 1h-4a1 1 0 01-1-1v-7z"),
        ("signals", "Today's Signals", "M13 10V3L4 14h7v7l9-11h-7z"),
        ("positions", "Open Positions", "M20 7l-8-4-8 4m16 0l-8 4m8-4v10l-8 4m0-10L4 7m8 4v10M4 7v10l8 4"),
        ("history", "Trade History", "M12 8v4l3 3m6-3a9 9 0 11-18 0 9 9 0 0118 0z"),
        ("market", "Market Indices", "M2 5a1 1 0 011-1h4a1 1 0 011 1v4a1 1 0 01-1 1H3a1 1 0 01-1-1V5zM12 5a1 1 0 011-1h4a1 1 0 011 1v4a1 1 0 01-1 1h-4a1 1 0 01-1-1V5zM2 15a1 1 0 011-1h4a1 1 0 011 1v4a1 1 0 01-1 1H3a1 1 0 01-1-1v-4zM12 15a1 1 0 011-1h4a1 1 0 011 1v4a1 1 0 01-1 1h-4a1 1 0 01-1-1v-4z"),
        ("performance", "Performance", "M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z"),
        ("engine", "Engine Control", "M10.325 4.317c.426-1.756 2.924-1.756 3.35 0a1.724 1.724 0 002.573 1.066c1.543-.94 3.31.826 2.37 2.37a1.724 1.724 0 001.066 2.573c1.756.426 1.756 2.924 0 3.35a1.724 1.724 0 00-1.066 2.573c.94 1.543-.826 3.31-2.37 2.37a1.724 1.724 0 00-2.573 1.066c-.426 1.756-2.924 1.756-3.35 0a1.724 1.724 0 00-2.573-1.066c-1.543.94-3.31-.826-2.37-2.37a1.724 1.724 0 00-1.066-2.573c-1.756-.426-1.756-2.924 0-3.35a1.724 1.724 0 001.066-2.573c-.94-1.543.826-3.31 2.37-2.37.996.608 2.296.07 2.572-1.065z"),
        ("feedback", "Feedback Loop", "M4 4v5h.582m15.356 2A8.001 8.001 0 004.582 9m0 0H9m11 11v-5h-.581m0 0a8.003 8.003 0 01-15.357-2m15.357 2H15"),
        ("filters", "Filters", "M3 4a1 1 0 011-1h16a1 1 0 011 1v2.586a1 1 0 01-.293.707l-6.414 6.414a1 1 0 00-.293.707V17l-4 4v-6.586a1 1 0 00-.293-.707L3.293 7.293A1 1 0 013 6.586V4z"),
    ]

    nav_items = ""
    for key, label, icon_path in tabs:
        is_active = key == active_tab
        cls = "bg-blue-50 text-blue-700 border border-blue-100" if is_active else "text-gray-500 hover:text-gray-800 hover:bg-gray-50"
        extra_icon = ""
        if key == "engine":
            extra_icon = f'<path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.8" d="M15 12a3 3 0 11-6 0 3 3 0 016 0z" />'
        nav_items += f'''
        <a href="/{key}" class="flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium transition-all {cls}">
          <svg class="w-5 h-5 flex-shrink-0" fill="none" stroke="currentColor" viewBox="0 0 24 24">
            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.8" d="{icon_path}" />{extra_icon}
          </svg>
          {label}
        </a>'''

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{_e(title)} — Traqo</title>
  <link rel="icon" type="image/png" href="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALQAAAC0CAYAAAA9zQYyAAAN20lEQVR4nO3dzXIVxxkG4HdUugUTUSy1gSyxXbHFT64BDBLCrgoXQFWC8WVYmFTlEgxYEnZWMdkmVRFIuHIHWuuHe8hkwRmdnp6vu7+e356Z99sY0TPf25Ie2nPmzJnJMKL6/Z/Oc+fgYiQL9MjMjR3dnD1yxTbmuHu27h4RGQCQqTKsjXJpG3f99tOaZipJVLITvWbh9U6UmD0ZzTAjl7d5v5sm8qQmde3ReS5BIGahx4CY7R5HCeEefCLXHhkrMTGPDrPd/2hvWNyDhV99dJ5LPxCziFnokTBms4aC3Xvo1cWKTMxxGcB4MJvbHPYMu7ewq8ahBTHHZQDjxGxWX7A7DzEhVwKJeRaYzX0O97uFvdJlc2JulgFMC3MG4Iv7p4HvqFl18q/FhlwJIuZZYrbrXQerdesrNDETszaji9W6VdDETMyxGW2jbmXJlyBXmhMzMft65MDb180PQRqv0MTs6UHMupxFxpf3mq/WjUATs6cHMetyrIymqGuDJmZPD2LW5TgyNhqgrgWamD09iFmX483IsXHvpBbqaNDE7OlBzLqcAOai6qBu5bQdMcdlAMQsjzc/gxcFmueZiTk2o9SjBuaNr+JWaTVoYibm2IxSjzor8+Kvb0SgVoEmZmKOzSj1aIC56KFFXesYmpjjMgBilsd1mGMqCJqXgDbLAIhZHo/HrFmlvaCJuVkGQMzyeP2VOYRafchBzHEZADHL4+0fZpjlBM3PANbPAIhZHm+OOQNw8657lQ6u0MQclwEQszzeDubQ708EzVsN1MsAiFkebx+za5V2rtDEHJcBELM83s/KXFQFNO9oFJ8BELM83i1maZWugCbmuAyAmOXx7ldmqUcJNG+cGJcBELM83h/mW9YqLR9DEzMx+3okglmqKmhiJmZfj4QxAwZo3mxclwEQszw+HOZbd5aHHcsVmpiJ2dcjUcz2Prrz0I4mxCxlELP5131ivtiWD+gJ15gwH+7+7uLPX2ydRWeUckaAudjm33+/nPnPQzuaELOUkQbmua3M9jYl0MQs9CBmYTxNzIABmpiFHsQsjKeLGViAJmahBzEL42ljBoAVPm5Y6DEVzJp5YDqYb985yd0X+BOzJ2McmDU1FczF+KqvyVQx//fFpYs/X//6fLkNMQvj48GM3HMtx1QxO3tMBXPg+6j0mBBmwAZNzJ6McWCe0zGzlFG5loOYpQxiNv86VcxAAXqmmENFzNV5pYwZAFaI2ZdBzOZfp44Z0LyxAmKWviRmR58BMQPAytww85hZGp8GZiBw5yRirn45Jsy6jOlgzhC8wJ+Yg/OwtidmKcc/rzaNiKCJufrlKDF75zE9zIDw1jcxV79MAfM741MoMXWwr9/vxr3T4DyKShEzIF7gPx/MlabClylgDv2s2qwxY0ZeucB/bpjdGaqcER1maGrsmIHSeWhiDs7D2n5KmDWVOmbg4jz0fDCHMlQ5xCyPS9UjZgBYmR/mcRwzE7MiR/jdrUoDwUkkiNm8aD+mfnulPwvw+YMzYpbGpRoAM5AvL04KNhAapYK51jFzjSJmYVyqgTADjo9gEXN8TRXz1tNLx/bY7s75eoqYAecbK0IRs7OmiHnr2yrkogrkezvn66WBgTEDFugpYf7s4Xl1IpEvAN8rjq/nhtmszaeXji9QJ4AZ0Nw5aYSYK02FLzVnM4LzUGzfBua+wMdgLmrz6aXjVDADoQv8R4vZnaHKSejUXJ+rdyzmoja/+7jf0JgB3wX+xOyeh6e/KicWcyKvFXyVAmZA8dAgYvb0mBDmvx38r9bqXNT979yre1+Yszzw0KCxYA5lqHJmjLnL6hMz4LjRjBgiTSIZzIFvNhQSe6aBmFXVN2bA8dAgYvZvH51DzO5tzGqIGbBuNCOGSJMgZn0OMbu3MasFzADK13IQc3z1gbmPU3ePN1bWw1u5a//7j2+wDIkZ0D6SAoljjl3VYl8A1tmmVczjWLaHxgzE3DmJmPXbzAzz/veLi5UGxgxo75xEzPptRoq57nnolDADmjsnzRXzgC8AiVnoofhxZHA9kgLjwjylN01SwLxnvMCT3gFM5QWglOO+wJ+Y9TVRzMASr10pYgaEQw5ijqw+MHdgXIPZValiRi5d4E/M+hohZtfx8uONlfU9jBszYJ+HngpmxTZjwqzJ0ZQPsyYndczAYoWeGubgPDz9VTkTwRzz7uAYMAPFeWhi1ucQs3sbswbADMB9Xw6xgdCImD09iNkY7x5zBtUjKaxKEXMLPwxNDjE7tjFrQMxA8JEUViWKuY0XgK3+H4CYjfH+MAPeR1JYRcy6HGI2xvvFDLjeWLGLmHU5xGyM948ZEB9JYdVcMSvPTROzND4MZiB0gf9IMfv6q3ISO5uh2aYpZnNO6nkkhhnwXeBPzN7qFXPgZ9sKZs08zEoQM+C6wJ+YvUXM0vjwmJE3vNEMkChm71yaZRCzNJ4G5gwNbjQDpIO58XEoMbvnYVbimIGaN5oB5ofZ3N6dQcylbXrGDAgPDZoK5sOajxL2FjEb4+lhBiJvNAMkill5zriLImahx0CYAeMTK8QcX31hzuD+yNReaS4f61+/rF38+Y93jQfST+A8c7GPaxvVjWYAYrarT8ybwpOo9owPr6p+ZjPADACrU8H8h62zcM4i4+3e8vj6y80zd87Ax8zEbJXCiPd6aGAcmFU5Izs11wZmc07eeXi2HxPm6nloewNiFsZHhFkzD7NGjhnw3TmJmIXx7jFLT24FiFmbIV/gP3HMxRmD4r9Twhz1j7eoiWBGLj0aeeKYv/7LJyXMxdcA8PKHD5VzukNi3ts5X495n2BuLwCljPKdkyaM2YTrqodPyriJWeiRMGbAvMB/5pjtevjkk+PtJ479iFmX0zNmoLhz0kQxhyA/3lhZ37h/VlqZ7dp+sgT36oclNGIO5AyAGQBW54jZvv7BPHbW4v7pmXyb2dJcidk9D7taMuI8bbdsMC7MoVX55fMP64835CvxMixWYZTxSvXgWzfuVDBXauKYM4Tu4D8hzC+fL1Zh5am5AjYQh3v3WRmmXcRsVctG3HfwHxFmzaocznCfzXj1bPEsEZTxSrW1GN/dqR6SELNVHRipnoeW9hwp5hfPP1xAbOtNE/PwwofbxLu7c77OY2arOjKSXd8+y8uD48D8TWBV9mE+2F8eQ9+4f+rMCM4Dy222Aiu3XcTcLMOVY13gP37MLxaHF6GV2Vs1Ts0Vhxiut6/NIuZmGb6c7Pr2WT43zLGHGa5K5WwGMRv9P90+y8eA+Zs/hyGXegyM2e6x+fTSMTE3y/DmFL+7T7dPc2kg2MDYnpilHHeGJoeYI3KM392qNBBsYGzf5UR9kAFidvVfjs8LM5AboEeE2YRc6kHMxvj8MAPCjWa8DYxGXU00ZlUu9SBmY3yemAFgdSyYbcilHsRsjM8XMyA9GtlXHU00dlUu9SBmY3zemIGYOycNgFmCXOpBzMY4MQPmo5F91cFE66zKpR7EbIwTc1HhOyf1jNkFudSDmI1xYjbLf+ekjib6419ltMTs6UHM7p2NHu5PrHQ9UaN8kEs9iNkYJ2aph7xC94D5xwViYg7kELN7Z7tHDmSfPahey9H5yqzIKPUgZmOcmMUei80qDw0iZsc8HBmaHGKOyGkB88XY51un3g7EXJ0XMcdleHNawvzmzZUsfAd/Idi7jV3ErMshZvfOdg/PZv47+AvB3m3sImZdDjG7d7Z7BL5vJ2hirs6LmOMyvDktYy62ke/gLwR7t7GLmHU5xOze2e6h/P2tAMD73bVMGiRmXQ4xR+R0hPnXN1cywFqhibk6L2KOy/DmdLgyF7UiDhKzKoeYI3J6wAwYoI+Kww5iVuUQc0ROx5iLww1AuJZD08Deh5gVOcTs3tnuUef3t6i4N1bsImZdDjG7d7Z7xP7+rO1LoI/21kr7EHNchtiDmN072z1qYP71n1dKQ7o3VoRGxKzIIWb3znaPhitzURXQR3trGTHHZYg9iNm9s92jJmZ7dQYUx9B2I2JW5BCze2e7R0src1Ei6EPrWLpoRMyKHGJ272z3aIBZWp0B7QpNzLocYnbvbPdogNn3vThBX6zSxKzLIWb3znaPhpj/4VidgRZuNFOaBDEb48Qs9uhoZS7KC/pwXziWdk2CmI1xYhZ79IwZiDzkuFipidkYJ2axxwCYgTrH0MRsjBOz2KPHY2a7okG/fV0960HMQg9i1uV4fnexqzNQ8yyHiZqYhR7ErMtpGTPQ4LTd29fFZabEXNqGmHU5HWAGGoAGgAPh8AMAMWtziLlVzEBD0ABw8PpyeQLErMsh5tYxOzPr1sZXJ7m2KTEH5mHXhDG3Abmoxiu0WQc/X/Z/fGtRxByYh13ErK5WQQPAf36+7J0gMQfmYRcxR1XrDc26sTgEKYURs3sedk0UcxeQi2p9hTbLXK2JOTAPu4i5VnXa3Kybd0+cPw5itmqCmLuG7J1Ll2XDJmarJobZ9ensrqp30EXdvHuSE7NVE8LcN+SiBgNd1C37UISYa2d4c3rCPBTkogYHbdatO0vcxByX4c3pGLN5f+ahK5mJ2HX7TuBFJDHrcjrCnBJis5KclKtu31kcdxOzLqclzG8SxSvV/wFp3TQhq3WHEgAAAABJRU5ErkJggg==">
  <script src="https://cdn.tailwindcss.com"></script>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
  <script>
    tailwind.config = {{
      darkMode: 'class',
      theme: {{
        extend: {{
          fontFamily: {{ sans: ['Inter', 'sans-serif'] }},
        }}
      }}
    }}
  </script>
  <style>
    body {{ font-family: 'Inter', sans-serif; }}
    .scrollbar-thin::-webkit-scrollbar {{ width: 6px; height: 6px; }}
    .scrollbar-thin::-webkit-scrollbar-track {{ background: transparent; }}
    .scrollbar-thin::-webkit-scrollbar-thumb {{ background: #cbd5e1; border-radius: 3px; }}
    .glass {{ background: #ffffff; border: 1px solid #e2e8f0; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
    @keyframes fade-in {{ from {{ opacity: 0; transform: translateY(8px); }} to {{ opacity: 1; transform: translateY(0); }} }}
    .fade-in {{ animation: fade-in 0.25s ease-out; }}
    @keyframes pulse-dot {{ 0%,100% {{ box-shadow: 0 0 0 0 rgba(16,185,129,0.4); }} 50% {{ box-shadow: 0 0 0 6px rgba(16,185,129,0); }} }}
    .pulse-dot {{ animation: pulse-dot 2s infinite; }}
  </style>
</head>
<body class="bg-gray-50 text-gray-800 min-h-screen">
  <!-- Sidebar -->
  <aside class="fixed top-0 left-0 h-screen w-60 bg-white border-r border-gray-200 flex flex-col z-50 shadow-sm">
    <div class="px-5 py-5 border-b border-gray-200">
      <div class="flex items-center gap-3">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALQAAAC0CAYAAAA9zQYyAAAN20lEQVR4nO3dzXIVxxkG4HdUugUTUSy1gSyxXbHFT64BDBLCrgoXQFWC8WVYmFTlEgxYEnZWMdkmVRFIuHIHWuuHe8hkwRmdnp6vu7+e356Z99sY0TPf25Ie2nPmzJnJMKL6/Z/Oc+fgYiQL9MjMjR3dnD1yxTbmuHu27h4RGQCQqTKsjXJpG3f99tOaZipJVLITvWbh9U6UmD0ZzTAjl7d5v5sm8qQmde3ReS5BIGahx4CY7R5HCeEefCLXHhkrMTGPDrPd/2hvWNyDhV99dJ5LPxCziFnokTBms4aC3Xvo1cWKTMxxGcB4MJvbHPYMu7ewq8ahBTHHZQDjxGxWX7A7DzEhVwKJeRaYzX0O97uFvdJlc2JulgFMC3MG4Iv7p4HvqFl18q/FhlwJIuZZYrbrXQerdesrNDETszaji9W6VdDETMyxGW2jbmXJlyBXmhMzMft65MDb180PQRqv0MTs6UHMupxFxpf3mq/WjUATs6cHMetyrIymqGuDJmZPD2LW5TgyNhqgrgWamD09iFmX483IsXHvpBbqaNDE7OlBzLqcAOai6qBu5bQdMcdlAMQsjzc/gxcFmueZiTk2o9SjBuaNr+JWaTVoYibm2IxSjzor8+Kvb0SgVoEmZmKOzSj1aIC56KFFXesYmpjjMgBilsd1mGMqCJqXgDbLAIhZHo/HrFmlvaCJuVkGQMzyeP2VOYRafchBzHEZADHL4+0fZpjlBM3PANbPAIhZHm+OOQNw8657lQ6u0MQclwEQszzeDubQ708EzVsN1MsAiFkebx+za5V2rtDEHJcBELM83s/KXFQFNO9oFJ8BELM83i1maZWugCbmuAyAmOXx7ldmqUcJNG+cGJcBELM83h/mW9YqLR9DEzMx+3okglmqKmhiJmZfj4QxAwZo3mxclwEQszw+HOZbd5aHHcsVmpiJ2dcjUcz2Prrz0I4mxCxlELP5131ivtiWD+gJ15gwH+7+7uLPX2ydRWeUckaAudjm33+/nPnPQzuaELOUkQbmua3M9jYl0MQs9CBmYTxNzIABmpiFHsQsjKeLGViAJmahBzEL42ljBoAVPm5Y6DEVzJp5YDqYb985yd0X+BOzJ2McmDU1FczF+KqvyVQx//fFpYs/X//6fLkNMQvj48GM3HMtx1QxO3tMBXPg+6j0mBBmwAZNzJ6McWCe0zGzlFG5loOYpQxiNv86VcxAAXqmmENFzNV5pYwZAFaI2ZdBzOZfp44Z0LyxAmKWviRmR58BMQPAytww85hZGp8GZiBw5yRirn45Jsy6jOlgzhC8wJ+Yg/OwtidmKcc/rzaNiKCJufrlKDF75zE9zIDw1jcxV79MAfM741MoMXWwr9/vxr3T4DyKShEzIF7gPx/MlabClylgDv2s2qwxY0ZeucB/bpjdGaqcER1maGrsmIHSeWhiDs7D2n5KmDWVOmbg4jz0fDCHMlQ5xCyPS9UjZgBYmR/mcRwzE7MiR/jdrUoDwUkkiNm8aD+mfnulPwvw+YMzYpbGpRoAM5AvL04KNhAapYK51jFzjSJmYVyqgTADjo9gEXN8TRXz1tNLx/bY7s75eoqYAecbK0IRs7OmiHnr2yrkogrkezvn66WBgTEDFugpYf7s4Xl1IpEvAN8rjq/nhtmszaeXji9QJ4AZ0Nw5aYSYK02FLzVnM4LzUGzfBua+wMdgLmrz6aXjVDADoQv8R4vZnaHKSejUXJ+rdyzmoja/+7jf0JgB3wX+xOyeh6e/KicWcyKvFXyVAmZA8dAgYvb0mBDmvx38r9bqXNT979yre1+Yszzw0KCxYA5lqHJmjLnL6hMz4LjRjBgiTSIZzIFvNhQSe6aBmFXVN2bA8dAgYvZvH51DzO5tzGqIGbBuNCOGSJMgZn0OMbu3MasFzADK13IQc3z1gbmPU3ePN1bWw1u5a//7j2+wDIkZ0D6SAoljjl3VYl8A1tmmVczjWLaHxgzE3DmJmPXbzAzz/veLi5UGxgxo75xEzPptRoq57nnolDADmjsnzRXzgC8AiVnoofhxZHA9kgLjwjylN01SwLxnvMCT3gFM5QWglOO+wJ+Y9TVRzMASr10pYgaEQw5ijqw+MHdgXIPZValiRi5d4E/M+hohZtfx8uONlfU9jBszYJ+HngpmxTZjwqzJ0ZQPsyYndczAYoWeGubgPDz9VTkTwRzz7uAYMAPFeWhi1ucQs3sbswbADMB9Xw6xgdCImD09iNkY7x5zBtUjKaxKEXMLPwxNDjE7tjFrQMxA8JEUViWKuY0XgK3+H4CYjfH+MAPeR1JYRcy6HGI2xvvFDLjeWLGLmHU5xGyM948ZEB9JYdVcMSvPTROzND4MZiB0gf9IMfv6q3ISO5uh2aYpZnNO6nkkhhnwXeBPzN7qFXPgZ9sKZs08zEoQM+C6wJ+YvUXM0vjwmJE3vNEMkChm71yaZRCzNJ4G5gwNbjQDpIO58XEoMbvnYVbimIGaN5oB5ofZ3N6dQcylbXrGDAgPDZoK5sOajxL2FjEb4+lhBiJvNAMkill5zriLImahx0CYAeMTK8QcX31hzuD+yNReaS4f61+/rF38+Y93jQfST+A8c7GPaxvVjWYAYrarT8ybwpOo9owPr6p+ZjPADACrU8H8h62zcM4i4+3e8vj6y80zd87Ax8zEbJXCiPd6aGAcmFU5Izs11wZmc07eeXi2HxPm6nloewNiFsZHhFkzD7NGjhnw3TmJmIXx7jFLT24FiFmbIV/gP3HMxRmD4r9Twhz1j7eoiWBGLj0aeeKYv/7LJyXMxdcA8PKHD5VzukNi3ts5X495n2BuLwCljPKdkyaM2YTrqodPyriJWeiRMGbAvMB/5pjtevjkk+PtJ479iFmX0zNmoLhz0kQxhyA/3lhZ37h/VlqZ7dp+sgT36oclNGIO5AyAGQBW54jZvv7BPHbW4v7pmXyb2dJcidk9D7taMuI8bbdsMC7MoVX55fMP64835CvxMixWYZTxSvXgWzfuVDBXauKYM4Tu4D8hzC+fL1Zh5am5AjYQh3v3WRmmXcRsVctG3HfwHxFmzaocznCfzXj1bPEsEZTxSrW1GN/dqR6SELNVHRipnoeW9hwp5hfPP1xAbOtNE/PwwofbxLu7c77OY2arOjKSXd8+y8uD48D8TWBV9mE+2F8eQ9+4f+rMCM4Dy222Aiu3XcTcLMOVY13gP37MLxaHF6GV2Vs1Ts0Vhxiut6/NIuZmGb6c7Pr2WT43zLGHGa5K5WwGMRv9P90+y8eA+Zs/hyGXegyM2e6x+fTSMTE3y/DmFL+7T7dPc2kg2MDYnpilHHeGJoeYI3KM392qNBBsYGzf5UR9kAFidvVfjs8LM5AboEeE2YRc6kHMxvj8MAPCjWa8DYxGXU00ZlUu9SBmY3yemAFgdSyYbcilHsRsjM8XMyA9GtlXHU00dlUu9SBmY3zemIGYOycNgFmCXOpBzMY4MQPmo5F91cFE66zKpR7EbIwTc1HhOyf1jNkFudSDmI1xYjbLf+ekjib6419ltMTs6UHM7p2NHu5PrHQ9UaN8kEs9iNkYJ2aph7xC94D5xwViYg7kELN7Z7tHDmSfPahey9H5yqzIKPUgZmOcmMUei80qDw0iZsc8HBmaHGKOyGkB88XY51un3g7EXJ0XMcdleHNawvzmzZUsfAd/Idi7jV3ErMshZvfOdg/PZv47+AvB3m3sImZdDjG7d7Z7BL5vJ2hirs6LmOMyvDktYy62ke/gLwR7t7GLmHU5xOze2e6h/P2tAMD73bVMGiRmXQ4xR+R0hPnXN1cywFqhibk6L2KOy/DmdLgyF7UiDhKzKoeYI3J6wAwYoI+Kww5iVuUQc0ROx5iLww1AuJZD08Deh5gVOcTs3tnuUef3t6i4N1bsImZdDjG7d7Z7xP7+rO1LoI/21kr7EHNchtiDmN072z1qYP71n1dKQ7o3VoRGxKzIIWb3znaPhitzURXQR3trGTHHZYg9iNm9s92jJmZ7dQYUx9B2I2JW5BCze2e7R0src1Ei6EPrWLpoRMyKHGJ272z3aIBZWp0B7QpNzLocYnbvbPdogNn3vThBX6zSxKzLIWb3znaPhpj/4VidgRZuNFOaBDEb48Qs9uhoZS7KC/pwXziWdk2CmI1xYhZ79IwZiDzkuFipidkYJ2axxwCYgTrH0MRsjBOz2KPHY2a7okG/fV0960HMQg9i1uV4fnexqzNQ8yyHiZqYhR7ErMtpGTPQ4LTd29fFZabEXNqGmHU5HWAGGoAGgAPh8AMAMWtziLlVzEBD0ABw8PpyeQLErMsh5tYxOzPr1sZXJ7m2KTEH5mHXhDG3Abmoxiu0WQc/X/Z/fGtRxByYh13ErK5WQQPAf36+7J0gMQfmYRcxR1XrDc26sTgEKYURs3sedk0UcxeQi2p9hTbLXK2JOTAPu4i5VnXa3Kybd0+cPw5itmqCmLuG7J1Ll2XDJmarJobZ9ensrqp30EXdvHuSE7NVE8LcN+SiBgNd1C37UISYa2d4c3rCPBTkogYHbdatO0vcxByX4c3pGLN5f+ahK5mJ2HX7TuBFJDHrcjrCnBJis5KclKtu31kcdxOzLqclzG8SxSvV/wFp3TQhq3WHEgAAAABJRU5ErkJggg==" alt="Traqo" class="w-9 h-9 rounded-xl">
        <div>
          <div class="text-base font-bold text-gray-800">Traqo</div>
          <div class="text-[10px] text-blue-500 font-medium tracking-wide uppercase">Quantitative Candlestick Intelligence</div>
        </div>
      </div>
    </div>
    <nav class="flex-1 px-3 py-4 space-y-1 overflow-y-auto scrollbar-thin">
      {nav_items}
    </nav>
    <div class="px-4 py-4 border-t border-gray-200">
      <div class="flex items-center gap-2 text-xs text-gray-400 mb-2">
        <div class="w-2 h-2 rounded-full bg-emerald-500 pulse-dot"></div>
        System Online
      </div>
      <div class="text-[10px] text-gray-300">by Prateek Tyagi</div>
    </div>
  </aside>

  <!-- Main Content -->
  <main class="ml-60 min-h-screen flex flex-col">
    <div class="p-8 max-w-[1400px] mx-auto fade-in flex-1 w-full">
      {body_html}
    </div>
    <footer class="border-t border-gray-200 py-4 px-8 text-center">
      <p class="text-sm text-blue-400"><span class="font-bold text-blue-600">TRAQO</span> &mdash; RAG Powered Quantitative Candlestick Intelligence by <span class="font-medium text-blue-500">Prateek Tyagi</span></p>
    </footer>
  </main>
</body>
</html>'''


def stat_card(label, value, subtitle="", color="indigo"):
    color_map = {
        "indigo": "bg-white border-gray-200 shadow-sm",
        "green": "bg-white border-emerald-200 shadow-sm",
        "red": "bg-white border-red-200 shadow-sm",
        "amber": "bg-white border-amber-200 shadow-sm",
        "cyan": "bg-white border-blue-200 shadow-sm",
    }
    label_color = {
        "indigo": "text-blue-600",
        "green": "text-emerald-600",
        "red": "text-red-600",
        "amber": "text-amber-600",
        "cyan": "text-blue-600",
    }
    sub_html = f'<p class="mt-1 text-xs text-gray-400">{_e(subtitle)}</p>' if subtitle else ""
    return f'''<div class="rounded-xl {color_map.get(color, color_map["indigo"])} border p-5">
      <p class="text-xs font-medium uppercase tracking-wider {label_color.get(color, "text-blue-600")}">{_e(label)}</p>
      <p class="mt-2 text-2xl font-bold text-gray-800">{_e(str(value))}</p>
      {sub_html}
    </div>'''


def badge(text, variant="default"):
    styles = {
        "default": "bg-gray-100 text-gray-600 border border-gray-200",
        "success": "bg-emerald-50 text-emerald-700 border border-emerald-200",
        "danger": "bg-red-50 text-red-700 border border-red-200",
        "warning": "bg-amber-50 text-amber-700 border border-amber-200",
        "info": "bg-blue-50 text-blue-700 border border-blue-200",
        "bullish": "bg-emerald-50 text-emerald-600",
    }
    return f'<span class="inline-flex items-center px-2.5 py-0.5 rounded-full text-xs font-medium {styles.get(variant, styles["default"])}">{_e(text)}</span>'


def status_badge(status):
    cls, label = _status_classes(status)
    return f'<span class="inline-flex items-center px-2.5 py-0.5 rounded-full text-xs font-medium border {cls}">{_e(label)}</span>'


# ============================================================
# PAGE RENDERERS
# ============================================================
def render_dashboard():
    s = q_stats()
    cards = f'''
    <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-4 gap-4">
      {stat_card("Open Trades", s["open_trades"], f'{s["today_entered"]} entered today', "indigo")}
      {stat_card("Closed Trades", s["closed_trades"], f'{s["wins"]}W / {s["losses"]}L', "cyan")}
      {stat_card("Win Rate", f'{s["win_rate"]}%', "of closed trades" if s["closed_trades"] else "no closed trades",
                 "green" if s["win_rate"] >= 55 else "amber" if s["win_rate"] >= 45 else "red")}
      {stat_card("Profit Factor", s["profit_factor"],
                 f'Avg W: {_pct(s["avg_win_pct"])} | L: {_pct(s["avg_loss_pct"])}',
                 "green" if s["profit_factor"] >= 1.5 else "amber" if s["profit_factor"] >= 1.0 else "red")}
    </div>
    <div class="grid grid-cols-1 md:grid-cols-3 gap-4 mt-4">
      {stat_card("Total Return", _pct(s["total_return_pct"]), "", "green" if s["total_return_pct"] >= 0 else "red")}
      {stat_card("Last Scan", _date(s["last_scan"]) if s["last_scan"] != "Never" else "Never", "", "indigo")}
      {stat_card("Total Trades", s["total_trades"], "all time", "cyan")}
    </div>
    <div class="grid grid-cols-1 md:grid-cols-1 gap-4 mt-4">
      {_significance_progress_bar(s.get("significance_progress", {}))}
    </div>'''

    date_range_html = '''
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <div class="glass rounded-xl border border-gray-200 shadow-sm p-5 mt-6">
      <div class="mb-4">
        <h3 class="text-sm font-semibold text-gray-700 uppercase tracking-wider">Date Range Analysis</h3>
        <p class="text-xs text-gray-400 mt-0.5">Analyse closed trade performance for a specific exit-date window</p>
      </div>
      <div class="flex flex-wrap items-end gap-4">
        <div>
          <label class="block text-xs font-medium text-gray-500 mb-1">From</label>
          <input type="date" id="dr-from" class="px-3 py-2 text-sm border border-gray-200 rounded-lg bg-white text-gray-800 focus:outline-none focus:ring-2 focus:ring-blue-400">
        </div>
        <div>
          <label class="block text-xs font-medium text-gray-500 mb-1">To</label>
          <input type="date" id="dr-to" class="px-3 py-2 text-sm border border-gray-200 rounded-lg bg-white text-gray-800 focus:outline-none focus:ring-2 focus:ring-blue-400">
        </div>
        <button onclick="executeAnalysis()" id="dr-execute-btn" class="px-5 py-2 rounded-lg bg-blue-600 hover:bg-blue-700 text-white text-sm font-medium transition shadow-sm disabled:opacity-60">Execute</button>
        <button onclick="clearDateFilter()" id="dr-clear-btn" style="display:none" class="px-4 py-2 rounded-lg bg-white border border-gray-200 hover:bg-gray-50 text-gray-600 text-sm transition">Clear</button>
      </div>
    </div>
    <div id="date-range-panel" style="display:none" class="mt-6">
      <div class="flex items-baseline gap-3 mb-4">
        <h3 class="text-lg font-bold text-gray-800" id="dr-range-label">&mdash;</h3>
        <span class="text-sm text-gray-400" id="dr-range-subtitle"></span>
      </div>
      <div class="grid grid-cols-4 gap-4 mb-4">
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-gray-500 uppercase tracking-wider">Closed</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-total">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-emerald-600 uppercase tracking-wider">Wins</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-wins">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-red-500 uppercase tracking-wider">Losses</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-losses">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-amber-600 uppercase tracking-wider">Win Rate</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-wr">&mdash;</p></div>
      </div>
      <div class="grid grid-cols-4 gap-4 mb-6">
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-blue-600 uppercase tracking-wider">Profit Factor</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-pf">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-indigo-600 uppercase tracking-wider">Total Return</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-ret">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-emerald-600 uppercase tracking-wider">Avg Win</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-avgw">&mdash;</p></div>
        <div class="glass rounded-xl p-4"><p class="text-xs font-medium text-red-500 uppercase tracking-wider">Avg Loss</p><p class="mt-1 text-2xl font-bold text-gray-800" id="dr-avgl">&mdash;</p></div>
      </div>
      <div class="grid grid-cols-4 gap-4 mb-6">
        <div class="col-span-3 glass rounded-xl p-5">
          <p class="text-xs font-semibold text-gray-500 uppercase tracking-wider mb-3">Equity Curve (Cumulative Return %)</p>
          <canvas id="equityChart" height="100"></canvas>
        </div>
        <div class="glass rounded-xl p-5 flex flex-col items-center justify-center">
          <p class="text-xs font-semibold text-gray-500 uppercase tracking-wider mb-3 text-center">Current Streak</p>
          <div id="dr-streak-badge" class="text-4xl font-black text-gray-300">&mdash;</div>
          <div id="dr-streak-label" class="mt-2 text-sm text-gray-400"></div>
        </div>
      </div>
      <div class="grid grid-cols-3 gap-4">
        <div class="glass rounded-xl overflow-hidden">
          <div class="px-4 py-3 bg-gray-50 border-b border-gray-200"><p class="text-xs font-semibold text-gray-600 uppercase tracking-wider">Top Patterns</p></div>
          <div class="overflow-x-auto scrollbar-thin"><table class="w-full text-xs">
            <thead><tr class="border-b border-gray-100 bg-gray-50 text-gray-500"><th class="px-3 py-2 text-left">Pattern</th><th class="px-2 py-2 text-right">W</th><th class="px-2 py-2 text-right">L</th><th class="px-2 py-2 text-right">Win%</th><th class="px-2 py-2 text-right">Avg%</th></tr></thead>
            <tbody id="dr-pattern-tbody"></tbody>
          </table></div>
        </div>
        <div class="glass rounded-xl overflow-hidden">
          <div class="px-4 py-3 bg-gray-50 border-b border-gray-200"><p class="text-xs font-semibold text-gray-600 uppercase tracking-wider">Horizon Breakdown</p></div>
          <div class="overflow-x-auto scrollbar-thin"><table class="w-full text-xs">
            <thead><tr class="border-b border-gray-100 bg-gray-50 text-gray-500"><th class="px-3 py-2 text-left">Horizon</th><th class="px-2 py-2 text-right">W</th><th class="px-2 py-2 text-right">L</th><th class="px-2 py-2 text-right">Win%</th><th class="px-2 py-2 text-right">Total%</th></tr></thead>
            <tbody id="dr-horizon-tbody"></tbody>
          </table></div>
        </div>
        <div class="glass rounded-xl overflow-hidden">
          <div class="px-4 py-3 bg-gray-50 border-b border-gray-200"><p class="text-xs font-semibold text-gray-600 uppercase tracking-wider">Sector Breakdown</p></div>
          <div class="overflow-x-auto scrollbar-thin"><table class="w-full text-xs">
            <thead><tr class="border-b border-gray-100 bg-gray-50 text-gray-500"><th class="px-3 py-2 text-left">Sector</th><th class="px-2 py-2 text-right">W</th><th class="px-2 py-2 text-right">L</th><th class="px-2 py-2 text-right">Win%</th><th class="px-2 py-2 text-right">Avg%</th></tr></thead>
            <tbody id="dr-sector-tbody"></tbody>
          </table></div>
        </div>
      </div>
    </div>
    <script>
    var activeDateFrom = null, activeDateTo = null;
    var equityChart = null;
    function executeAnalysis() {
      var from = document.getElementById('dr-from').value;
      var to   = document.getElementById('dr-to').value;
      if (!from || !to) { alert('Please select both From and To dates.'); return; }
      if (from > to) { alert('From date must be on or before To date.'); return; }
      var btn = document.getElementById('dr-execute-btn');
      btn.textContent = 'Loading\u2026';
      btn.disabled = true;
      fetch('/api/stats-by-date?from=' + from + '&to=' + to)
        .then(function(r) { return r.json(); })
        .then(function(data) {
          if (data.error) { alert('Error: ' + data.error); return; }
          var s = data.stats;
          var fmtD = function(iso) {
            if (!iso) return '\u2014';
            var p = iso.split('-');
            var months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
            return p[2] + ' ' + months[parseInt(p[1],10)-1] + ' ' + p[0].slice(2);
          };
          document.getElementById('dr-range-label').textContent = fmtD(from) + ' \u2192 ' + fmtD(to);
          document.getElementById('dr-range-subtitle').textContent = s.total + ' closed trade' + (s.total !== 1 ? 's' : '') + ' in selected range';
          document.getElementById('dr-total').textContent   = s.total;
          document.getElementById('dr-wins').textContent    = s.wins;
          document.getElementById('dr-losses').textContent  = s.losses;
          var wrEl = document.getElementById('dr-wr');
          wrEl.textContent = s.win_rate + '%';
          wrEl.className   = 'mt-1 text-2xl font-bold ' + (s.win_rate >= 55 ? 'text-emerald-600' : s.win_rate >= 45 ? 'text-amber-600' : 'text-red-600');
          var pfEl = document.getElementById('dr-pf');
          pfEl.textContent = s.profit_factor;
          pfEl.className   = 'mt-1 text-2xl font-bold ' + (s.profit_factor >= 1.5 ? 'text-emerald-600' : s.profit_factor >= 1.0 ? 'text-amber-600' : 'text-red-600');
          var retEl = document.getElementById('dr-ret');
          retEl.textContent = (s.total_return >= 0 ? '+' : '') + s.total_return.toFixed(2) + '%';
          retEl.className   = 'mt-1 text-2xl font-bold ' + (s.total_return >= 0 ? 'text-emerald-600' : 'text-red-600');
          document.getElementById('dr-avgw').textContent = (s.avg_win >= 0 ? '+' : '') + s.avg_win.toFixed(2) + '%';
          document.getElementById('dr-avgl').textContent = s.avg_loss.toFixed(2) + '%';
          var ctx = document.getElementById('equityChart').getContext('2d');
          if (equityChart) { equityChart.destroy(); }
          var eq = data.equity_curve;
          var labels = eq.map(function(p) { return p.date; });
          var values = eq.map(function(p) { return p.cumulative_return; });
          var lastVal   = values.length ? values[values.length - 1] : 0;
          var lineColor = lastVal >= 0 ? 'rgb(16,185,129)' : 'rgb(239,68,68)';
          var fillColor = lastVal >= 0 ? 'rgba(16,185,129,0.1)' : 'rgba(239,68,68,0.08)';
          equityChart = new Chart(ctx, {
            type: 'line',
            data: { labels: labels, datasets: [{ label: 'Cumulative Return %', data: values, borderColor: lineColor, backgroundColor: fillColor, borderWidth: 2, pointRadius: eq.length > 60 ? 0 : 3, pointHoverRadius: 5, fill: true, tension: 0.3 }] },
            options: { responsive: true, interaction: { mode: 'index', intersect: false }, plugins: { legend: { display: false }, tooltip: { callbacks: { label: function(c) { return ' ' + c.parsed.y.toFixed(2) + '%'; } } } }, scales: { x: { ticks: { maxTicksLimit: 8, font: { size: 10 } }, grid: { display: false } }, y: { ticks: { font: { size: 10 }, callback: function(v) { return v.toFixed(1) + '%'; } }, grid: { color: 'rgba(0,0,0,0.04)' } } } }
          });
          var sk = data.streak;
          var skBadge = document.getElementById('dr-streak-badge');
          var skLabel = document.getElementById('dr-streak-label');
          if (sk.count > 0) {
            var isWin = sk.type === 'WIN';
            skBadge.textContent = sk.count + (isWin ? 'W' : 'L');
            skBadge.className   = 'text-4xl font-black ' + (isWin ? 'text-emerald-500' : 'text-red-500');
            skLabel.textContent = isWin ? 'consecutive wins' : 'consecutive losses';
          } else {
            skBadge.textContent = '\u2014';
            skBadge.className   = 'text-4xl font-black text-gray-300';
            skLabel.textContent = 'No closed trades in range';
          }
          var patHtml = '';
          data.pattern_breakdown.forEach(function(r) {
            var wc = r.win_rate >= 60 ? 'text-emerald-600 font-semibold' : r.win_rate >= 45 ? 'text-amber-600' : 'text-red-500';
            var rc = r.avg_ret >= 0 ? 'text-emerald-600' : 'text-red-500';
            patHtml += '<tr class="border-b border-gray-100 hover:bg-gray-50"><td class="px-3 py-2 text-gray-700 max-w-[110px] truncate" title="' + r.pattern + '">' + r.pattern + '</td><td class="px-2 py-2 text-right text-emerald-600">' + r.wins + '</td><td class="px-2 py-2 text-right text-red-500">' + r.losses + '</td><td class="px-2 py-2 text-right ' + wc + '">' + r.win_rate + '%</td><td class="px-2 py-2 text-right ' + rc + '">' + (r.avg_ret >= 0 ? '+' : '') + r.avg_ret + '%</td></tr>';
          });
          document.getElementById('dr-pattern-tbody').innerHTML = patHtml || '<tr><td colspan="5" class="px-3 py-4 text-center text-gray-400 text-xs">No data</td></tr>';
          var hzHtml = '';
          data.horizon_breakdown.forEach(function(r) {
            var wc = r.win_rate >= 60 ? 'text-emerald-600 font-semibold' : r.win_rate >= 45 ? 'text-amber-600' : 'text-red-500';
            var rc = r.total_ret >= 0 ? 'text-emerald-600' : 'text-red-500';
            hzHtml += '<tr class="border-b border-gray-100 hover:bg-gray-50"><td class="px-3 py-2 text-gray-700 font-medium">' + r.horizon + '</td><td class="px-2 py-2 text-right text-emerald-600">' + r.wins + '</td><td class="px-2 py-2 text-right text-red-500">' + r.losses + '</td><td class="px-2 py-2 text-right ' + wc + '">' + r.win_rate + '%</td><td class="px-2 py-2 text-right ' + rc + '">' + (r.total_ret >= 0 ? '+' : '') + r.total_ret + '%</td></tr>';
          });
          document.getElementById('dr-horizon-tbody').innerHTML = hzHtml || '<tr><td colspan="5" class="px-3 py-4 text-center text-gray-400 text-xs">No data</td></tr>';
          var secHtml = '';
          data.sector_breakdown.forEach(function(r) {
            var wc = r.win_rate >= 60 ? 'text-emerald-600 font-semibold' : r.win_rate >= 45 ? 'text-amber-600' : 'text-red-500';
            var rc = r.avg_ret >= 0 ? 'text-emerald-600' : 'text-red-500';
            secHtml += '<tr class="border-b border-gray-100 hover:bg-gray-50"><td class="px-3 py-2 text-gray-700">' + r.sector + '</td><td class="px-2 py-2 text-right text-emerald-600">' + r.wins + '</td><td class="px-2 py-2 text-right text-red-500">' + r.losses + '</td><td class="px-2 py-2 text-right ' + wc + '">' + r.win_rate + '%</td><td class="px-2 py-2 text-right ' + rc + '">' + (r.avg_ret >= 0 ? '+' : '') + r.avg_ret + '%</td></tr>';
          });
          document.getElementById('dr-sector-tbody').innerHTML = secHtml || '<tr><td colspan="5" class="px-3 py-4 text-center text-gray-400 text-xs">No data</td></tr>';
          document.getElementById('date-range-panel').style.display = '';
          document.getElementById('dr-clear-btn').style.display = '';
        })
        .catch(function(e) { alert('Request failed: ' + e); })
        .finally(function() { btn.textContent = 'Execute'; btn.disabled = false; });
    }
    function clearDateFilter() {
      document.getElementById('date-range-panel').style.display = 'none';
      document.getElementById('dr-clear-btn').style.display = 'none';
      if (equityChart) { equityChart.destroy(); equityChart = null; }
    }
    (function() {
      var today = new Date();
      var from  = new Date(today);
      from.setDate(from.getDate() - 30);
      var fmt = function(d) { return d.getFullYear() + '-' + String(d.getMonth()+1).padStart(2,'0') + '-' + String(d.getDate()).padStart(2,'0'); };
      var fromEl = document.getElementById('dr-from');
      var toEl   = document.getElementById('dr-to');
      if (fromEl) fromEl.value = fmt(from);
      if (toEl)   toEl.value   = fmt(today);
    })();
    </script>'''

    body = f'''
    <div class="flex items-center justify-between mb-6">
      <div>
        <h2 class="text-2xl font-bold text-gray-800">Dashboard</h2>
        <p class="text-sm text-gray-500 mt-1">Overview of your paper trading engine</p>
      </div>
      <a href="/dashboard" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-white border border-gray-200 hover:bg-gray-50 text-gray-600 text-sm transition shadow-sm">
        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 4v5h.582m15.356 2A8.001 8.001 0 004.582 9m0 0H9m11 11v-5h-.581m0 0a8.003 8.003 0 01-15.357-2m15.357 2H15"/></svg>
        Refresh
      </a>
    </div>
    {cards}
    {date_range_html}'''

    return page_shell("Dashboard", "dashboard", body)


def _history_xlsx_bytes():
    """Generate Excel (.xlsx) bytes for all closed trades."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    trades = q_closed_trades()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closed Trades"

    headers = [
        "ID", "Ticker", "Status", "Horizon", "Direction",
        "Entry Price", "Exit Price", "Target Price", "SL Price",
        "Target %", "SL %", "RR Ratio",
        "Predicted Win%", "Actual Return%",
        "Exit Reason", "Entry Date", "Exit Date",
        "Patterns", "Confidence", "Sector",
    ]

    # Header row styling
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2563EB")  # blue-600
    hdr_align = Alignment(horizontal="center")
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    win_fill  = PatternFill("solid", fgColor="D1FAE5")  # green-100
    loss_fill = PatternFill("solid", fgColor="FEE2E2")  # red-100

    for t in trades:
        status = t.get("status", "")
        row = [
            t.get("id"),
            _ticker(t.get("ticker", "")),
            status,
            t.get("horizon_label", ""),
            t.get("direction", ""),
            t.get("entry_price"),
            t.get("exit_price"),
            t.get("target_price"),
            t.get("sl_price"),
            t.get("target_pct"),
            t.get("sl_pct"),
            t.get("rr_ratio"),
            t.get("predicted_win_rate"),
            t.get("actual_return_pct"),
            t.get("exit_reason", ""),
            t.get("entry_date", ""),
            t.get("exit_date", ""),
            t.get("patterns", ""),
            t.get("confidence", ""),
            t.get("sector", ""),
        ]
        ws.append(row)
        # Colour row by outcome
        if status in ("WON", "EXPIRED_WIN"):
            fill = win_fill
        elif status in ("LOST", "EXPIRED_LOSS"):
            fill = loss_fill
        else:
            fill = None
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_signals():
    trades = q_today_trades()
    entry_date = trades[0]["entry_date"] if trades else date.today().isoformat()

    # group by horizon
    by_hz = {}
    for t in trades:
        h = t.get("horizon_label") or f'{t["horizon_days"]}d'
        if h not in by_hz:
            by_hz[h] = []
        by_hz[h].append(t)

    summary_badges = f'''
    <div class="flex gap-3 flex-wrap mb-6">
      {badge(f'{len(trades)} signals', "info")}
    </div>'''

    tables = ""
    if not trades:
        tables = '''<div class="flex flex-col items-center justify-center py-16 text-center">
          <p class="text-lg font-medium text-gray-600">No signals yet</p>
          <p class="mt-1 text-sm text-gray-400">Run the engine to scan for signals</p>
        </div>'''
    else:
        for hz, hz_trades in by_hz.items():
            rows = ""
            for t in hz_trades:
                upside = ((t["target_price"] - t["entry_price"]) / t["entry_price"] * 100) if t["entry_price"] else 0
                dir_bdg = badge(t["direction"], "bullish")
                conf_v = "success" if t.get("confidence") == "HIGH" else "warning" if t.get("confidence") == "MEDIUM" else "danger"
                patterns = (t.get("patterns") or "").replace(",", ", ")
                rows += f'''
                <tr class="hover:bg-blue-50/50 transition border-b border-gray-100">
                  <td class="px-4 py-3 font-semibold text-gray-800">{_e(_ticker(t["ticker"]))}</td>
                  <td class="px-4 py-3">{dir_bdg}</td>
                  <td class="px-4 py-3 text-right font-mono text-gray-700">{_price(t["entry_price"])}</td>
                  <td class="px-4 py-3 text-right"><span class="font-mono text-emerald-600">{_price(t["target_price"])}</span> <span class="text-xs text-gray-400">({_pct(upside)})</span></td>
                  <td class="px-4 py-3 text-right font-mono text-red-600">{_price(t["sl_price"])}</td>
                  <td class="px-4 py-3 text-right font-semibold text-gray-800">{t["rr_ratio"]:.1f}x</td>
                  <td class="px-4 py-3 text-right font-semibold text-gray-800">{t["predicted_win_rate"]:.0f}%</td>
                  <td class="px-4 py-3 text-gray-600 text-xs max-w-[200px] truncate">{_e(patterns)}</td>
                  <td class="px-4 py-3 text-center">{badge(t.get("confidence",""), conf_v)}</td>
                </tr>'''

            tables += f'''
            <div class="glass rounded-xl overflow-hidden mb-4">
              <div class="px-5 py-3 bg-gray-50 border-b border-gray-200 flex items-center justify-between">
                <h3 class="text-sm font-semibold text-gray-800">{_e(hz)}</h3>
                {badge(f'{len(hz_trades)} trades', "default")}
              </div>
              <div class="overflow-x-auto scrollbar-thin">
                <table class="w-full text-sm">
                  <thead><tr class="border-b border-gray-200">
                    <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Stock</th>
                    <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Dir</th>
                    <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Entry</th>
                    <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Target</th>
                    <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Stop Loss</th>
                    <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">R:R</th>
                    <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Win Rate</th>
                    <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Pattern</th>
                    <th class="px-4 py-3 text-center text-xs font-medium text-gray-500 uppercase">Conf</th>
                  </tr></thead>
                  <tbody>{rows}</tbody>
                </table>
              </div>
            </div>'''

    # ---- Global Bearish Score widget ----
    try:
        from global_sentiment import get_overnight_bearish_score
        bs = get_overnight_bearish_score()
        if bs >= 70:
            bs_bg = "bg-red-50 border-red-200"; bs_icon = "\u26a0\ufe0f"
            bs_label = "RED ALERT — BTST trims active"; bs_tc = "text-red-700"
        elif bs >= 40:
            bs_bg = "bg-amber-50 border-amber-200"; bs_icon = "\u26a1"
            bs_label = "CAUTION — Elevated bearish risk"; bs_tc = "text-amber-700"
        else:
            bs_bg = "bg-green-50 border-green-200"; bs_icon = "\u2705"
            bs_label = "SAFE — Global markets neutral"; bs_tc = "text-green-700"
    except Exception:
        bs = 30; bs_bg = "bg-gray-50 border-gray-200"
        bs_icon = "\u2014"; bs_label = "Score unavailable"; bs_tc = "text-gray-500"

    bearish_widget = f'''<div class="mb-5 p-4 rounded-xl border {bs_bg} flex items-center justify-between shadow-sm">
      <div class="flex items-center gap-3">
        <span class="text-2xl">{bs_icon}</span>
        <div>
          <p class="text-xs font-medium text-gray-500 uppercase tracking-wide">Global Bearish Score</p>
          <p class="text-xl font-bold {bs_tc}">{bs} / 100 &nbsp;&mdash;&nbsp; {bs_label}</p>
          <p class="text-xs text-gray-400 mt-0.5">S&amp;P Futures &middot; VIX &middot; DXY &middot; Oil &middot; Nikkei &middot; Hang Seng &middot; ASX</p>
        </div>
      </div>
      <div class="text-right text-xs text-gray-400 space-y-1">
        <p>BTST auto-trim: score &gt; 70</p>
        <p>Intraday trim: delta &gt; 25 pts</p>
        <p>Early-exit: trajectory &le; 40</p>
        <p>SHORT_1d gate: score &ge; 70</p>
        <a href="/api/bearish-score" target="_blank" class="text-blue-500 hover:underline">Live JSON →</a>
        <a href="/api/pending-trims" target="_blank" class="text-blue-500 hover:underline block">Pending trims →</a>
        <a href="/api/early-exits" target="_blank" class="text-blue-500 hover:underline block">Dead trades →</a>
        <a href="/api/short-trades" target="_blank" class="text-blue-500 hover:underline block">Short trades →</a>
        <a href="/api/short-force-closes" target="_blank" class="text-blue-500 hover:underline block">Short close queue →</a>
      </div>
    </div>'''

    body = f'''{bearish_widget}
    <div class="flex items-center justify-between mb-4">
      <div>
        <h2 class="text-2xl font-bold text-gray-800">Today's Signals</h2>
        <p class="text-sm text-gray-500 mt-1">Signals from {_date(entry_date)} — {len(trades)} total</p>
      </div>
      <a href="/signals" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-white border border-gray-200 hover:bg-gray-50 text-gray-600 text-sm transition shadow-sm">Refresh</a>
    </div>
    {summary_badges}
    {tables}'''
    return page_shell("Today's Signals", "signals", body)


def render_positions():
    trades = q_open_trades()
    today_str = date.today().isoformat()

    # Fetch live prices for all open tickers
    tickers_raw = [t["ticker"] for t in trades if t.get("ticker")]
    live_prices = fetch_live_prices(tickers_raw)
    price_ts = datetime.now().strftime("%H:%M")

    cards = ""
    if not trades:
        cards = '''<div class="flex flex-col items-center justify-center py-16 text-center">
          <p class="text-lg font-medium text-gray-600">No open positions</p>
          <p class="mt-1 text-sm text-gray-400">Trades will appear here after running the engine</p>
        </div>'''
    else:
        for i, t in enumerate(trades):
            days_held = _days_between(t["entry_date"], today_str)
            days_left = max(0, _days_between(today_str, t["expiry_date"]))
            total_days = t.get("horizon_days") or max(1, _days_between(t["entry_date"], t["expiry_date"]))
            pct_done = min(100, int(days_held / total_days * 100)) if total_days > 0 else 0
            upside = ((t["target_price"] - t["entry_price"]) / t["entry_price"] * 100) if t["entry_price"] else 0
            downside = ((t["entry_price"] - t["sl_price"]) / t["entry_price"] * 100) if t["entry_price"] else 0
            dir_cls = "bg-emerald-50 text-emerald-600"
            dir_arrow = "↑"
            bar_color = "bg-amber-500" if days_left <= 1 else "bg-blue-500"
            conf_v = "success" if t.get("confidence") == "HIGH" else "warning" if t.get("confidence") == "MEDIUM" else "danger"
            patterns_display = (t.get("patterns") or "").replace(",", " · ")

            # ---- Live price & P&L computation ----
            cur_price = live_prices.get(t["ticker"])
            entry_p = t["entry_price"] or 0
            target_p = t["target_price"] or 0
            sl_p = t["sl_price"] or 0
            is_bull = t["direction"] == "BULLISH"

            if cur_price and entry_p:
                pnl_pct = (cur_price - entry_p) / entry_p * 100
                pnl_sign = "+" if pnl_pct >= 0 else ""
                pnl_color = "text-emerald-600" if pnl_pct >= 0 else "text-red-600"
                pnl_bg = "bg-emerald-50" if pnl_pct >= 0 else "bg-red-50"

                # Distance gauge: where is cur_price between SL and Target?
                if is_bull:
                    total_range = target_p - sl_p if target_p != sl_p else 1
                    position_in_range = (cur_price - sl_p) / total_range * 100
                    dist_to_target = ((target_p - cur_price) / cur_price * 100) if cur_price else 0
                    dist_to_sl = ((cur_price - sl_p) / cur_price * 100) if cur_price else 0
                else:
                    total_range = sl_p - target_p if sl_p != target_p else 1
                    position_in_range = (sl_p - cur_price) / total_range * 100
                    dist_to_target = ((cur_price - target_p) / cur_price * 100) if cur_price else 0
                    dist_to_sl = ((sl_p - cur_price) / cur_price * 100) if cur_price else 0
                position_in_range = max(0, min(100, position_in_range))

                # Gauge bar color based on position
                if position_in_range >= 70:
                    gauge_color = "bg-emerald-500"  # near target
                elif position_in_range >= 30:
                    gauge_color = "bg-amber-400"    # mid-range
                else:
                    gauge_color = "bg-red-500"      # near SL

                live_price_html = f'''
                <div class="rounded-lg {pnl_bg} border border-gray-100 p-3 mb-4">
                  <div class="flex items-center justify-between mb-2">
                    <div>
                      <p class="text-[10px] uppercase tracking-wider text-gray-400 mb-0.5">Current Price</p>
                      <p class="text-lg font-mono font-bold {pnl_color}">{_price(cur_price)}</p>
                    </div>
                    <div class="text-right">
                      <p class="text-[10px] uppercase tracking-wider text-gray-400 mb-0.5">Unrealized P&L</p>
                      <p class="text-lg font-bold {pnl_color}">{pnl_sign}{pnl_pct:.2f}%</p>
                    </div>
                  </div>
                  <div class="mb-1">
                    <div class="flex justify-between text-[10px] text-gray-400 mb-1">
                      <span>SL ({_price(sl_p)})</span>
                      <span>Target ({_price(target_p)})</span>
                    </div>
                    <div class="w-full h-2 bg-gray-200 rounded-full overflow-hidden relative">
                      <div class="h-full rounded-full {gauge_color} transition-all" style="width:{position_in_range:.0f}%"></div>
                    </div>
                  </div>
                  <div class="flex justify-between text-[10px] mt-1">
                    <span class="text-red-500">{dist_to_sl:.1f}% to SL</span>
                    <span class="text-emerald-500">{dist_to_target:.1f}% to Target</span>
                  </div>
                </div>'''
            else:
                live_price_html = '''
                <div class="rounded-lg bg-gray-50 border border-dashed border-gray-200 p-3 mb-4 text-center">
                  <p class="text-xs text-gray-400">Live price unavailable</p>
                </div>'''

            hz_label = t.get("horizon_label", "") or ""
            cards += f'''
            <div class="glass rounded-xl p-5 hover:border-blue-300 transition-all position-card relative" data-horizon="{_e(hz_label)}" data-expiry="{t.get('expiry_date','')}" data-id="{t['id']}">
              <label class="select-checkbox-wrap hidden absolute top-3 left-3 z-10 cursor-pointer" title="Select">
                <input type="checkbox" class="pos-checkbox w-5 h-5 rounded border-gray-400 text-red-600 focus:ring-red-500 cursor-pointer" data-id="{t['id']}">
              </label>
              <div class="flex items-center justify-between mb-4">
                <div class="flex items-center gap-3">
                  <div class="w-10 h-10 rounded-xl flex items-center justify-center text-sm font-bold {dir_cls}">{dir_arrow}</div>
                  <div>
                    <p class="font-bold text-gray-800 text-base">{_e(_ticker(t["ticker"]))}</p>
                    <p class="text-xs text-gray-400">{_e((t.get("sector") or "NSE").upper())} · Entered {_date(t["entry_date"])}</p>
                  </div>
                </div>
                {badge(hz_label, "info")}
              </div>
              {live_price_html}
              <div class="grid grid-cols-3 gap-3 mb-4">
                <div>
                  <p class="text-[10px] uppercase tracking-wider text-gray-400 mb-0.5">Entry</p>
                  <p class="text-sm font-mono font-semibold text-gray-800">{_price(t["entry_price"])}</p>
                </div>
                <div>
                  <p class="text-[10px] uppercase tracking-wider text-gray-400 mb-0.5">Target</p>
                  <p class="text-sm font-mono font-semibold text-emerald-600">{_price(t["target_price"])}</p>
                  <p class="text-[10px] text-emerald-500">+{upside:.1f}%</p>
                </div>
                <div>
                  <p class="text-[10px] uppercase tracking-wider text-gray-400 mb-0.5">Stop Loss</p>
                  <p class="text-sm font-mono font-semibold text-red-600">{_price(t["sl_price"])}</p>
                  <p class="text-[10px] text-red-500">-{downside:.1f}%</p>
                </div>
              </div>
              <div class="mb-3">
                <div class="flex justify-between text-xs text-gray-400 mb-1.5">
                  <span>Day {days_held} of {total_days}</span>
                  <span class="font-medium text-gray-500">Expires {_date(t["expiry_date"])}</span>
                  <span>{days_left}d left</span>
                </div>
                <div class="w-full h-1.5 bg-gray-100 rounded-full overflow-hidden">
                  <div class="h-full rounded-full {bar_color} transition-all" style="width:{pct_done}%"></div>
                </div>
              </div>
              <div class="flex items-center justify-between pt-3 border-t border-gray-100">
                <span class="text-xs text-gray-500">R:R {t.get("rr_ratio",0):.1f}x</span>
                <span class="text-xs text-gray-500">WR {t.get("predicted_win_rate",0):.0f}%</span>
                {badge(t.get("confidence",""), conf_v)}
              </div>
              <div class="mt-2">
                <p class="text-[10px] text-gray-400 truncate" title="{_e(t.get('patterns',''))}">{_e(patterns_display)}</p>
              </div>
              <form method="POST" action="/trade/cancel?id={t['id']}" onsubmit="return confirm('Cancel this trade and erase all RAG imprints? This cannot be undone.')" class="cancel-trade-form mt-3 pt-3 border-t border-gray-100">
                <button type="submit" class="w-full py-1.5 text-xs font-medium text-red-500 border border-red-200 rounded-lg hover:bg-red-50 transition-all">&#x2715; Cancel Trade &amp; Remove from RAG</button>
              </form>
            </div>'''

    price_note = f'<span class="text-xs text-gray-400 ml-2">Prices as of {price_ts}</span>' if live_prices else ''

    # Build horizon filter buttons from actual data
    horizon_counts = {}
    for t in trades:
        hz = t.get("horizon_label", "") or ""
        horizon_counts[hz] = horizon_counts.get(hz, 0) + 1

    # Sort horizons by horizon_days
    hz_order = sorted(horizon_counts.keys(), key=lambda h: next((t.get("horizon_days", 0) for t in trades if (t.get("horizon_label") or "") == h), 0))

    hz_buttons = ''
    for hz in hz_order:
        cnt = horizon_counts[hz]
        hz_buttons += f'<button onclick="filterPositions(this, \'{_e(hz)}\')" class="hz-filter-btn px-3 py-1.5 rounded-lg text-xs font-medium border border-gray-200 bg-white text-gray-600 hover:bg-blue-50 hover:border-blue-300 hover:text-blue-700 transition-all" data-hz="{_e(hz)}">{_e(hz)} <span class="text-gray-400 ml-1">({cnt})</span></button>'

    # Collect unique expiry dates with counts, sorted chronologically
    expiry_counts = {}
    for t in trades:
        ed = t.get("expiry_date", "")
        if ed:
            expiry_counts[ed] = expiry_counts.get(ed, 0) + 1
    expiry_order = sorted(expiry_counts.keys())
    expiry_buttons = ''
    for ed in expiry_order:
        cnt = expiry_counts[ed]
        # Color-code: today = red/urgent, tomorrow = amber, rest = default
        is_today = (ed == today_str)
        is_tomorrow = False
        try:
            from datetime import timedelta
            is_tomorrow = (ed == (date.today() + timedelta(days=1)).isoformat())
        except Exception:
            pass
        if is_today:
            exp_cls = "border-red-300 bg-red-50 text-red-600 hover:bg-red-100"
        elif is_tomorrow:
            exp_cls = "border-amber-300 bg-amber-50 text-amber-600 hover:bg-amber-100"
        else:
            exp_cls = "border-gray-200 bg-white text-gray-600 hover:bg-blue-50 hover:border-blue-300 hover:text-blue-700"
        expiry_buttons += f'<button onclick="filterByExpiry(this, \'{ed}\')" class="exp-filter-btn px-3 py-1.5 rounded-lg text-xs font-medium border {exp_cls} transition-all" data-exp="{ed}">{_date(ed)}{" ⚠️" if is_today else ""} <span class="opacity-60 ml-1">({cnt})</span></button>'

    filter_bar = f'''
    <div class="glass rounded-xl p-4 mb-5 space-y-2">
      <div class="flex flex-wrap items-center gap-2">
        <span class="text-xs font-semibold text-gray-500 uppercase tracking-wider mr-1 w-16">Horizon:</span>
        <button onclick="filterPositions(this, 'ALL')" class="hz-filter-btn active-filter px-3 py-1.5 rounded-lg text-xs font-medium border border-blue-300 bg-blue-50 text-blue-700 transition-all" data-hz="ALL">All <span class="text-blue-400 ml-1">({len(trades)})</span></button>
        {hz_buttons}
      </div>

      <div class="flex flex-wrap items-center gap-2">
        <span class="text-xs font-semibold text-gray-500 uppercase tracking-wider mr-1 w-16">Expiry:</span>
        <button onclick="filterByExpiry(this, 'ALL')" class="exp-filter-btn active-exp-filter px-3 py-1.5 rounded-lg text-xs font-medium border border-blue-300 bg-blue-50 text-blue-700 transition-all" data-exp="ALL">All</button>
        {expiry_buttons}
      </div>
      <p class="text-[10px] text-gray-400 mt-2" id="filter-summary">Showing all {len(trades)} positions</p>
    </div>''' if trades else ''

    filter_js = '''
    <script>
    (function() {
      var activeHz = 'ALL', activeExp = 'ALL';

      function applyFilters() {
        var cards = document.querySelectorAll('.position-card');
        var shown = 0;
        cards.forEach(function(card) {
          var hz = card.getAttribute('data-horizon');
          var exp = card.getAttribute('data-expiry');
          var hzMatch = (activeHz === 'ALL' || hz === activeHz);
          var expMatch = (activeExp === 'ALL' || exp === activeExp);
          if (hzMatch && expMatch) {
            card.style.display = '';
            shown++;
          } else {
            card.style.display = 'none';
          }
        });
        var summary = document.getElementById('filter-summary');
        if (summary) {
          var parts = [];
          if (activeHz !== 'ALL') parts.push(activeHz);
          if (activeExp !== 'ALL') parts.push('expiry ' + activeExp);
          var label = parts.length ? parts.join(' + ') : 'all';
          summary.textContent = 'Showing ' + shown + ' of ' + cards.length + ' positions' + (parts.length ? ' \u2014 ' + label : '');
        }
      }

      window.filterPositions = function(btn, hz) {
        activeHz = hz;
        document.querySelectorAll('.hz-filter-btn').forEach(function(b) {
          b.classList.remove('active-filter', 'bg-blue-50', 'border-blue-300', 'text-blue-700');
          b.classList.add('bg-white', 'text-gray-600', 'border-gray-200');
        });
        btn.classList.add('active-filter', 'bg-blue-50', 'border-blue-300', 'text-blue-700');
        btn.classList.remove('bg-white', 'text-gray-600', 'border-gray-200');
        applyFilters();
      };

      window.filterByExpiry = function(btn, exp) {
        activeExp = exp;
        document.querySelectorAll('.exp-filter-btn').forEach(function(b) {
          b.classList.remove('active-exp-filter', 'bg-blue-50', 'border-blue-300', 'text-blue-700');
          if (!b.classList.contains('bg-red-50') && !b.classList.contains('bg-amber-50')) {
            b.classList.add('bg-white', 'border-gray-200');
          }
        });
        btn.classList.add('active-exp-filter', 'bg-blue-50', 'border-blue-300', 'text-blue-700');
        btn.classList.remove('bg-white', 'border-gray-200', 'bg-red-50', 'border-red-300', 'bg-amber-50', 'border-amber-300');
        applyFilters();
      };
    })();
    </script>'''

    bulk_bar = '''
    <div id="bulk-cancel-bar" class="hidden fixed bottom-6 left-1/2 -translate-x-1/2 z-50 flex items-center gap-3 px-5 py-3 bg-white border border-gray-200 rounded-2xl shadow-2xl">
      <span id="bulk-count" class="text-sm font-medium text-gray-700">0 selected</span>
      <button onclick="selectAllVisible()" class="px-3 py-2 rounded-lg bg-gray-100 text-gray-700 text-sm hover:bg-gray-200 transition font-medium">Select All Visible</button>
      <button onclick="cancelSelected()" class="px-4 py-2 rounded-lg bg-red-500 text-white text-sm font-semibold hover:bg-red-600 transition">&#x2715; Cancel Selected</button>
      <button onclick="toggleSelectMode()" class="px-3 py-2 rounded-lg bg-gray-100 text-gray-500 text-sm hover:bg-gray-200 transition">Exit Select</button>
    </div>'''

    multiselect_js = '''
    <script>
    (function() {
      var selectMode = false;
      window.toggleSelectMode = function() {
        selectMode = !selectMode;
        var btn = document.getElementById('select-mode-btn');
        var checkboxWraps = document.querySelectorAll('.select-checkbox-wrap');
        var cancelForms = document.querySelectorAll('.cancel-trade-form');
        var bar = document.getElementById('bulk-cancel-bar');
        if (selectMode) {
          btn.innerHTML = '&#x2715; Exit Select';
          btn.classList.add('bg-blue-50', 'border-blue-300', 'text-blue-700');
          btn.classList.remove('text-gray-600');
          checkboxWraps.forEach(function(w) { w.classList.remove('hidden'); });
          cancelForms.forEach(function(f) { f.classList.add('hidden'); });
          bar.classList.remove('hidden');
        } else {
          btn.innerHTML = '&#x2611; Select';
          btn.classList.remove('bg-blue-50', 'border-blue-300', 'text-blue-700');
          btn.classList.add('text-gray-600');
          checkboxWraps.forEach(function(w) {
            w.classList.add('hidden');
            w.querySelector('input').checked = false;
          });
          cancelForms.forEach(function(f) { f.classList.remove('hidden'); });
          bar.classList.add('hidden');
          updateBulkCount();
        }
      };
      window.updateBulkCount = function() {
        var checked = document.querySelectorAll('.pos-checkbox:checked');
        document.getElementById('bulk-count').textContent = checked.length + ' selected';
      };
      window.selectAllVisible = function() {
        document.querySelectorAll('.position-card').forEach(function(card) {
          if (card.style.display !== 'none') {
            var cb = card.querySelector('.pos-checkbox');
            if (cb) cb.checked = true;
          }
        });
        updateBulkCount();
      };
      window.cancelSelected = function() {
        var checked = document.querySelectorAll('.pos-checkbox:checked');
        if (!checked.length) { alert('No trades selected.'); return; }
        if (!confirm('Cancel ' + checked.length + ' selected trade(s) and remove their RAG imprints? This cannot be undone.')) return;
        var ids = Array.from(checked).map(function(cb) { return cb.getAttribute('data-id'); }).join(',');
        var form = document.createElement('form');
        form.method = 'POST';
        form.action = '/trade/cancel-bulk';
        var inp = document.createElement('input');
        inp.type = 'hidden'; inp.name = 'ids'; inp.value = ids;
        form.appendChild(inp);
        document.body.appendChild(form);
        form.submit();
      };
      document.addEventListener('change', function(e) {
        if (e.target && e.target.classList.contains('pos-checkbox')) {
          updateBulkCount();
        }
      });
    })();
    </script>'''

    body = f'''
    <div class="flex items-center justify-between mb-6">
      <div>
        <h2 class="text-2xl font-bold text-gray-800">Open Positions</h2>
        <p class="text-sm text-gray-500 mt-1">{len(trades)} active trades {price_note}</p>
      </div>
      <div class="flex items-center gap-2">
        <button id="select-mode-btn" onclick="toggleSelectMode()" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-white border border-gray-200 hover:bg-blue-50 hover:border-blue-300 text-gray-600 text-sm transition shadow-sm">&#x2611; Select</button>
        <a href="/positions" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-white border border-gray-200 hover:bg-gray-50 text-gray-600 text-sm transition shadow-sm">
          <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 4v5h.582m15.356 2A8.001 8.001 0 004.582 9m0 0H9m11 11v-5h-.581m0 0a8.003 8.003 0 01-15.357-2m15.357 2H15"/></svg>
          Refresh Prices
        </a>
      </div>
    </div>
    {filter_bar}
    <div class="grid grid-cols-1 lg:grid-cols-2 xl:grid-cols-3 gap-4" id="positions-grid">
      {cards}
    </div>
    {filter_js}
    {bulk_bar}
    {multiselect_js}'''
    return page_shell("Open Positions", "positions", body)


def render_history():
    trades = q_closed_trades()

    wins = sum(1 for t in trades if t["status"] in ("WON", "EXPIRED_WIN"))
    losses = sum(1 for t in trades if t["status"] in ("LOST", "EXPIRED_LOSS"))
    wr = (wins / len(trades) * 100) if trades else 0
    wr_color = "text-emerald-600" if wr >= 55 else "text-amber-600" if wr >= 45 else "text-red-600"

    summary = f'''
    <div class="grid grid-cols-4 gap-4 mb-6">
      <div class="rounded-xl bg-white border border-gray-200 shadow-sm p-5">
        <p class="text-xs font-medium uppercase tracking-wider text-blue-600">Total</p>
        <p class="mt-2 text-2xl font-bold text-gray-800" id="hist-total">{len(trades)}</p>
      </div>
      <div class="rounded-xl bg-white border border-emerald-200 shadow-sm p-5">
        <p class="text-xs font-medium uppercase tracking-wider text-emerald-600">Wins</p>
        <p class="mt-2 text-2xl font-bold text-gray-800" id="hist-wins">{wins}</p>
      </div>
      <div class="rounded-xl bg-white border border-red-200 shadow-sm p-5">
        <p class="text-xs font-medium uppercase tracking-wider text-red-600">Losses</p>
        <p class="mt-2 text-2xl font-bold text-gray-800" id="hist-losses">{losses}</p>
      </div>
      <div class="rounded-xl bg-white border border-amber-200 shadow-sm p-5">
        <p class="text-xs font-medium uppercase tracking-wider text-amber-600">Win Rate</p>
        <p class="mt-2 text-2xl font-bold {wr_color}" id="hist-wr">{wr:.1f}%</p>
      </div>
    </div>'''

    if not trades:
        table = '''<div class="flex flex-col items-center justify-center py-16 text-center">
          <p class="text-lg font-medium text-gray-600">No closed trades yet</p>
          <p class="mt-1 text-sm text-gray-400">Trades will appear here once they hit SL, target, or expire</p>
        </div>'''
    else:
        rows = ""
        for t in trades:
            ret = t.get("actual_return_pct", 0) or 0
            ret_cls = "text-emerald-600" if ret >= 0 else "text-red-600"

        # Build dropdown option lists from actual data
        horizons = sorted(set(t.get("horizon_label", "") for t in trades if t.get("horizon_label")))
        statuses_raw = {"WON": "Won", "LOST": "Lost", "EXPIRED_WIN": "Exp Win", "EXPIRED_LOSS": "Exp Loss", "CANCELLED": "Cancelled"}
        statuses_present = sorted(set(t.get("status", "") for t in trades if t.get("status")))
        reasons_present = sorted(set(t.get("exit_reason", "") for t in trades if t.get("exit_reason")))

        sel_cls = "w-full text-xs border border-gray-200 rounded-md px-1 py-1 bg-white text-gray-700 focus:outline-none focus:ring-1 focus:ring-blue-400"

        hz_opts = "<option value=''>All</option>" + "".join(f"<option value='{hz}'>{hz}</option>" for hz in horizons)
        st_opts = "<option value=''>All</option>" + "".join(f"<option value='{s}'>{statuses_raw.get(s,s)}</option>" for s in statuses_present)
        re_opts = "<option value=''>All</option>" + "".join(f"<option value='{r}'>{r}</option>" for r in reasons_present)

        rows = ""
        for t in trades:
            ret = t.get("actual_return_pct", 0) or 0
            ret_cls = "text-emerald-600" if ret >= 0 else "text-red-600"

            # Win% — show whenever predicted_win_rate is available
            win_pct = t.get("predicted_win_rate")
            if win_pct is not None:
                win_pct_str = f"{win_pct:.0f}%"
                win_pct_color = "text-emerald-600" if win_pct >= 50 else "text-amber-600"
            else:
                win_pct_str = "—"
                win_pct_color = "text-gray-400"

            trade_id = t.get("id", 0)
            row_horizon = _e(t.get("horizon_label", ""))
            row_status = _e(t.get("status", ""))
            row_ticker = _e(_ticker(t["ticker"]))
            row_reason = _e(t.get("exit_reason", ""))
            row_exit_date = t.get("exit_date", "") or ""
            rows += f'''
            <tr class="hover:bg-blue-50/50 transition border-b border-gray-100 history-row" data-trade-id="{trade_id}" data-horizon="{row_horizon}" data-status="{row_status}" data-ticker="{row_ticker}" data-reason="{row_reason}" data-exit-date="{row_exit_date}">
              <td class="px-3 py-3">
                <input type="checkbox" class="trade-checkbox w-4 h-4 rounded border-gray-300 text-blue-600 cursor-pointer" value="{trade_id}" data-ticker="{row_ticker}">
              </td>
              <td class="px-4 py-3">{status_badge(t["status"])}</td>
              <td class="px-4 py-3 font-semibold text-gray-800">{row_ticker}</td>
              <td class="px-4 py-3 text-gray-600">{row_horizon}</td>
              <td class="px-4 py-3 text-center font-semibold {win_pct_color}">{win_pct_str}</td>
              <td class="px-4 py-3 text-right font-mono text-gray-600">{_price(t["entry_price"])}</td>
              <td class="px-4 py-3 text-right font-mono text-gray-600">{_price(t.get("exit_price"))}</td>
              <td class="px-4 py-3 text-right font-mono font-semibold {ret_cls}">{_pct(ret)}</td>
              <td class="px-4 py-3 text-xs text-gray-500">{row_reason}</td>
              <td class="px-4 py-3 text-xs text-gray-500">{_date(t["entry_date"])}</td>
              <td class="px-4 py-3 text-xs text-gray-500">{_date(t.get("exit_date"))}</td>
              <td class="px-4 py-3 text-xs text-gray-500 max-w-[150px] truncate">{_e(t.get("patterns",""))}</td>
            </tr>'''

        table = f'''
        <div class="glass rounded-xl overflow-hidden">
          <div class="px-6 py-4 border-b border-gray-200 flex items-center justify-between bg-gray-50">
            <div class="flex items-center gap-3">
              <input type="checkbox" id="select-all-trades" class="w-4 h-4 rounded border-gray-300 text-blue-600 cursor-pointer">
              <label for="select-all-trades" class="text-sm font-medium text-gray-700 cursor-pointer">Select All</label>
              <span id="selection-count" class="text-sm text-gray-500">(0 / {len(trades)} selected)</span>
            </div>
            <button id="delete-selected-btn" onclick="deleteSelectedTrades()" disabled class="px-4 py-2 rounded-lg bg-red-600 text-white text-sm font-medium hover:bg-red-700 disabled:opacity-50 disabled:cursor-not-allowed transition">
              Delete Selected
            </button>
          </div>
          <div class="overflow-x-auto scrollbar-thin">
            <table class="w-full text-sm">
              <thead>
                <tr class="border-b border-gray-200 bg-gray-50">
                  <th class="px-3 py-3 w-8"></th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Status</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Stock</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Horizon</th>
                  <th class="px-4 py-3 text-center text-xs font-medium text-gray-500 uppercase">Win%</th>
                  <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Entry</th>
                  <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Exit</th>
                  <th class="px-4 py-3 text-right text-xs font-medium text-gray-500 uppercase">Return</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Reason</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Entry</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Exit</th>
                  <th class="px-4 py-3 text-left text-xs font-medium text-gray-500 uppercase">Pattern</th>
                </tr>
                <tr class="border-b border-gray-100 bg-gray-50">
                  <th class="px-3 py-2"></th>
                  <th class="px-2 py-2"><select id="f-status" onchange="applyHistoryFilters()" class="{sel_cls}">{st_opts}</select></th>
                  <th class="px-2 py-2"><input id="f-stock" oninput="applyHistoryFilters()" type="text" placeholder="Search…" class="{sel_cls}"></th>
                  <th class="px-2 py-2"><select id="f-horizon" onchange="applyHistoryFilters()" class="{sel_cls}">{hz_opts}</select></th>
                  <th></th><th></th><th></th><th></th>
                  <th class="px-2 py-2"><select id="f-reason" onchange="applyHistoryFilters()" class="{sel_cls}">{re_opts}</select></th>
                  <th></th><th></th><th></th>
                </tr>
              </thead>
              <tbody>{rows}</tbody>
            </table>
          </div>
        </div>

        <script>
        function applyHistoryFilters() {{
          var fStatus  = document.getElementById('f-status').value.toLowerCase();
          var fStock   = document.getElementById('f-stock').value.toLowerCase().trim();
          var fHorizon = document.getElementById('f-horizon').value.toLowerCase();
          var fReason  = document.getElementById('f-reason').value.toLowerCase();
          var total = 0, wins = 0, losses = 0;
          document.querySelectorAll('.history-row').forEach(function(row) {{
            var s  = (row.dataset.status  || '').toLowerCase();
            var tk = (row.dataset.ticker  || '').toLowerCase();
            var hz = (row.dataset.horizon || '').toLowerCase();
            var re = (row.dataset.reason  || '').toLowerCase();
            var show = (!fStatus  || s  === fStatus)
                    && (!fStock   || tk.includes(fStock))
                    && (!fHorizon || hz === fHorizon)
                    && (!fReason  || re === fReason);
            row.style.display = show ? '' : 'none';
            if (show) {{
              total++;
              if (s === 'won' || s === 'expired_win') wins++;
              else if (s === 'lost' || s === 'expired_loss') losses++;
            }}
          }});
          document.getElementById('hist-total').textContent   = total;
          document.getElementById('hist-wins').textContent    = wins;
          document.getElementById('hist-losses').textContent  = losses;
          var wr = total > 0 ? (wins / total * 100).toFixed(1) : '0.0';
          document.getElementById('hist-wr').textContent = wr + '%';
        }}

        function updateSelectionCount() {{
          const checkboxes = document.querySelectorAll('.trade-checkbox');
          const selected = Array.from(checkboxes).filter(cb => cb.checked);
          const count = selected.length;
          document.getElementById('selection-count').textContent = `(${{count}} / {len(trades)} selected)`;
          document.getElementById('delete-selected-btn').disabled = count === 0;

          // Update "Select All" checkbox state
          const selectAllCheckbox = document.getElementById('select-all-trades');
          if (count === 0) {{
            selectAllCheckbox.checked = false;
            selectAllCheckbox.indeterminate = false;
          }} else if (count === checkboxes.length) {{
            selectAllCheckbox.checked = true;
            selectAllCheckbox.indeterminate = false;
          }} else {{
            selectAllCheckbox.indeterminate = true;
          }}
        }}

        document.querySelectorAll('.trade-checkbox').forEach(checkbox => {{
          checkbox.addEventListener('change', function(e) {{
            updateSelectionCount();
          }});
        }});

        document.getElementById('select-all-trades').addEventListener('change', function(e) {{
          const checkboxes = document.querySelectorAll('.trade-checkbox');
          if (this.checked) {{
            checkboxes.forEach(cb => cb.checked = true);
          }} else {{
            checkboxes.forEach(cb => cb.checked = false);
          }}
          updateSelectionCount();
        }});

        function deleteSelectedTrades() {{
          const selected = Array.from(document.querySelectorAll('.trade-checkbox:checked'));
          if (selected.length === 0) {{
            alert('No trades selected');
            return;
          }}
          
          const tickers = selected.map(cb => cb.dataset.ticker).join(', ');
          const confirmMsg = `Are you sure you want to permanently delete ${{selected.length}} trade(s)?\\n\\nTrades: ${{tickers}}\\n\\nThis action:\\n- Deletes trades from database\\n- Removes from RAG memory\\n- Clears all feedback entries\\n- Excludes from all calculations\\n\\nThis CANNOT be undone.`;
          
          if (!confirm(confirmMsg)) {{
            return;
          }}

          const ids = selected.map(cb => cb.value).join(',');
          const btn = document.getElementById('delete-selected-btn');
          btn.disabled = true;
          btn.textContent = 'Deleting...';

          // Use a hidden form to do a real browser POST + redirect
          // This avoids fetch() swallowing the server redirect
          const form = document.createElement('form');
          form.method = 'POST';
          form.action = '/trade/purge';
          const input = document.createElement('input');
          input.type = 'hidden';
          input.name = 'ids';
          input.value = ids;
          form.appendChild(input);
          document.body.appendChild(form);
          form.submit();
        }}

        </script>'''

    body = f'''
    <div class="flex items-center justify-between mb-6">
      <div>
        <h2 class="text-2xl font-bold text-gray-800">Trade History</h2>
        <p class="text-sm text-gray-500 mt-1">{len(trades)} closed trades</p>
      </div>
      <div class="flex items-center gap-2">
        <a href="/history/export" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-emerald-600 hover:bg-emerald-700 text-white text-sm font-medium transition shadow-sm">
          <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>
          Download Excel
        </a>
        <a href="/history" class="flex items-center gap-2 px-3 py-2 rounded-lg bg-white border border-gray-200 hover:bg-gray-50 text-gray-600 text-sm transition shadow-sm">Refresh</a>
      </div>
    </div>
    {summary}
    {table}'''
    return page_shell("Trade History", "history", body)


def render_performance():
    s = q_stats()
    hz_stats = q_stats_by_horizon()
    pat_stats = q_stats_by_pattern()
    stock_stats = q_stats_by_stock()
    sec_stats = q_stats_by_sector()

    kpis = f'''
    <div class="grid grid-cols-2 md:grid-cols-5 gap-4 mb-6">
      {stat_card("Total Trades", s["total_trades"], "", "indigo")}
      {stat_card("Win Rate", f'{s["win_rate"]}%', "", "green" if s["win_rate"] >= 55 else "amber")}
      {stat_card("Profit Factor", s["profit_factor"], "", "green" if s["profit_factor"] >= 1.5 else "amber")}
      {stat_card("Avg Win", _pct(s["avg_win_pct"]), "", "green")}
      {stat_card("Avg Loss", _pct(s["avg_loss_pct"]), "", "red")}
    </div>'''

    # Horizon table
    hz_html = ""
    if hz_stats:
        hz_rows = ""
        for h in hz_stats:
            hz_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 font-medium text-gray-800">{_e(h.get("horizon_label",""))}</td>
              <td class="px-4 py-2 text-right text-gray-600" data-val="{h["total"]}">{h["total"]}</td>
              <td class="px-4 py-2 text-right text-emerald-600" data-val="{h["wins"]}">{h["wins"]}</td>
              <td class="px-4 py-2 text-right text-red-600" data-val="{h["losses"]}">{h["losses"]}</td>
              <td class="px-4 py-2 text-right font-semibold text-gray-800" data-val="{h["win_rate"]}">{h["win_rate"]}%</td>
              <td class="px-4 py-2 text-right text-emerald-600" data-val="{round(h.get('avg_win') or 0, 4)}">{_pct(h.get("avg_win"))}</td>
              <td class="px-4 py-2 text-right text-red-600" data-val="{round(h.get('avg_loss') or 0, 4)}">{_pct(h.get("avg_loss"))}</td>
            </tr>'''
        hz_html = f'''
        <div class="glass rounded-xl p-6 mb-6">
          <div class="flex items-center justify-between mb-4">
            <h3 class="text-lg font-semibold text-gray-800">Performance by Horizon</h3>
            <button onclick="hzDownload()" title="Download as Excel/CSV" class="inline-flex items-center gap-1.5 px-3 py-1.5 text-xs font-medium text-white bg-emerald-600 hover:bg-emerald-700 active:bg-emerald-800 rounded-lg transition-colors shadow-sm">
              <svg class="w-3.5 h-3.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>
              Download Excel
            </button>
          </div>
          <table id="hz-table" class="w-full text-sm"><thead><tr class="border-b border-gray-200">
            <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Horizon</th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="1" onclick="hzSort(this)"><span>Trades</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="2" onclick="hzSort(this)"><span>Wins</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="3" onclick="hzSort(this)"><span>Losses</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="4" onclick="hzSort(this)"><span>Win Rate</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="5" onclick="hzSort(this)"><span>Avg Win</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="6" onclick="hzSort(this)"><span>Avg Loss</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
          </tr></thead><tbody>{hz_rows}</tbody></table>
        </div>
        <script>
        window.hzDownload = function() {{
          var headers = ['Horizon','Trades','Wins','Losses','Win Rate %','Avg Win %','Avg Loss %'];
          var rows = [headers];
          document.querySelectorAll('#hz-table tbody tr').forEach(function(tr) {{
            var cells = tr.querySelectorAll('td');
            rows.push([
              cells[0].textContent.trim(),
              cells[1].textContent.trim(),
              cells[2].textContent.trim(),
              cells[3].textContent.trim(),
              cells[4].getAttribute('data-val') || '',
              cells[5].getAttribute('data-val') || '',
              cells[6].getAttribute('data-val') || ''
            ]);
          }});
          var csv = rows.map(function(r) {{
            return r.map(function(v) {{
              var s = String(v);
              return (s.indexOf(',')>=0||s.indexOf('"')>=0||s.indexOf('\\n')>=0)
                ? '"'+s.replace(/"/g,'""')+'"' : s;
            }}).join(',');
          }}).join('\\r\\n');
          var blob = new Blob(['\\uFEFF'+csv], {{type:'text/csv;charset=utf-8;'}});
          var url = URL.createObjectURL(blob);
          var a = document.createElement('a');
          a.href = url;
          a.download = 'horizon_performance_'+new Date().toISOString().slice(0,10)+'.csv';
          document.body.appendChild(a); a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);
        }};
        (function(){{
          var _hzCol = -1, _hzDir = 1;
          window.hzSort = function(th) {{
            var col = parseInt(th.getAttribute('data-col'));
            if (_hzCol === col) {{ _hzDir *= -1; }} else {{ _hzCol = col; _hzDir = -1; }}
            var icon = _hzDir === -1 ? '&#8595;' : '&#8593;';
            document.querySelectorAll('#hz-table thead th').forEach(function(h) {{
              var si = h.querySelector('.sort-icon');
              if (si) si.innerHTML = h === th ? icon : '&#8645;';
            }});
            var tbody = document.querySelector('#hz-table tbody');
            var rows = Array.from(tbody.querySelectorAll('tr'));
            rows.sort(function(a, b) {{
              var av = parseFloat(a.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              var bv = parseFloat(b.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              return (av - bv) * _hzDir;
            }});
            rows.forEach(function(r) {{ tbody.appendChild(r); }});
          }};
        }})();
        </script>'''

    # Pattern table
    pat_html = ""
    if pat_stats:
        pat_rows = ""
        for p in pat_stats:
            ret_cls = "text-emerald-600" if (p.get("avg_ret") or 0) >= 0 else "text-red-600"
            pat_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-gray-800 text-xs">{_e((p.get("patterns","") or "").replace(",", " · "))}</td>
              <td class="px-4 py-2 text-right text-gray-600" data-val="{p["total"]}">{p["total"]}</td>
              <td class="px-4 py-2 text-right text-gray-600" data-val="{p["wins"]}">{p["wins"]} / {p["losses"]}</td>
              <td class="px-4 py-2 text-right font-semibold text-gray-800" data-val="{p["win_rate"]}">{p["win_rate"]}%</td>
              <td class="px-4 py-2 text-right font-mono {ret_cls}" data-val="{round(p.get('avg_ret') or 0, 4)}">{_pct(p.get("avg_ret"))}</td>
            </tr>'''
        pat_html = f'''
        <div class="glass rounded-xl p-6 mb-6">
          <div class="flex items-center justify-between mb-4">
            <h3 class="text-lg font-semibold text-gray-800">Performance by Pattern</h3>
            <button onclick="patDownload()" title="Download as Excel/CSV" class="inline-flex items-center gap-1.5 px-3 py-1.5 text-xs font-medium text-white bg-emerald-600 hover:bg-emerald-700 active:bg-emerald-800 rounded-lg transition-colors shadow-sm">
              <svg class="w-3.5 h-3.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>
              Download Excel
            </button>
          </div>
          <table id="pat-table" class="w-full text-sm"><thead><tr class="border-b border-gray-200">
            <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Pattern</th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="1" onclick="patSort(this)"><span>Trades</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="2" onclick="patSort(this)"><span>W / L</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="3" onclick="patSort(this)"><span>Win Rate</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="4" onclick="patSort(this)"><span>Avg Return</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
          </tr></thead><tbody>{pat_rows}</tbody></table>
        </div>
        <script>
        window.patDownload = function() {{
          var headers = ['Pattern','Trades','W','L','Win Rate %','Avg Return %'];
          var rows = [headers];
          document.querySelectorAll('#pat-table tbody tr').forEach(function(tr) {{
            var cells = tr.querySelectorAll('td');
            var wl = cells[2].textContent.trim().split(' / ');
            rows.push([
              cells[0].textContent.trim(),
              cells[1].textContent.trim(),
              wl[0] || '',
              wl[1] || '',
              cells[3].getAttribute('data-val') || '',
              cells[4].getAttribute('data-val') || ''
            ]);
          }});
          var csv = rows.map(function(r) {{
            return r.map(function(v) {{
              var s = String(v);
              return (s.indexOf(',')>=0||s.indexOf('"')>=0||s.indexOf('\\n')>=0)
                ? '"'+s.replace(/"/g,'""')+'"' : s;
            }}).join(',');
          }}).join('\\r\\n');
          var blob = new Blob(['\\uFEFF'+csv], {{type:'text/csv;charset=utf-8;'}});
          var url = URL.createObjectURL(blob);
          var a = document.createElement('a');
          a.href = url;
          a.download = 'pattern_performance_'+new Date().toISOString().slice(0,10)+'.csv';
          document.body.appendChild(a); a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);
        }};
        (function(){{
          var _patCol = -1, _patDir = 1;
          window.patSort = function(th) {{
            var col = parseInt(th.getAttribute('data-col'));
            if (_patCol === col) {{ _patDir *= -1; }} else {{ _patCol = col; _patDir = -1; }}
            var icon = _patDir === -1 ? '&#8595;' : '&#8593;';
            document.querySelectorAll('#pat-table thead th').forEach(function(h) {{
              var si = h.querySelector('.sort-icon');
              if (si) si.innerHTML = h === th ? icon : '&#8645;';
            }});
            var tbody = document.querySelector('#pat-table tbody');
            var rows = Array.from(tbody.querySelectorAll('tr'));
            rows.sort(function(a, b) {{
              var av = parseFloat(a.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              var bv = parseFloat(b.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              return (av - bv) * _patDir;
            }});
            rows.forEach(function(r) {{ tbody.appendChild(r); }});
          }};
        }})();
        </script>'''

    # Sector table
    sec_html = ""
    if sec_stats:
        sec_rows = ""
        for sc in sec_stats:
            wr_cls = "text-emerald-600" if (sc.get("win_rate") or 0) >= 55 else "text-amber-600" if (sc.get("win_rate") or 0) >= 45 else "text-red-600"
            avg_ret_cls = "text-emerald-600" if (sc.get("avg_ret") or 0) >= 0 else "text-red-600"
            tot_ret_cls = "text-emerald-600" if (sc.get("total_ret") or 0) >= 0 else "text-red-600"
            pf_val = sc.get("profit_factor")
            pf_display = "∞" if pf_val == "∞" else (f"{pf_val:.2f}" if pf_val is not None else "—")
            pf_data = "9999" if pf_val == "∞" else (str(round(pf_val, 4)) if pf_val is not None else "0")
            sec_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-3 py-2 font-semibold text-gray-800 text-xs">{_e((sc.get("sector") or "").upper())}</td>
              <td class="px-3 py-2 text-right text-gray-600" data-val="{sc["stocks"]}">{sc["stocks"]}</td>
              <td class="px-3 py-2 text-right text-gray-600" data-val="{sc["total"]}">{sc["total"]}</td>
              <td class="px-3 py-2 text-right text-gray-500">{sc["wins"]} / {sc["losses"]}</td>
              <td class="px-3 py-2 text-right font-semibold {wr_cls}" data-val="{sc["win_rate"]}">{sc["win_rate"]}%</td>
              <td class="px-3 py-2 text-right text-emerald-600" data-val="{round(sc.get('avg_win') or 0, 4)}">{_pct(sc.get("avg_win"))}</td>
              <td class="px-3 py-2 text-right text-red-600" data-val="{round(sc.get('avg_loss') or 0, 4)}">{_pct(sc.get("avg_loss"))}</td>
              <td class="px-3 py-2 text-right font-mono {avg_ret_cls}" data-val="{round(sc.get('avg_ret') or 0, 4)}">{_pct(sc.get("avg_ret"))}</td>
              <td class="px-3 py-2 text-right font-mono font-semibold {tot_ret_cls}" data-val="{round(sc.get('total_ret') or 0, 4)}">{_pct(sc.get("total_ret"))}</td>
              <td class="px-3 py-2 text-right text-gray-700" data-val="{pf_data}">{pf_display}</td>
              <td class="px-3 py-2 text-right text-emerald-600" data-val="{round(sc.get('best_trade') or 0, 4)}">{_pct(sc.get("best_trade"))}</td>
              <td class="px-3 py-2 text-right text-red-600" data-val="{round(sc.get('worst_trade') or 0, 4)}">{_pct(sc.get("worst_trade"))}</td>
              <td class="px-3 py-2 text-xs text-gray-500">{_e(sc.get("dom_pattern") or "—")}</td>
              <td class="px-3 py-2 text-xs text-gray-500">{_e(sc.get("dom_horizon") or "—")}</td>
            </tr>'''
        sec_html = f'''
        <div class="glass rounded-xl p-6 mb-6">
          <div class="flex items-center justify-between mb-4">
            <h3 class="text-lg font-semibold text-gray-800">Performance by Sector</h3>
            <button onclick="secDownload()" title="Download as Excel/CSV" class="inline-flex items-center gap-1.5 px-3 py-1.5 text-xs font-medium text-white bg-emerald-600 hover:bg-emerald-700 active:bg-emerald-800 rounded-lg transition-colors shadow-sm">
              <svg class="w-3.5 h-3.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>
              Download Excel
            </button>
          </div>
          <div class="overflow-x-auto">
          <table id="sec-table" class="w-full text-sm min-w-max"><thead><tr class="border-b border-gray-200">
            <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Sector</th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="1" onclick="secSort(this)"><span>Stocks</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="2" onclick="secSort(this)"><span>Trades</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">W / L</th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="4" onclick="secSort(this)"><span>Win Rate</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="5" onclick="secSort(this)"><span>Avg Win</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="6" onclick="secSort(this)"><span>Avg Loss</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="7" onclick="secSort(this)"><span>Avg Ret</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="8" onclick="secSort(this)"><span>Total Ret</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="9" onclick="secSort(this)"><span>Prof. Factor</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="10" onclick="secSort(this)"><span>Best</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="11" onclick="secSort(this)"><span>Worst</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span></th>
            <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Dom. Pattern</th>
            <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Dom. Horizon</th>
          </tr></thead><tbody>{sec_rows}</tbody></table>
          </div>
        </div>
        <script>
        window.secDownload = function() {{
          var headers = ['Sector','Stocks','Trades','W','L','Win Rate %','Avg Win %','Avg Loss %','Avg Return %','Total Return %','Profit Factor','Best Trade %','Worst Trade %','Dom. Pattern','Dom. Horizon'];
          var rows = [headers];
          document.querySelectorAll('#sec-table tbody tr').forEach(function(tr) {{
            var cells = tr.querySelectorAll('td');
            var wl = cells[3].textContent.trim().split(' / ');
            rows.push([
              cells[0].textContent.trim(),
              cells[1].getAttribute('data-val') || '',
              cells[2].getAttribute('data-val') || '',
              wl[0] || '',
              wl[1] || '',
              cells[4].getAttribute('data-val') || '',
              cells[5].getAttribute('data-val') || '',
              cells[6].getAttribute('data-val') || '',
              cells[7].getAttribute('data-val') || '',
              cells[8].getAttribute('data-val') || '',
              cells[9].textContent.trim(),
              cells[10].getAttribute('data-val') || '',
              cells[11].getAttribute('data-val') || '',
              cells[12].textContent.trim(),
              cells[13].textContent.trim()
            ]);
          }});
          var csv = rows.map(function(r) {{
            return r.map(function(v) {{
              var s = String(v);
              return (s.indexOf(',')>=0||s.indexOf('"')>=0||s.indexOf('\\n')>=0)
                ? '"'+s.replace(/"/g,'""')+'"' : s;
            }}).join(',');
          }}).join('\\r\\n');
          var blob = new Blob(['\\uFEFF'+csv], {{type:'text/csv;charset=utf-8;'}});
          var url = URL.createObjectURL(blob);
          var a = document.createElement('a');
          a.href = url;
          a.download = 'sector_performance_'+new Date().toISOString().slice(0,10)+'.csv';
          document.body.appendChild(a); a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);
        }};
        (function(){{
          var _secCol = -1, _secDir = 1;
          window.secSort = function(th) {{
            var col = parseInt(th.getAttribute('data-col'));
            if (_secCol === col) {{ _secDir *= -1; }} else {{ _secCol = col; _secDir = -1; }}
            var icon = _secDir === -1 ? '&#8595;' : '&#8593;';
            document.querySelectorAll('#sec-table thead th').forEach(function(h) {{
              var si = h.querySelector('.sort-icon');
              if (si) si.innerHTML = h === th ? icon : '&#8645;';
            }});
            var tbody = document.querySelector('#sec-table tbody');
            var rows = Array.from(tbody.querySelectorAll('tr'));
            rows.sort(function(a, b) {{
              var av = parseFloat(a.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              var bv = parseFloat(b.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              return (av - bv) * _secDir;
            }});
            rows.forEach(function(r) {{ tbody.appendChild(r); }});
          }};
        }})();
        </script>'''

    # Stock table
    stk_html = ""
    if stock_stats:
        stk_rows = ""
        for st in stock_stats:
            avg_cls = "text-emerald-600" if (st.get("avg_ret") or 0) >= 0 else "text-red-600"
            tot_cls = "text-emerald-600" if (st.get("total_ret") or 0) >= 0 else "text-red-600"
            avg_val = round(st.get("avg_ret") or 0, 4)
            tot_val = round(st.get("total_ret") or 0, 4)
            stk_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 font-semibold text-gray-800">{_e(_ticker(st["ticker"]))}</td>
              <td class="px-4 py-2 text-xs text-gray-500 uppercase">{_e((st.get("sector") or "—").upper())}</td>
              <td class="px-4 py-2 text-right text-gray-600" data-val="{st["total"]}">{st["total"]}</td>
              <td class="px-4 py-2 text-right text-gray-600">{st["wins"]} / {st["losses"]}</td>
              <td class="px-4 py-2 text-right font-semibold text-gray-800" data-val="{st["win_rate"]}">{st["win_rate"]}%</td>
              <td class="px-4 py-2 text-right font-mono {avg_cls}" data-val="{avg_val}">{_pct(st.get("avg_ret"))}</td>
              <td class="px-4 py-2 text-right font-mono font-semibold {tot_cls}" data-val="{tot_val}">{_pct(st.get("total_ret"))}</td>
            </tr>'''
        stk_html = f'''
        <div class="glass rounded-xl p-6 mb-6">
          <div class="flex items-center justify-between mb-4">
            <h3 class="text-lg font-semibold text-gray-800">Performance by Stock</h3>
            <div class="flex items-center gap-3">
              <span class="text-xs text-gray-400">{len(stock_stats)} stocks</span>
              <button onclick="stkDownload()" title="Download as Excel/CSV" class="inline-flex items-center gap-1.5 px-3 py-1.5 text-xs font-medium text-white bg-emerald-600 hover:bg-emerald-700 active:bg-emerald-800 rounded-lg transition-colors shadow-sm">
                <svg class="w-3.5 h-3.5" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>
                Download Excel
              </button>
            </div>
          </div>
          <table id="stk-table" class="w-full text-sm">
            <thead><tr class="border-b border-gray-200">
              <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Stock</th>
              <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Sector</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="2" onclick="stkSort(this)">
                <span>Trades</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span>
              </th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">W / L</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="4" onclick="stkSort(this)">
                <span>Win Rate</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span>
              </th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="5" onclick="stkSort(this)">
                <span>Avg Return</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span>
              </th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase cursor-pointer select-none group" data-col="6" onclick="stkSort(this)">
                <span>Total Return</span> <span class="sort-icon text-gray-300 group-hover:text-blue-400">&#8645;</span>
              </th>
            </tr></thead>
            <tbody>{stk_rows}</tbody>
          </table>
        </div>
        <script>
        (function(){{
          var _stkCol = -1, _stkDir = 1;
          window.stkDownload = function() {{
            var headers = ['Stock','Sector','Trades','W','L','Win Rate %','Avg Return %','Total Return %'];
            var rows = [headers];
            document.querySelectorAll('#stk-table tbody tr').forEach(function(tr) {{
              var cells = tr.querySelectorAll('td');
              var wl = cells[3].textContent.trim().split(' / ');
              rows.push([
                cells[0].textContent.trim(),
                cells[1].textContent.trim(),
                cells[2].textContent.trim(),
                wl[0] || '',
                wl[1] || '',
                cells[4].getAttribute('data-val') || '',
                cells[5].getAttribute('data-val') || '',
                cells[6].getAttribute('data-val') || ''
              ]);
            }});
            var csv = rows.map(function(r) {{
              return r.map(function(v) {{
                var s = String(v);
                return (s.indexOf(',') >= 0 || s.indexOf('"') >= 0 || s.indexOf('\\n') >= 0)
                  ? '"' + s.replace(/"/g, '""') + '"' : s;
              }}).join(',');
            }}).join('\\r\\n');
            var blob = new Blob(['\\uFEFF' + csv], {{type: 'text/csv;charset=utf-8;'}});
            var url = URL.createObjectURL(blob);
            var a = document.createElement('a');
            a.href = url;
            a.download = 'stock_performance_' + new Date().toISOString().slice(0, 10) + '.csv';
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            URL.revokeObjectURL(url);
          }};
          window.stkSort = function(th) {{
            var col = parseInt(th.getAttribute('data-col'));
            if (_stkCol === col) {{ _stkDir *= -1; }}
            else {{ _stkCol = col; _stkDir = -1; }}
            // reset all icons
            document.querySelectorAll('#stk-table thead th[data-col] .sort-icon').forEach(function(el) {{
              el.innerHTML = '&#8645;';
              el.classList.remove('text-blue-500');
              el.classList.add('text-gray-300');
            }});
            var icon = th.querySelector('.sort-icon');
            icon.innerHTML = _stkDir === -1 ? '&#8595;' : '&#8593;';
            icon.classList.remove('text-gray-300');
            icon.classList.add('text-blue-500');
            var tbody = document.querySelector('#stk-table tbody');
            var rows = Array.from(tbody.querySelectorAll('tr'));
            rows.sort(function(a, b) {{
              var av = parseFloat(a.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              var bv = parseFloat(b.querySelectorAll('td')[col].getAttribute('data-val')) || 0;
              return _stkDir * (bv - av);
            }});
            rows.forEach(function(r) {{ tbody.appendChild(r); }});
          }};
        }})();
        </script>'''

    empty = ""
    if s["closed_trades"] == 0:
        empty = '''<div class="flex flex-col items-center justify-center py-16 text-center">
          <p class="text-lg font-medium text-gray-600">No performance data yet</p>
          <p class="mt-1 text-sm text-gray-400">Analytics will appear once trades are closed</p>
        </div>'''

    # A2 + A3: async-loaded analytics panel (benchmark alpha + bootstrap CI)
    analytics_html = '''
    <div class="glass rounded-xl p-6 mb-6">
      <div class="flex items-center justify-between mb-4">
        <h3 class="text-lg font-semibold text-gray-800">Advanced Analytics</h3>
        <span class="text-xs text-gray-400">Benchmark Alpha &amp; Statistical Confidence</span>
      </div>
      <div id="analytics-loading" class="flex items-center gap-2 text-sm text-gray-400 py-4">
        <svg class="animate-spin h-4 w-4 text-blue-400" viewBox="0 0 24 24" fill="none">
          <circle class="opacity-25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"/>
          <path class="opacity-75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"/>
        </svg>
        Loading advanced analytics (fetching Nifty50)...
      </div>
      <div id="analytics-content" class="hidden">
        <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
          <!-- A2: Benchmark Alpha -->
          <div class="bg-gray-50 rounded-lg p-4">
            <p class="text-xs font-semibold text-gray-500 uppercase tracking-wider mb-3">Benchmark Alpha vs Nifty50</p>
            <div class="grid grid-cols-3 gap-3 text-center">
              <div>
                <p class="text-xs text-gray-400">Portfolio Return</p>
                <p id="a2-portfolio" class="text-xl font-bold mt-1 text-gray-800">—</p>
              </div>
              <div>
                <p class="text-xs text-gray-400">Nifty B&amp;H</p>
                <p id="a2-benchmark" class="text-xl font-bold mt-1 text-gray-600">—</p>
              </div>
              <div>
                <p class="text-xs text-gray-400">Alpha</p>
                <p id="a2-alpha" class="text-xl font-bold mt-1">—</p>
              </div>
            </div>
            <p id="a2-period" class="text-xs text-gray-400 mt-3 text-center"></p>
          </div>
          <!-- A3: Bootstrap CI -->
          <div class="bg-gray-50 rounded-lg p-4">
            <p class="text-xs font-semibold text-gray-500 uppercase tracking-wider mb-3">95% Bootstrap CI (n=<span id="a3-n">—</span> trades)</p>
            <div class="grid grid-cols-2 gap-4">
              <div>
                <p class="text-xs text-gray-400 mb-1">Win Rate</p>
                <p id="a3-wr" class="text-xl font-bold text-gray-800">—</p>
                <p id="a3-wr-ci" class="text-xs text-gray-400 mt-0.5">CI: —</p>
              </div>
              <div>
                <p class="text-xs text-gray-400 mb-1">Profit Factor</p>
                <p id="a3-pf" class="text-xl font-bold text-gray-800">—</p>
                <p id="a3-pf-ci" class="text-xs text-gray-400 mt-0.5">CI: —</p>
              </div>
            </div>
          </div>
        </div>
      </div>
      <div id="analytics-error" class="hidden text-xs text-red-400 py-2"></div>
    </div>
    <script>
    (function() {
      fetch('/api/analytics')
        .then(r => r.json())
        .then(data => {
          document.getElementById('analytics-loading').classList.add('hidden');
          var ba = data.benchmark_alpha || {};
          var ci = data.bootstrap_ci || {};
          if (ba.error && ci.error) {
            document.getElementById('analytics-error').textContent = 'Analytics: ' + (ba.error || ci.error);
            document.getElementById('analytics-error').classList.remove('hidden');
            return;
          }
          document.getElementById('analytics-content').classList.remove('hidden');
          // A2
          if (!ba.error) {
            var port = ba.portfolio_return_pct || 0;
            var bench = ba.benchmark_return_pct || 0;
            var alpha = ba.alpha_pct || 0;
            document.getElementById('a2-portfolio').textContent = (port >= 0 ? '+' : '') + port.toFixed(2) + '%';
            document.getElementById('a2-portfolio').className = 'text-xl font-bold mt-1 ' + (port >= 0 ? 'text-emerald-600' : 'text-red-500');
            document.getElementById('a2-benchmark').textContent = (bench >= 0 ? '+' : '') + bench.toFixed(2) + '%';
            document.getElementById('a2-alpha').textContent = (alpha >= 0 ? '+' : '') + alpha.toFixed(2) + '%';
            document.getElementById('a2-alpha').className = 'text-xl font-bold mt-1 ' + (alpha >= 0 ? 'text-emerald-600' : 'text-red-500');
            document.getElementById('a2-period').textContent = (ba.first_trade_date || '') + ' → ' + (ba.last_trade_date || '');
          } else {
            document.getElementById('a2-portfolio').textContent = 'N/A';
          }
          // A3
          if (!ci.error) {
            document.getElementById('a3-n').textContent = ci.n_trades || '—';
            document.getElementById('a3-wr').textContent = (ci.win_rate || 0) + '%';
            document.getElementById('a3-wr').className = 'text-xl font-bold ' + (ci.win_rate >= 55 ? 'text-emerald-600' : ci.win_rate >= 45 ? 'text-amber-600' : 'text-red-500');
            document.getElementById('a3-wr-ci').textContent = 'CI: ' + ci.wr_ci_lower + '% – ' + ci.wr_ci_upper + '%';
            document.getElementById('a3-pf').textContent = (ci.profit_factor || 0).toFixed(2) + 'x';
            document.getElementById('a3-pf').className = 'text-xl font-bold ' + (ci.profit_factor >= 1.5 ? 'text-emerald-600' : ci.profit_factor >= 1.0 ? 'text-amber-600' : 'text-red-500');
            document.getElementById('a3-pf-ci').textContent = 'CI: ' + ci.pf_ci_lower + 'x – ' + ci.pf_ci_upper + 'x';
          } else {
            document.getElementById('a3-wr').textContent = ci.error;
          }
        })
        .catch(e => {
          document.getElementById('analytics-loading').classList.add('hidden');
          document.getElementById('analytics-error').textContent = 'Failed to load analytics: ' + e;
          document.getElementById('analytics-error').classList.remove('hidden');
        });
    })();
    </script>'''

    body = f'''
    <h2 class="text-2xl font-bold text-gray-800 mb-6">Performance Analytics</h2>
    {kpis}
    {analytics_html}
    {hz_html}
    {pat_html}
    {sec_html}
    {stk_html}
    {empty}'''
    return page_shell("Performance", "performance", body)


def _render_index_card(label, ticker, size="normal"):
    """Render a single index card. Simple: just use latest 2 available data points."""
    try:
        # Get last 30 days of data - guaranteed to have at least 2 points
        data = yf.download(ticker, period="30d", progress=False, interval="1d", multi_level_index=False)
        
        if data.empty or len(data) < 2:
            return f'''
            <div class="glass rounded-xl p-{'6' if size == 'normal' else '4'}">
              <div class="text-sm font-medium text-gray-500 uppercase tracking-wide">{_e(label)}</div>
              <div class="mt-2 text-gray-400 text-xs">No data available</div>
            </div>'''
        
        current = float(data["Close"].iloc[-1])
        prev_close = float(close_col.iloc[-2])
        pct_change = ((current - prev_close) / prev_close) * 100

        if pct_change >= 0:
            color_class = "text-emerald-600"
            badge_cls = "bg-emerald-50 text-emerald-700 border-emerald-200"
            arrow = "&#9650;"
        else:
            color_class = "text-red-600"
            badge_cls = "bg-red-50 text-red-700 border-red-200"
            arrow = "&#9660;"

        if size == "normal":
            return f'''
            <div class="glass rounded-xl p-6">
              <div class="text-sm font-medium text-gray-500 uppercase tracking-wide">{_e(label)}</div>
              <div class="mt-4 flex items-end gap-4">
                <div>
                  <div class="text-3xl font-bold text-gray-800">{current:,.2f}</div>
                  <div class="text-xs text-gray-400 mt-1">Latest Value</div>
                </div>
                <div class="ml-auto text-right">
                  <div class="{color_class} text-2xl font-bold flex items-center gap-1 justify-end">
                    {arrow} {abs(pct_change):.2f}%
                  </div>
                  <div class="text-xs text-gray-400 mt-1">Vs Previous</div>
                </div>
              </div>
              <div class="mt-4 pt-4 border-t border-gray-100">
                <div class="text-xs text-gray-500">Previous: <span class="font-medium text-gray-700">{prev_close:,.2f}</span></div>
              </div>
            </div>'''
        else:
            return f'''
            <div class="glass rounded-xl p-4 flex flex-col gap-2">
              <div class="flex items-center justify-between">
                <div class="text-xs font-semibold text-gray-500 uppercase tracking-wide">{_e(label)}</div>
                <span class="inline-flex items-center gap-1 px-2 py-0.5 rounded-full text-xs font-bold border {badge_cls}">
                  {arrow} {abs(pct_change):.2f}%
                </span>
              </div>
              </div>
              <div class="flex items-end justify-between">
                <div class="text-xl font-bold text-gray-800">{current:,.2f}</div>
                <div class="text-[10px] text-gray-400">Prev: {prev_close:,.2f}</div>
              </div>
            </div>'''

    except Exception as e:
        return f'''
        <div class="glass rounded-xl p-{'6' if size == 'normal' else '4'}">
          <div class="text-sm font-medium text-gray-500 uppercase tracking-wide">{_e(label)}</div>
          <div class="mt-2 text-red-600 text-xs">Error: {_e(str(e)[:60])}</div>
        </div>'''


def render_market_indices():
    """Render market indices dashboard showing broad-market and sector proxy indices."""
    try:
        if not _HAS_YF:
            return page_shell("Market Indices", "market",
                '''<div class="flex flex-col items-center justify-center py-16 text-center">
                  <p class="text-lg font-medium text-gray-600">yfinance module not available</p>
                  <p class="mt-1 text-sm text-gray-400">Please install yfinance to view market data</p>
                </div>''')

        # ---- Section 1: Broad Market Indices ----
        broad_indices = [
            ("NIFTY 50",           "^NSEI"),
            ("NIFTY Next 50",      "^NSMIDCP"),
            ("NIFTY Bank",         "^NSEBANK"),
        ]

        broad_cards = ""
        for label, ticker in broad_indices:
            broad_cards += _render_index_card(label, ticker, size="normal")

        # ---- Section 2: Sector Proxy Indices (from position_risk_monitor) ----
        sector_indices = [
            ("NIFTY Bank",    "^NSEBANK",    "Banking &amp; Finance"),
            ("NIFTY IT",      "^CNXIT",      "Information Technology"),
            ("NIFTY Auto",    "^CNXAUTO",    "Automobile"),
            ("NIFTY Pharma",  "^CNXPHARMA",  "Pharmaceuticals"),
            ("NIFTY Metal",   "^CNXMETAL",   "Metals &amp; Mining"),
            ("NIFTY FMCG",    "^CNXFMCG",    "Fast-Moving Consumer Goods"),
            ("NIFTY Energy",  "^CNXENERGY",  "Energy &amp; Oil/Gas"),
            ("NIFTY Realty",  "^CNXREALTY",   "Real Estate"),
            ("NIFTY Infra",   "^CNXINFRA",   "Infrastructure"),
        ]

        sector_cards = ""
        for label, ticker, _desc in sector_indices:
            sector_cards += _render_index_card(label, ticker, size="compact")

        # ---- Section 3: Sector–Index mapping reference ----
        mapping_rows = ""
        sector_map_display = [
            ("Banking",        "^NSEBANK",   "NIFTY Bank"),
            ("Finance",        "^NSEBANK",   "NIFTY Bank"),
            ("IT",             "^CNXIT",     "NIFTY IT"),
            ("Auto",           "^CNXAUTO",   "NIFTY Auto"),
            ("Pharma",         "^CNXPHARMA", "NIFTY Pharma"),
            ("Metals",         "^CNXMETAL",  "NIFTY Metal"),
            ("FMCG",           "^CNXFMCG",   "NIFTY FMCG"),
            ("Energy",         "^CNXENERGY", "NIFTY Energy"),
            ("Realty",         "^CNXREALTY",  "NIFTY Realty"),
            ("Infra",          "^CNXINFRA",  "NIFTY Infra"),
            ("Conglomerate",   "^NSEI",      "NIFTY 50 (fallback)"),
            ("Cement",         "^NSEI",      "NIFTY 50 (fallback)"),
            ("Telecom",        "^NSEI",      "NIFTY 50 (fallback)"),
            ("Media",          "^NSEI",      "NIFTY 50 (fallback)"),
            ("Chemicals",      "^NSEI",      "NIFTY 50 (fallback)"),
            ("Consumer",       "^NSEI",      "NIFTY 50 (fallback)"),
            ("Industrial",     "^NSEI",      "NIFTY 50 (fallback)"),
            ("Logistics",      "^NSEI",      "NIFTY 50 (fallback)"),
        ]
        for sector, ticker, idx_name in sector_map_display:
            fallback = "fallback" in idx_name
            tag_cls = "bg-gray-100 text-gray-500" if fallback else "bg-blue-50 text-blue-700"
            mapping_rows += f'''
              <tr class="border-b border-gray-50 hover:bg-gray-50/50">
                <td class="px-4 py-2 text-sm font-medium text-gray-700">{sector}</td>
                <td class="px-4 py-2 font-mono text-xs text-gray-500">{ticker}</td>
                <td class="px-4 py-2 text-sm">
                  <span class="inline-block px-2 py-0.5 rounded text-xs font-medium {tag_cls}">{idx_name}</span>
                </td>
              </tr>'''

        body = f'''
        <h2 class="text-2xl font-bold text-gray-800 mb-6">Market Indices</h2>

        <!-- Broad Market -->
        <div class="grid grid-cols-1 md:grid-cols-3 gap-6">
          {broad_cards}
        </div>

        <!-- Sector Indices -->
        <div class="mt-10">
          <h3 class="text-lg font-semibold text-gray-800 mb-1">Sector Indices</h3>
          <p class="text-xs text-gray-400 mb-4">Proxy tickers used by Position Risk Monitor for sector momentum checks</p>
          <div class="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-4">
            {sector_cards}
          </div>
        </div>

        <!-- Sector → Index Mapping -->
        <div class="mt-10 glass rounded-xl p-6">
          <h3 class="text-lg font-semibold text-gray-800 mb-4">Sector &rarr; Proxy Index Mapping</h3>
          <p class="text-xs text-gray-400 mb-3">Used by <span class="font-mono">position_risk_monitor.py</span> to calculate sector momentum penalties. Sectors without a dedicated index fall back to NIFTY 50.</p>
          <div class="overflow-x-auto">
            <table class="min-w-full text-left">
              <thead>
                <tr class="border-b border-gray-200">
                  <th class="px-4 py-2 text-xs font-semibold text-gray-500 uppercase">Sector</th>
                  <th class="px-4 py-2 text-xs font-semibold text-gray-500 uppercase">Ticker</th>
                  <th class="px-4 py-2 text-xs font-semibold text-gray-500 uppercase">Index Name</th>
                </tr>
              </thead>
              <tbody>
                {mapping_rows}
              </tbody>
            </table>
          </div>
        </div>

        <!-- Info Footer -->
        <div class="mt-6 p-4 bg-blue-50 rounded-lg border border-blue-100">
          <p class="text-xs text-blue-800">
            Last updated at market close. Data sourced from YFinance.
            Sector indices are used by the Tier-1 Position Risk Monitor to detect regime shifts and sector momentum divergence.
          </p>
        </div>'''

        return page_shell("Market Indices", "market", body)

    except Exception as e:
        body = f'''
        <div class="flex flex-col items-center justify-center py-16 text-center">
          <p class="text-lg font-medium text-red-600">Error loading market indices</p>
          <p class="mt-2 text-sm text-gray-500 font-mono">{_e(str(e))}</p>
        </div>'''
        return page_shell("Market Indices", "market", body)


# ============================================================
# FEEDBACK LOOP PAGE
# ============================================================
FEEDBACK_FILE = os.path.join("feedback", "feedback_log.json")
LEARNING_FILE = os.path.join("feedback", "learned_rules.json")


def _load_feedback_log():
    """Load feedback log entries."""
    try:
        with open(FEEDBACK_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return []


def _load_learned_rules():
    """Load learned rules / adjustments."""
    try:
        with open(LEARNING_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return {}


def _delete_feedback_entries(indices):
    """Delete feedback entries by index from feedback log AND database."""
    try:
        entries = _load_feedback_log()
        if not entries:
            return {"status": "error", "message": "No feedback entries to delete"}
        
        # Sort indices in reverse to avoid index shifting
        sorted_indices = sorted(set(indices), reverse=True)
        deleted_trades = []
        deleted_db_ids = []
        
        db = get_db()
        
        for idx in sorted_indices:
            if 0 <= idx < len(entries):
                entry = entries[idx]
                ticker = entry.get("ticker", "")
                entry_date = entry.get("timestamp", "")
                
                # Find matching trade in database by ticker and entry date
                if ticker and entry_date:
                    match = db.execute(
                        "SELECT id FROM trades WHERE ticker = ? AND entry_date = ?",
                        (ticker, entry_date)
                    ).fetchone()
                    if match:
                        deleted_db_ids.append(match[0])
                
                deleted_trades.append(entry.get("trade_id", f"entry_{idx}"))
                entries.pop(idx)
        
        # Save updated feedback log
        with open(FEEDBACK_FILE, "w") as f:
            json.dump(entries, f, indent=2)
        
        # Delete from database by id
        if deleted_db_ids:
            placeholders = ','.join('?' * len(deleted_db_ids))
            db.execute(f"DELETE FROM trades WHERE id IN ({placeholders})", deleted_db_ids)
            db.commit()
        
        return {
            "status": "success",
            "deleted_count": len(deleted_trades),
            "deleted_trades": deleted_trades,
            "deleted_db_count": len(deleted_db_ids),
            "remaining": len(entries)
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def _feedback_csv_bytes():
    """Generate CSV bytes from feedback log for download."""
    entries = _load_feedback_log()
    if not entries:
        return b"No feedback data available"
    # Column order
    cols = [
        "trade_id", "ticker", "sector", "direction", "patterns", "horizon_label",
        "horizon_days", "predicted_win_rate", "predicted_pf", "confidence",
        "outcome", "actual_return_pct", "exit_reason", "notes", "timestamp", "source",
    ]
    indicator_cols = ["ema_9", "ema_21", "ema_50", "rsi_14", "atr_14", "vol_ratio",
                      "price_vs_vwap", "trend_short", "rsi_zone"]
    import io, csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(cols + indicator_cols)
    for e in entries:
        row = []
        for c in cols:
            v = e.get(c, "")
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            row.append(v)
        ind = e.get("indicators_at_entry", {})
        for ic in indicator_cols:
            row.append(ind.get(ic, ""))
        w.writerow(row)
    return buf.getvalue().encode("utf-8")


def _get_shadow_trade_stats():
    """Compute shadow trade validation statistics."""
    db = get_db()
    
    # Shadow trade stats
    sh_total = db.execute("SELECT COUNT(*) FROM shadow_trades").fetchone()[0]
    sh_closed = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status NOT IN ('SHADOW_OPEN')").fetchone()[0]
    sh_wins = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status IN ('SHADOW_WON','SHADOW_EXPIRED_WIN')").fetchone()[0]
    sh_losses = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status IN ('SHADOW_LOST','SHADOW_EXPIRED_LOSS')").fetchone()[0]
    
    # Real trade stats
    real_total = db.execute("SELECT COUNT(*) FROM trades WHERE status NOT IN ('OPEN','CANCELLED')").fetchone()[0]
    real_wins = db.execute("SELECT COUNT(*) FROM trades WHERE status IN ('WON','EXPIRED_WIN')").fetchone()[0]
    real_losses = db.execute("SELECT COUNT(*) FROM trades WHERE status IN ('LOST','EXPIRED_LOSS')").fetchone()[0]
    
    # Win rates
    sh_wr = (sh_wins / sh_closed * 100) if sh_closed > 0 else 0
    real_wr = (real_wins / real_total * 100) if real_total > 0 else 0
    gap = real_wr - sh_wr  # positive = real outperforms shadow
    
    # Per-horizon comparison
    horizon_comp = []
    horizons = ['BTST_1d', 'Swing_3d', 'Swing_5d', 'Swing_10d']
    for hz in horizons:
        sh_hz = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status!='SHADOW_OPEN' AND horizon_label=?", (hz,)).fetchone()[0]
        sh_hz_wins = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status IN ('SHADOW_WON','SHADOW_EXPIRED_WIN') AND horizon_label=?", (hz,)).fetchone()[0]
        
        real_hz = db.execute("SELECT COUNT(*) FROM trades WHERE status NOT IN ('OPEN','CANCELLED') AND horizon_label=?", (hz,)).fetchone()[0]
        real_hz_wins = db.execute("SELECT COUNT(*) FROM trades WHERE status IN ('WON','EXPIRED_WIN') AND horizon_label=?", (hz,)).fetchone()[0]
        
        sh_hz_wr = (sh_hz_wins / sh_hz * 100) if sh_hz > 0 else None
        real_hz_wr = (real_hz_wins / real_hz * 100) if real_hz > 0 else None
        
        horizon_comp.append({
            'horizon': hz,
            'shadow_count': sh_hz,
            'shadow_wr': sh_hz_wr,
            'real_count': real_hz,
            'real_wr': real_hz_wr,
        })
    
    # Pattern comparison
    pattern_comp = []
    patterns_in_shadow = db.execute("SELECT DISTINCT patterns FROM shadow_trades WHERE status!='SHADOW_OPEN'").fetchall()
    
    for row in patterns_in_shadow[:10]:  # Top 10 patterns
        pat = row[0] if isinstance(row[0], str) else (row[0][0] if row[0] else '')
        if not pat:
            continue
        
        sh_pat = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status!='SHADOW_OPEN' AND patterns LIKE ?", (f'%{pat}%',)).fetchone()[0]
        sh_pat_wins = db.execute("SELECT COUNT(*) FROM shadow_trades WHERE status IN ('SHADOW_WON','SHADOW_EXPIRED_WIN') AND patterns LIKE ?", (f'%{pat}%',)).fetchone()[0]
        
        real_pat = db.execute("SELECT COUNT(*) FROM trades WHERE status NOT IN ('OPEN','CANCELLED') AND patterns LIKE ?", (f'%{pat}%',)).fetchone()[0]
        real_pat_wins = db.execute("SELECT COUNT(*) FROM trades WHERE status IN ('WON','EXPIRED_WIN') AND patterns LIKE ?", (f'%{pat}%',)).fetchone()[0]
        
        sh_pat_wr = (sh_pat_wins / sh_pat * 100) if sh_pat > 0 else None
        real_pat_wr = (real_pat_wins / real_pat * 100) if real_pat > 0 else None
        
        if sh_pat > 0 or real_pat > 0:
            pattern_comp.append({
                'pattern': pat,
                'shadow_wr': sh_pat_wr,
                'real_wr': real_pat_wr,
            })
    
    db.close()
    
    return {
        'shadow_total': sh_total,
        'shadow_closed': sh_closed,
        'shadow_wins': sh_wins,
        'shadow_losses': sh_losses,
        'shadow_wr': sh_wr,
        'real_total': real_total,
        'real_wins': real_wins,
        'real_losses': real_losses,
        'real_wr': real_wr,
        'gap': gap,  # positive = filtering working
        'efficiency': (real_wr / sh_wr) if sh_wr > 0 else 0,
        'horizon_comp': horizon_comp,
        'pattern_comp': pattern_comp,
    }


def render_feedback():
    """Render the Feedback Loop page — RAG learning visibility."""
    entries = _load_feedback_log()
    rules = _load_learned_rules()
    shadow_stats = _get_shadow_trade_stats()

    # Global Bearish Score widget (decision context should remain visible on feedback page)
    try:
        from global_sentiment import get_overnight_bearish_score
        bs = get_overnight_bearish_score()
        if bs >= 70:
            bs_bg = "bg-red-50 border-red-200"; bs_icon = "\u26a0\ufe0f"
            bs_label = "RED ALERT — BTST trims active"; bs_tc = "text-red-700"
        elif bs >= 40:
            bs_bg = "bg-amber-50 border-amber-200"; bs_icon = "\u26a1"
            bs_label = "CAUTION — Elevated bearish risk"; bs_tc = "text-amber-700"
        else:
            bs_bg = "bg-green-50 border-green-200"; bs_icon = "\u2705"
            bs_label = "SAFE — Global markets neutral"; bs_tc = "text-green-700"
    except Exception:
        bs = 30; bs_bg = "bg-gray-50 border-gray-200"
        bs_icon = "\u2014"; bs_label = "Score unavailable"; bs_tc = "text-gray-500"

    bearish_widget = f'''<div class="mb-5 p-4 rounded-xl border {bs_bg} flex items-center justify-between shadow-sm">
      <div class="flex items-center gap-3">
        <span class="text-2xl">{bs_icon}</span>
        <div>
          <p class="text-xs font-medium text-gray-500 uppercase tracking-wide">Global Bearish Score</p>
          <p class="text-xl font-bold {bs_tc}">{bs} / 100 &nbsp;&mdash;&nbsp; {bs_label}</p>
          <p class="text-xs text-gray-400 mt-0.5">S&amp;P Futures &middot; VIX &middot; DXY &middot; Oil &middot; Nikkei &middot; Hang Seng &middot; ASX</p>
        </div>
      </div>
      <div class="text-right text-xs text-gray-400 space-y-1">
        <p>BTST auto-trim: score &gt; 70</p>
        <p>Intraday trim: delta &gt; 25 pts</p>
        <p>Early-exit: trajectory &le; 40</p>
        <p>SHORT_1d gate: score &ge; 70</p>
        <a href="/api/bearish-score" target="_blank" class="text-blue-500 hover:underline">Live JSON →</a>
        <a href="/api/pending-trims" target="_blank" class="text-blue-500 hover:underline block">Pending trims →</a>
        <a href="/api/early-exits" target="_blank" class="text-blue-500 hover:underline block">Dead trades →</a>
        <a href="/api/short-trades" target="_blank" class="text-blue-500 hover:underline block">Short trades →</a>
        <a href="/api/short-force-closes" target="_blank" class="text-blue-500 hover:underline block">Short close queue →</a>
      </div>
    </div>'''

    # ---- Summary stats ----
    total_entries = len(entries)
    updated_at = rules.get("updated_at", "—")
    if updated_at and updated_at != "—":
        try:
            updated_at = datetime.fromisoformat(updated_at).strftime("%d %b %Y %H:%M")
        except Exception:
            pass

    pat_adj = rules.get("pattern_adjustments", {})
    regime_adj = rules.get("regime_adjustments", {})
    horizon_adj = rules.get("horizon_adjustments", {})
    triple_adj = rules.get("triple_adjustments", {})
    sector_adj = rules.get("sector_adjustments", {})

    filter_pen = rules.get("filter_penalties", {})
    filter_bst = rules.get("filter_boosts", {})
    hz_pen = rules.get("horizon_filter_penalties", {})
    hz_bst = rules.get("horizon_filter_boosts", {})
    sec_pen = rules.get("sector_filter_penalties", {})
    sec_bst = rules.get("sector_filter_boosts", {})
    active_rules = rules.get("rules", [])

    # Build set of penalized/boosted patterns
    penalized_patterns = set(filter_pen.keys()) | set(hz_pen.keys()) | set(sec_pen.keys())
    boosted_patterns = set(filter_bst.keys()) | set(hz_bst.keys()) | set(sec_bst.keys())

    # Outcome counts
    wins = sum(1 for e in entries if e.get("outcome") == "win")
    losses = sum(1 for e in entries if e.get("outcome") == "loss")
    fb_wr = f"{wins/total_entries*100:.1f}" if total_entries else "0"

    cards = f'''
    <div class="grid grid-cols-2 md:grid-cols-4 xl:grid-cols-6 gap-4 mb-6">
      {stat_card("Feedback Entries", total_entries, f"W: {wins} / L: {losses}", "indigo")}
      {stat_card("Feedback Win Rate", f"{fb_wr}%", "", "green" if float(fb_wr) >= 50 else "red")}
      {stat_card("Patterns Tracked", len(pat_adj), f"{len(penalized_patterns)} penalized · {len(boosted_patterns)} boosted", "amber")}
      {stat_card("Regime Rules", len(regime_adj), "", "cyan")}
      {stat_card("Learned Rules", len(active_rules), f"Updated: {_e(updated_at)}", "green")}
      {stat_card("Cross-Dim Keys", f"{len(triple_adj)} + {len(sector_adj)}", "triple + sector", "indigo")}
    </div>'''

    # ---- Shadow Trade Validation Section ----
    sh = shadow_stats
    sh_quality = "✓ Working" if sh['gap'] >= 10 else ("⚠ Moderate" if sh['gap'] >= 5 else "⚠ Weak")
    sh_quality_color = "emerald" if sh['gap'] >= 10 else ("amber" if sh['gap'] >= 5 else "red")
    gap_color = "emerald" if sh['gap'] > 0 else "red"
    # Dynamic gap message based on gap sign (positive = real outperforms, negative = shadow outperforms)
    if sh['gap'] >= 10:
        gap_msg = "✓ Filters are working as intended (real trades significantly outperform filtered signals)."
        gap_subtitle = "Real outperforms"
    elif sh['gap'] >= 5:
        gap_msg = "⚠ Monitor filter effectiveness (real trades moderately outperform filtered signals)."
        gap_subtitle = "Real outperforms"
    elif sh['gap'] > 0:
        gap_msg = "⚠ Horizon composition issue detected (real trades slightly underperform). Check real vs shadow horizon mix."
        gap_subtitle = "Real slightly ahead"
    else:
        gap_msg = "⚠ Horizon composition issue detected (filtered signals outperform real trades). Recent filter changes should improve this."
        gap_subtitle = "Shadow outperforms"
    
    # Horizon comparison rows
    hz_rows = ""
    for hz_data in sh['horizon_comp']:
        hz = hz_data['horizon']
        sh_wr = hz_data['shadow_wr']
        real_wr = hz_data['real_wr']
        
        if sh_wr is not None and real_wr is not None:
            sh_cls = "text-emerald-600" if sh_wr >= 50 else "text-red-600"
            real_cls = "text-emerald-600" if real_wr >= 50 else "text-red-600"
            gap_val = real_wr - sh_wr
            gap_cls = "text-emerald-600" if gap_val > 0 else "text-red-600"
            hz_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-xs font-medium text-gray-800">{_e(hz)}</td>
              <td class="px-4 py-2 text-right text-xs font-mono {sh_cls}">{sh_wr:.1f}%</td>
              <td class="px-4 py-2 text-right text-xs text-gray-600">{hz_data['shadow_count']}</td>
              <td class="px-4 py-2 text-right text-xs font-mono {real_cls}">{real_wr:.1f}%</td>
              <td class="px-4 py-2 text-right text-xs text-gray-600">{hz_data['real_count']}</td>
              <td class="px-4 py-2 text-right text-xs font-mono {gap_cls}">{gap_val:+.1f}pp</td>
            </tr>'''
        elif sh_wr is not None:
            sh_cls = "text-emerald-600" if sh_wr >= 50 else "text-red-600"
            hz_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-xs font-medium text-gray-800">{_e(hz)}</td>
              <td class="px-4 py-2 text-right text-xs font-mono {sh_cls}">{sh_wr:.1f}%</td>
              <td class="px-4 py-2 text-right text-xs text-gray-600">{hz_data['shadow_count']}</td>
              <td class="px-4 py-2 text-right text-xs text-gray-400">—</td>
              <td class="px-4 py-2 text-right text-xs text-gray-600">0</td>
              <td class="px-4 py-2 text-right text-xs text-gray-400">N/A</td>
            </tr>'''
    
    shadow_section = f'''
    <div class="glass rounded-xl p-6 mb-6">
      <div class="flex items-center justify-between mb-4">
        <div>
          <h3 class="text-lg font-semibold text-gray-800">Shadow Trade Validation</h3>
          <p class="text-xs text-gray-500 mt-1">Filter quality assessment: Real trades vs Filtered signal performance</p>
        </div>
      </div>
      
      <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-6">
        {stat_card("Shadow Win Rate", f"{sh['shadow_wr']:.1f}%", f"{sh['shadow_closed']} closed", "amber")}
        {stat_card("Real Win Rate", f"{sh['real_wr']:.1f}%", f"{sh['real_total']} closed", "green")}
        {stat_card("Filter Gap", f"{sh['gap']:+.1f}pp", gap_subtitle, gap_color)}
        {stat_card("Filter Efficiency", f"{sh['efficiency']:.2f}x", "Real / Shadow ratio", sh_quality_color)}
      </div>
      
      <div class="mb-4">
        <h4 class="text-sm font-semibold text-gray-700 mb-2">Interpretation</h4>
        <div class="text-sm text-gray-600 bg-{sh_quality_color}-50 border border-{sh_quality_color}-200 rounded-lg p-4">
          {sh_quality}: {'Filtered signals underperform real trades by' if sh['gap'] >= 0 else 'Real trades underperform filtered signals by'} <span class="font-semibold">{abs(sh['gap']):.1f} percentage points</span>.
          Your filters are <span class="font-semibold">{sh['efficiency']:.2f}x more effective</span> than random selections.
          {gap_msg}
        </div>
      </div>
      
      <h4 class="text-sm font-semibold text-gray-700 mb-3">Performance by Horizon</h4>
      <div class="overflow-x-auto">
        <table class="w-full text-sm"><thead><tr class="border-b border-gray-200">
          <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Horizon</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Shadow WR</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Shadow N</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Real WR</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Real N</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Gap</th>
        </tr></thead><tbody>{hz_rows}</tbody></table>
      </div>
    </div>'''

    # ---- Section 2: Pattern Adjustments ----
    pat_rows = ""
    if pat_adj:
        for pname, pdata in sorted(pat_adj.items()):
            actual_wr = pdata.get("actual_win_rate", 0)
            decay_wr = pdata.get("decay_weighted_win_rate", 0)
            avg_ret = pdata.get("avg_return", 0)
            total_t = pdata.get("total_trades", 0)
            vol = pdata.get("volume_breakdown", {})
            vol_str = " · ".join(f"{k}: {v}" for k, v in sorted(vol.items())) if vol else "—"

            # Status
            if pname in penalized_patterns:
                st_badge = '<span class="inline-flex items-center px-2 py-0.5 rounded-full text-xs font-medium bg-red-50 text-red-700 border border-red-200">Penalized</span>'
            elif pname in boosted_patterns:
                st_badge = '<span class="inline-flex items-center px-2 py-0.5 rounded-full text-xs font-medium bg-emerald-50 text-emerald-700 border border-emerald-200">Boosted</span>'
            else:
                st_badge = '<span class="inline-flex items-center px-2 py-0.5 rounded-full text-xs font-medium bg-gray-100 text-gray-600 border border-gray-200">Neutral</span>'

            wr_cls = "text-emerald-600" if actual_wr >= 50 else "text-red-600"
            ret_cls = "text-emerald-600" if avg_ret >= 0 else "text-red-600"

            pat_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2.5 font-medium text-gray-800 text-xs">{_e(pname.replace("_", " ").title())}</td>
              <td class="px-4 py-2.5 text-right font-mono {wr_cls}">{actual_wr:.1f}%</td>
              <td class="px-4 py-2.5 text-right font-mono text-gray-600">{decay_wr:.1f}%</td>
              <td class="px-4 py-2.5 text-right font-mono {ret_cls}">{avg_ret:+.2f}%</td>
              <td class="px-4 py-2.5 text-right text-gray-600">{total_t}</td>
              <td class="px-4 py-2.5 text-xs text-gray-500">{_e(vol_str)}</td>
              <td class="px-4 py-2.5 text-center">{st_badge}</td>
            </tr>'''

    pat_section = f'''
    <div class="glass rounded-xl p-6 mb-6">
      <div class="flex items-center justify-between mb-4">
        <h3 class="text-lg font-semibold text-gray-800">Pattern Adjustments</h3>
        <span class="text-xs text-gray-400">{len(pat_adj)} patterns · temporal decay half-life 60 days</span>
      </div>
      <div class="overflow-x-auto">
        <table class="w-full text-sm"><thead><tr class="border-b border-gray-200">
          <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Pattern</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Actual WR</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Decay WR</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Avg Return</th>
          <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Trades</th>
          <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Volume Breakdown</th>
          <th class="px-4 py-2 text-center text-xs text-gray-500 uppercase">Status</th>
        </tr></thead><tbody>{pat_rows}</tbody></table>
      </div>
    </div>''' if pat_adj else ""

    # ---- Section 3: Cross-Dimensional Intelligence ----

    def _adj_table(title, adj_dict, key_label="Key"):
        """Build a collapsible table for adjustment dicts."""
        if not adj_dict:
            return ""
        rows = ""
        for k, v in sorted(adj_dict.items()):
            actual_wr = v.get("actual_win_rate", 0)
            decay_wr = v.get("decay_weighted_win_rate", 0)
            avg_ret = v.get("avg_return", 0)
            total_t = v.get("total_trades", 0)
            wr_cls = "text-emerald-600" if actual_wr >= 50 else "text-red-600"
            ret_cls = "text-emerald-600" if avg_ret >= 0 else "text-red-600"
            # Format the key nicely
            display_key = k.replace("_", " ").replace("|", " → ").title()
            rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-xs text-gray-800">{_e(display_key)}</td>
              <td class="px-4 py-2 text-right font-mono {wr_cls}">{actual_wr:.1f}%</td>
              <td class="px-4 py-2 text-right font-mono text-gray-600">{decay_wr:.1f}%</td>
              <td class="px-4 py-2 text-right font-mono {ret_cls}">{avg_ret:+.2f}%</td>
              <td class="px-4 py-2 text-right text-gray-600">{total_t}</td>
            </tr>'''
        uid = title.lower().replace(" ", "_").replace("-", "_")
        return f'''
        <div class="glass rounded-xl mb-4 overflow-hidden">
          <button onclick="document.getElementById('adj_{uid}').classList.toggle('hidden'); this.querySelector('.chevron').classList.toggle('rotate-90')" class="w-full flex items-center justify-between px-6 py-4 hover:bg-gray-50 transition-colors">
            <div class="flex items-center gap-3">
              <h4 class="text-sm font-semibold text-gray-700">{_e(title)}</h4>
              <span class="text-xs font-medium text-gray-400 bg-gray-100 px-2 py-0.5 rounded-full">{len(adj_dict)} entries</span>
            </div>
            <svg class="w-4 h-4 text-gray-400 chevron transition-transform" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 5l7 7-7 7"/></svg>
          </button>
          <div id="adj_{uid}" class="hidden">
            <div class="overflow-x-auto px-6 pb-4">
              <table class="w-full text-sm"><thead><tr class="border-b border-gray-200">
                <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">{_e(key_label)}</th>
                <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Actual WR</th>
                <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Decay WR</th>
                <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Avg Return</th>
                <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Trades</th>
              </tr></thead><tbody>{rows}</tbody></table>
            </div>
          </div>
        </div>'''

    cross_dim = f'''
    <div class="mb-6">
      <h3 class="text-lg font-semibold text-gray-800 mb-4">Cross-Dimensional Intelligence</h3>
      {_adj_table("Regime Adjustments", regime_adj, "Regime")}
      {_adj_table("Horizon Adjustments", horizon_adj, "Horizon")}
      {_adj_table("Sector Adjustments", sector_adj, "Sector")}
      {_adj_table("Triple Adjustments (Pattern × Regime × Horizon)", triple_adj, "Combination")}
    </div>'''

    # ---- Section 3b: Filter Penalties & Boosts ----
    def _filter_table(title, pen_dict, bst_dict, uid_prefix):
        if not pen_dict and not bst_dict:
            return ""
        rows = ""
        all_keys = sorted(set(list(pen_dict.keys()) + list(bst_dict.keys())))
        for k in all_keys:
            pen_val = pen_dict.get(k)
            bst_val = bst_dict.get(k)
            display = k.replace("_", " ").replace("|", " → ").title()
            # Values can be dicts with {actual_wr, trades, action, reason} or plain floats
            if isinstance(pen_val, dict):
                wr = pen_val.get("actual_wr", 0)
                action = pen_val.get("action", "")
                reason = pen_val.get("reason", "")
                pen_html = f'<span class="font-mono text-red-600">{wr:.1f}% WR</span> <span class="text-red-400 text-[10px]">({_e(action)})</span>'
                pen_tip = reason
            elif pen_val is not None:
                pen_html = f'<span class="font-mono text-red-600">{float(pen_val):.3f}</span>'
                pen_tip = ""
            else:
                pen_html = '<span class="text-gray-300">—</span>'
                pen_tip = ""

            if isinstance(bst_val, dict):
                wr = bst_val.get("actual_wr", 0)
                action = bst_val.get("action", "")
                reason = bst_val.get("reason", "")
                bst_html = f'<span class="font-mono text-emerald-600">{wr:.1f}% WR</span> <span class="text-emerald-400 text-[10px]">({_e(action)})</span>'
                bst_tip = reason
            elif bst_val is not None:
                bst_html = f'<span class="font-mono text-emerald-600">{float(bst_val):.3f}</span>'
                bst_tip = ""
            else:
                bst_html = '<span class="text-gray-300">—</span>'
                bst_tip = ""

            rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-xs text-gray-800">{_e(display)}</td>
              <td class="px-4 py-2 text-center" title="{_e(pen_tip)}">{pen_html}</td>
              <td class="px-4 py-2 text-center" title="{_e(bst_tip)}">{bst_html}</td>
            </tr>'''
        return f'''
        <div class="glass rounded-xl mb-4 overflow-hidden">
          <button onclick="document.getElementById('{uid_prefix}_tbl').classList.toggle('hidden'); this.querySelector('.chevron').classList.toggle('rotate-90')" class="w-full flex items-center justify-between px-6 py-4 hover:bg-gray-50 transition-colors">
            <div class="flex items-center gap-3">
              <h4 class="text-sm font-semibold text-gray-700">{_e(title)}</h4>
              <span class="text-xs font-medium text-gray-400 bg-gray-100 px-2 py-0.5 rounded-full">{len(all_keys)} entries</span>
            </div>
            <svg class="w-4 h-4 text-gray-400 chevron transition-transform" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 5l7 7-7 7"/></svg>
          </button>
          <div id="{uid_prefix}_tbl" class="hidden">
            <div class="overflow-x-auto px-6 pb-4">
              <table class="w-full text-sm"><thead><tr class="border-b border-gray-200">
                <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Key</th>
                <th class="px-4 py-2 text-center text-xs text-gray-500 uppercase">Penalty</th>
                <th class="px-4 py-2 text-center text-xs text-gray-500 uppercase">Boost</th>
              </tr></thead><tbody>{rows}</tbody></table>
            </div>
          </div>
        </div>'''

    filter_section = f'''
    <div class="mb-6">
      <h3 class="text-lg font-semibold text-gray-800 mb-4">Active Filters</h3>
      {_filter_table("Pattern Filters", filter_pen, filter_bst, "flt_pat")}
      {_filter_table("Horizon Filters", hz_pen, hz_bst, "flt_hz")}
      {_filter_table("Sector Filters", sec_pen, sec_bst, "flt_sec")}
    </div>'''

    # ---- Section 4: Raw Feedback Log ----
    log_rows = ""
    for idx, e in enumerate(reversed(entries)):  # newest first
        outcome = e.get("outcome", "")
        out_cls = "text-emerald-600 font-semibold" if outcome == "win" else "text-red-600 font-semibold"
        ret_val = e.get("actual_return_pct")
        ret_cls = "text-emerald-600" if (ret_val or 0) >= 0 else "text-red-600"
        pats = e.get("patterns", [])
        pat_str = ", ".join(pats) if isinstance(pats, list) else str(pats)
        ts = e.get("timestamp", "")
        if ts:
            try:
                ts = datetime.fromisoformat(ts).strftime("%d %b %y %H:%M")
            except Exception:
                pass
        actual_idx = total_entries - 1 - idx  # Map to original list index
        log_rows += f'''
        <tr class="hover:bg-blue-50/50 border-b border-gray-100 feedback-row" data-index="{actual_idx}">
          <td class="px-3 py-2 text-xs"><input type="checkbox" class="feedback-checkbox" data-index="{actual_idx}" onchange="updateSelectionInfo()"></td>
          <td class="px-3 py-2 text-xs text-gray-500 whitespace-nowrap">{_e(ts)}</td>
          <td class="px-3 py-2 text-xs font-medium text-gray-800">{_e(_ticker(e.get("ticker","")))}</td>
          <td class="px-3 py-2 text-xs text-gray-600">{_e(e.get("direction",""))}</td>
          <td class="px-3 py-2 text-xs text-gray-600 max-w-[200px] truncate" title="{_e(pat_str)}">{_e(pat_str)}</td>
          <td class="px-3 py-2 text-xs text-gray-600">{_e(e.get("horizon_label",""))}</td>
          <td class="px-3 py-2 text-xs text-right text-gray-600">{e.get("predicted_win_rate","—")}</td>
          <td class="px-3 py-2 text-xs text-right {out_cls}">{_e(outcome)}</td>
          <td class="px-3 py-2 text-xs text-right font-mono {ret_cls}">{_pct(ret_val)}</td>
          <td class="px-3 py-2 text-xs text-gray-500">{_e(e.get("exit_reason",""))}</td>
        </tr>'''

    log_section = f'''
    <div class="glass rounded-xl p-6 mb-6">
      <div class="flex items-center justify-between mb-4">
        <div>
          <h3 class="text-lg font-semibold text-gray-800">Raw Feedback Log</h3>
          <p class="text-xs text-gray-500 mt-1" id="selection-info">No entries selected</p>
        </div>
        <div class="flex gap-2">
          <a href="/feedback/download" class="inline-flex items-center gap-2 px-4 py-2 bg-blue-600 text-white text-xs font-medium rounded-lg hover:bg-blue-700 transition-colors shadow-sm">
            <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 10v6m0 0l-3-3m3 3l3-3m2 8H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>
            Download CSV
          </a>
          <button onclick="deleteFeedback()" id="delete-btn" class="inline-flex items-center gap-2 px-4 py-2 bg-red-600 text-white text-xs font-medium rounded-lg hover:bg-red-700 transition-colors shadow-sm disabled:opacity-50 disabled:cursor-not-allowed" disabled>
            <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M19 7l-.867 12.142A2 2 0 0116.138 21H7.862a2 2 0 01-1.995-1.858L5 7m5 4v6m4-6v6m1-10V4a1 1 0 00-1-1h-4a1 1 0 00-1 1v3M4 7h16"/></svg>
            Delete Selected
          </button>
          <button onclick="toggleSelectAll()" class="inline-flex items-center gap-2 px-4 py-2 bg-gray-600 text-white text-xs font-medium rounded-lg hover:bg-gray-700 transition-colors shadow-sm">
            <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/></svg>
            <span id="select-all-text">Select All</span>
          </button>
        </div>
      </div>
      <div class="overflow-x-auto max-h-[500px] overflow-y-auto scrollbar-thin">
        <table class="w-full text-sm"><thead class="sticky top-0 bg-white"><tr class="border-b border-gray-200">
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">✓</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Timestamp</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Ticker</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Dir</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Patterns</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Horizon</th>
          <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Pred WR</th>
          <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Outcome</th>
          <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Return</th>
          <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Exit Reason</th>
        </tr></thead><tbody>{log_rows}</tbody></table>
      </div>
    </div>
    <script>
    function toggleSelectAll() {{
      const checkboxes = document.querySelectorAll('.feedback-checkbox');
      const btn = document.getElementById('select-all-text');
      const allChecked = Array.from(checkboxes).every(cb => cb.checked);
      checkboxes.forEach(cb => cb.checked = !allChecked);
      btn.textContent = allChecked ? 'Select All' : 'Deselect All';
      updateSelectionInfo();
    }}
    function updateSelectionInfo() {{
      const checkboxes = document.querySelectorAll('.feedback-checkbox:checked');
      const count = checkboxes.length;
      const info = document.getElementById('selection-info');
      const btn = document.getElementById('delete-btn');
      if (count === 0) {{
        info.textContent = 'No entries selected';
        btn.disabled = true;
      }} else if (count <= 10) {{
        info.textContent = `${{count}} entries selected`;
        btn.disabled = false;
      }} else {{
        info.textContent = `${{count}} entries selected (Max 10 to delete)`;
        btn.disabled = true;
      }}
    }}
    function deleteFeedback() {{
      const checkboxes = document.querySelectorAll('.feedback-checkbox:checked');
      const count = checkboxes.length;
      if (count === 0) {{
        alert('Please select at least one entry to delete');
        return;
      }}
      if (count > 10) {{
        alert('Maximum 10 entries can be deleted at once');
        return;
      }}
      if (!confirm(`Delete ${{count}} feedback entries? This cannot be undone.`)) return;
      const indices = Array.from(checkboxes).map(cb => cb.dataset.index).join(',');
      fetch('/feedback/delete', {{
        method: 'POST',
        body: new URLSearchParams({{indices: indices}})
      }}).then(r => r.json()).then(data => {{
        if (data.status === 'success') {{
          alert(`Deleted ${{data.deleted_count}} entries`);
          location.reload();
        }} else {{
          alert(`Error: ${{data.message || 'Unknown error'}}`);
        }}
      }}).catch(e => {{
        alert(`Error: ${{e.message}}`);
      }});
    }}
    </script>'''

    # ---- Learned Rules ----
    rules_html = ""
    if active_rules:
        rule_items = ""
        for r in active_rules:
            rule_items += f'''
            <div class="glass rounded-lg p-4 mb-2">
              <div class="text-sm text-gray-800">{_e(json.dumps(r, indent=2) if isinstance(r, dict) else str(r))}</div>
            </div>'''
        rules_html = f'''
        <div class="mb-6">
          <h3 class="text-lg font-semibold text-gray-800 mb-4">Learned Rules</h3>
          {rule_items}
        </div>'''

    body = f'''
    <div class="flex items-center justify-between mb-6">
      <div>
        <h2 class="text-2xl font-bold text-gray-800">RAG Feedback Loop</h2>
        <p class="text-sm text-gray-500 mt-1">How the system learns from trade outcomes — pattern penalties, regime adjustments, and cross-dimensional intelligence</p>
      </div>
    </div>
    {bearish_widget}
    {cards}
    {shadow_section}
    {pat_section}
    {filter_section}
    {cross_dim}
    {rules_html}
    {log_section}'''
    return page_shell("Feedback Loop", "feedback", body)


def render_engine(action_result=None):
    scan_log = q_scan_log()
    log_lines = get_engine_log()
    status = get_engine_status()
    pending = get_pending_signals()

    # Global Bearish Score widget (for decision-making context during scan/approval)
    try:
        from global_sentiment import get_overnight_bearish_score
        bs = get_overnight_bearish_score()
        if bs >= 70:
            bs_bg = "bg-red-50 border-red-200"; bs_icon = "\u26a0\ufe0f"
            bs_label = "RED ALERT — BTST trims active"; bs_tc = "text-red-700"
        elif bs >= 40:
            bs_bg = "bg-amber-50 border-amber-200"; bs_icon = "\u26a1"
            bs_label = "CAUTION — Elevated bearish risk"; bs_tc = "text-amber-700"
        else:
            bs_bg = "bg-green-50 border-green-200"; bs_icon = "\u2705"
            bs_label = "SAFE — Global markets neutral"; bs_tc = "text-green-700"
    except Exception:
        bs = 30; bs_bg = "bg-gray-50 border-gray-200"
        bs_icon = "\u2014"; bs_label = "Score unavailable"; bs_tc = "text-gray-500"

    bearish_widget = f'''<div class="mb-5 p-4 rounded-xl border {bs_bg} flex items-center justify-between shadow-sm">
      <div class="flex items-center gap-3">
        <span class="text-2xl">{bs_icon}</span>
        <div>
          <p class="text-xs font-medium text-gray-500 uppercase tracking-wide">Global Bearish Score (Decision Context)</p>
          <p class="text-xl font-bold {bs_tc}">{bs} / 100 &nbsp;&mdash;&nbsp; {bs_label}</p>
          <p class="text-xs text-gray-400 mt-0.5">S&amp;P Futures &middot; VIX &middot; DXY &middot; Oil &middot; Nikkei &middot; Hang Seng &middot; ASX</p>
        </div>
      </div>
      <div class="text-right text-xs text-gray-400 space-y-1">
        <p>BTST auto-trim: score &gt; 70</p>
        <p>Intraday trim: delta &gt; 25 pts</p>
        <p>Early-exit: trajectory &le; 40</p>
        <p>SHORT_1d gate: score &ge; 70</p>
      </div>
    </div>'''

    # Action buttons - disabled while engine is running
    disabled = 'opacity-50 pointer-events-none' if status['running'] else ''
    buttons = f'''
    <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-4 gap-4 mb-6">
      <form method="POST" action="/engine?action=run">
        <button type="submit" class="w-full glass rounded-xl p-6 text-left hover:border-blue-400 transition-all group {disabled}">
          <div class="flex items-center gap-3 mb-2">
            <div class="w-10 h-10 rounded-xl bg-blue-50 flex items-center justify-center">
              <svg class="w-5 h-5 text-blue-600" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M14.752 11.168l-3.197-2.132A1 1 0 0010 9.87v4.263a1 1 0 001.555.832l3.197-2.132a1 1 0 000-1.664z"/><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M21 12a9 9 0 11-18 0 9 9 0 0118 0z"/></svg>
            </div>
            <div>
              <p class="font-semibold text-gray-800 group-hover:text-blue-600 transition">Full Run</p>
              <p class="text-xs text-gray-400">Catch-up + Scan + Monitor + Report</p>
            </div>
          </div>
        </button>
      </form>
      <form method="POST" action="/engine?action=scan">
        <button type="submit" class="w-full glass rounded-xl p-6 text-left hover:border-emerald-400 transition-all group {disabled}">
          <div class="flex items-center gap-3 mb-2">
            <div class="w-10 h-10 rounded-xl bg-emerald-50 flex items-center justify-center">
              <svg class="w-5 h-5 text-emerald-600" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M21 21l-6-6m2-5a7 7 0 11-14 0 7 7 0 0114 0z"/></svg>
            </div>
            <div>
              <p class="font-semibold text-gray-800 group-hover:text-emerald-600 transition">Scan Only</p>
              <p class="text-xs text-gray-400">Scan + auto-enter signals</p>
            </div>
          </div>
        </button>
      </form>
      <form method="POST" action="/engine?action=scan_preview">
        <button type="submit" class="w-full glass rounded-xl p-6 text-left hover:border-purple-400 transition-all group {disabled}">
          <div class="flex items-center gap-3 mb-2">
            <div class="w-10 h-10 rounded-xl bg-purple-50 flex items-center justify-center">
              <svg class="w-5 h-5 text-purple-600" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 5H7a2 2 0 00-2 2v12a2 2 0 002 2h10a2 2 0 002-2V7a2 2 0 00-2-2h-2M9 5a2 2 0 002 2h2a2 2 0 002-2M9 5a2 2 0 012-2h2a2 2 0 012 2m-6 9l2 2 4-4"/></svg>
            </div>
            <div>
              <p class="font-semibold text-gray-800 group-hover:text-purple-600 transition">Scan & Review</p>
              <p class="text-xs text-gray-400">Scan signals, approve manually</p>
            </div>
          </div>
        </button>
      </form>
      <form method="POST" action="/engine?action=monitor">
        <button type="submit" class="w-full glass rounded-xl p-6 text-left hover:border-amber-400 transition-all group {disabled}">
          <div class="flex items-center gap-3 mb-2">
            <div class="w-10 h-10 rounded-xl bg-amber-50 flex items-center justify-center">
              <svg class="w-5 h-5 text-amber-600" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 12a3 3 0 11-6 0 3 3 0 016 0z"/><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M2.458 12C3.732 7.943 7.523 5 12 5c4.478 0 8.268 2.943 9.542 7-1.274 4.057-5.064 7-9.542 7-4.477 0-8.268-2.943-9.542-7z"/></svg>
            </div>
            <div>
              <p class="font-semibold text-gray-800 group-hover:text-amber-600 transition">Monitor Only</p>
              <p class="text-xs text-gray-400">Check open positions for SL/target</p>
            </div>
          </div>
        </button>
      </form>
    </div>'''

    # Live output panel (shows when engine is running OR just finished)
    live_html = ""
    if status['running'] or status['done']:
        status_label = '<span class="inline-flex items-center gap-2"><span class="w-2 h-2 rounded-full bg-blue-500 animate-pulse"></span> Running...</span>' if status['running'] else (
            '<span class="text-emerald-600 font-semibold">Completed Successfully</span>' if status['success'] else '<span class="text-red-600 font-semibold">Failed</span>'
        )
        lines_text = _e("\n".join(status['lines']))
        live_html = f'''
        <div id="live-output" class="glass rounded-xl p-5 border-blue-300 mb-6 fade-in">
          <div class="flex items-center justify-between mb-3">
            <div class="flex items-center gap-3">
              <h3 class="text-lg font-semibold text-gray-800">Engine Output</h3>
              <span class="text-xs px-2 py-1 rounded-full bg-blue-50 text-blue-700">{_e(status['action']).upper()}</span>
            </div>
            <div id="status-indicator" class="text-sm">{status_label}</div>
          </div>
          <pre id="output-log" class="text-xs text-gray-700 bg-gray-50 rounded-lg p-4 max-h-96 overflow-y-auto scrollbar-thin whitespace-pre-wrap font-mono leading-relaxed border border-gray-200">{lines_text}</pre>
          <div class="flex items-center justify-between mt-3">
            <span id="line-count" class="text-xs text-gray-400">{len(status['lines'])} lines</span>
            <span class="text-xs text-gray-400">Started: {status['started_at'][:19] if status['started_at'] else ''}</span>
          </div>
        </div>
        '''
        # Add polling JS only while running
        if status['running']:
            live_html += '''
        <script>
        (function() {
          let prevLen = 0;
          function poll() {
            fetch('/engine/stream')
              .then(r => r.json())
              .then(data => {
                const el = document.getElementById('output-log');
                const si = document.getElementById('status-indicator');
                const lc = document.getElementById('line-count');
                if (el) {
                  el.textContent = data.lines.join('\\n');
                  el.scrollTop = el.scrollHeight;
                }
                if (lc) lc.textContent = data.lines.length + ' lines';
                if (data.done) {
                  if (si) si.innerHTML = data.success
                    ? '<span class="text-emerald-600 font-semibold">Completed Successfully</span>'
                    : '<span class="text-red-600 font-semibold">Failed</span>';
                  // Re-enable buttons
                  document.querySelectorAll('form button').forEach(b => {
                    b.classList.remove('opacity-50', 'pointer-events-none');
                  });
                  // If scan_preview just finished, check for pending signals and reload
                  if (data.action === 'scan_preview' && data.success) {
                    fetch('/engine/pending').then(r => r.json()).then(p => {
                      if (p.has_pending) setTimeout(() => location.reload(), 500);
                    });
                  }
                  // If approve just finished, reload to clear review panel
                  if (data.action === 'approve' && data.success) {
                    setTimeout(() => location.reload(), 500);
                  }
                } else {
                  setTimeout(poll, 800);
                }
              })
              .catch(() => setTimeout(poll, 2000));
          }
          setTimeout(poll, 800);
        })();
        </script>'''

    # Action result (legacy — for non-streaming fallback)
    result_html = ""

    # Scan history
    scan_html = ""
    if scan_log:
        scan_rows = ""
        for s in scan_log:
            scan_rows += f'''
            <tr class="hover:bg-blue-50/50 border-b border-gray-100">
              <td class="px-4 py-2 text-gray-800">{_date(s["scan_date"])}</td>
              <td class="px-4 py-2 text-right text-gray-600">{s["tickers_scanned"]}</td>
              <td class="px-4 py-2 text-right text-gray-600">{s["signals_found"]}</td>
              <td class="px-4 py-2 text-right text-emerald-600">{s["trades_entered"]}</td>
              <td class="px-4 py-2 text-right text-red-600">{s["errors"]}</td>
              <td class="px-4 py-2 text-right text-gray-600">{s.get("duration_seconds",0):.1f}s</td>
            </tr>'''
        scan_html = f'''
        <div class="glass rounded-xl p-6 mb-6">
          <h3 class="text-lg font-semibold text-gray-800 mb-4">Scan History</h3>
          <div class="overflow-x-auto scrollbar-thin">
            <table class="w-full text-sm"><thead><tr class="border-b border-gray-200">
              <th class="px-4 py-2 text-left text-xs text-gray-500 uppercase">Date</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Scanned</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Signals</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Entered</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Errors</th>
              <th class="px-4 py-2 text-right text-xs text-gray-500 uppercase">Duration</th>
            </tr></thead><tbody>{scan_rows}</tbody></table>
          </div>
        </div>'''

    # Engine log
    log_text = _e("".join(log_lines)) if log_lines else '<span class="italic text-gray-400">No log entries yet</span>'
    log_html = f'''
    <div class="glass rounded-xl p-6">
      <div class="flex items-center justify-between mb-4">
        <h3 class="text-lg font-semibold text-gray-800">Engine Log</h3>
        <a href="/engine" class="text-xs text-gray-400 hover:text-gray-800 transition">Refresh</a>
      </div>
      <div class="bg-gray-50 rounded-lg p-4 max-h-80 overflow-y-auto scrollbar-thin border border-gray-200">
        <pre class="text-xs text-gray-600 font-mono leading-relaxed whitespace-pre-wrap">{log_text}</pre>
      </div>
    </div>'''

    # Signal Review Panel (shows when pending signals exist)
    review_html = ""
    if pending and not status['running']:
        total = pending.get("total_signals", 0)
        qualifying = pending.get("qualifying", 0)
        filtered_out = pending.get("filtered_out", 0)
        scan_dt = pending.get("scan_date", "")
        skip_summary = pending.get("skip_reason_summary", {})
        signals = pending.get("signals", [])
        skipped = pending.get("skipped", [])

        # Skip reason badges
        skip_badges = ""
        for reason, cnt in sorted(skip_summary.items(), key=lambda x: -x[1]):
            colors = {"Low Win Rate": "red", "Low Confidence": "amber",
                      "Low R:R Ratio": "orange", "Duplicate Trade": "gray"}.get(reason, "gray")
            skip_badges += f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs bg-{colors}-50 text-{colors}-700">{_e(reason)}: {cnt}</span> '

        # Summary bar
        summary_bar = f'''
        <div class="glass rounded-xl p-5 mb-6 border-purple-300 fade-in">
          <div class="flex items-center justify-between mb-4">
            <div class="flex items-center gap-3">
              <h3 class="text-lg font-semibold text-gray-800">Signal Review</h3>
              <span class="text-xs px-2 py-1 rounded-full bg-purple-50 text-purple-700">Scan {_e(scan_dt)}</span>
            </div>
            <span class="text-xs text-gray-400">Awaiting your approval</span>
          </div>
          <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-4">
            <div class="bg-gray-50 rounded-lg p-3 text-center">
              <p class="text-2xl font-bold text-gray-800">{total}</p>
              <p class="text-xs text-gray-500">Total Signals</p>
            </div>
            <div class="bg-emerald-50 rounded-lg p-3 text-center">
              <p class="text-2xl font-bold text-emerald-700">{qualifying}</p>
              <p class="text-xs text-emerald-600">Qualifying</p>
            </div>
            <div class="bg-red-50 rounded-lg p-3 text-center">
              <p class="text-2xl font-bold text-red-700">{filtered_out}</p>
              <p class="text-xs text-red-600">Filtered Out</p>
            </div>
            <div class="bg-blue-50 rounded-lg p-3 text-center">
              <p class="text-2xl font-bold text-blue-700">{pending.get("duration", 0):.0f}s</p>
              <p class="text-xs text-blue-600">Scan Duration</p>
            </div>
          </div>
          <div class="flex flex-wrap gap-2 mb-1">
            <span class="text-xs text-gray-500 font-medium">Skip Reasons:</span>
            {skip_badges if skip_badges else '<span class="text-xs text-gray-400 italic">None</span>'}
          </div>
        </div>'''

        # Fetch live prices for qualifying signal tickers
        all_tickers = list(set(sig.get("ticker", "") for sig in signals + skipped))
        live_prices = fetch_live_prices(all_tickers) if all_tickers else {}

        # Qualifying signals table
        if signals:
            # Extract unique filter values from signals
            confidents = sorted(set(sig.get("confidence", "-") for sig in signals))
            sectors = sorted(set((sig.get("sector", "-") or "-") for sig in signals))
            horizons = sorted(set(sig.get("horizon_label", "-") for sig in signals))
            
            # Win rate ranges
            win_ranges = [
                ("All", None, None),
                ("\u2265 55%", 55, None),
                ("40-50%", 40, 50),
                ("50-60%", 50, 60),
                ("60-70%", 60, 70),
                ("70%+", 70, None),
            ]
            
            # Extract all patterns
            all_patterns = set()
            for sig in signals:
                patterns = sig.get("patterns", "-")
                if patterns and patterns != "-":
                    for p in patterns.split(","):
                        all_patterns.add(p.strip())
            patterns = sorted(all_patterns)
            
            sig_rows = ""
            for i, sig in enumerate(signals):
                dir_color = "emerald"
                dir_icon = "&#9650;"
                wr = sig.get("predicted_win_rate", 0)
                wr_color = "emerald" if wr >= 60 else ("amber" if wr >= 55 else "red")
                rr = sig.get("rr_ratio", 0)
                rr_val = f"{rr:.1f}x" if rr else "-"
                conf = sig.get("confidence", "-")
                conf_color = {"HIGH": "emerald", "MEDIUM": "amber", "LOW": "red"}.get(conf, "gray")
                sector = _e(sig.get("sector", "-") or "-")
                patterns_str = _e(sig.get("patterns", "-") or "-")
                if len(patterns_str) > 30:
                    patterns_str = patterns_str[:28] + ".."
                # Current market price
                cmp = live_prices.get(sig.get("ticker", ""))
                entry_p = sig.get("entry_price", 0)
                if cmp:
                    cmp_diff = ((cmp - entry_p) / entry_p * 100) if entry_p else 0
                    cmp_color = "emerald" if cmp_diff >= 0 else "red"
                    cmp_html = f'{cmp:.2f} <span class="text-xs text-{cmp_color}-500">({cmp_diff:+.1f}%)</span>'
                else:
                    cmp_html = '<span class="text-gray-400">-</span>'
                
                # Data attributes for filtering
                horizon_attr = _e(sig.get("horizon_label", ""))
                sector_attr = _e(sig.get("sector", "-") or "-")
                conf_attr = _e(conf)
                wr_attr = int(wr)
                patterns_attr = _e(sig.get("patterns", "-") or "-")

                sig_rows += f'''
                <tr class="hover:bg-purple-50/50 border-b border-gray-100 signal-row" 
                    data-conf="{conf_attr}" data-sector="{sector_attr}" data-win-rate="{wr_attr}" 
                    data-horizon="{horizon_attr}" data-patterns="{patterns_attr}">
                  <td class="px-3 py-2 text-center">
                    <input type="checkbox" name="sig_idx" value="{i}" checked
                      class="sig-checkbox w-4 h-4 rounded border-gray-300 text-purple-600 focus:ring-purple-500">
                  </td>
                  <td class="px-3 py-2 font-medium text-gray-800">{_e(sig.get("ticker",""))}</td>
                  <td class="px-3 py-2 text-center">
                    <span class="text-{dir_color}-600 font-semibold">{dir_icon} {_e(sig.get("direction",""))}</span>
                  </td>
                  <td class="px-3 py-2 text-center">
                    <span class="px-2 py-0.5 rounded-full text-xs bg-blue-50 text-blue-700">{horizon_attr}</span>
                  </td>
                  <td class="px-3 py-2 text-right text-gray-700">{sig.get("entry_price",0):.2f}</td>
                  <td class="px-3 py-2 text-right font-medium">{cmp_html}</td>
                  <td class="px-3 py-2 text-right text-emerald-600">{sig.get("target_price",0):.2f}</td>
                  <td class="px-3 py-2 text-right text-red-600">{sig.get("sl_price",0):.2f}</td>
                  <td class="px-3 py-2 text-center">
                    <span class="text-{wr_color}-600 font-medium">{wr:.0f}%</span>
                  </td>
                  <td class="px-3 py-2 text-center text-gray-700">{rr_val}</td>
                  <td class="px-3 py-2 text-center">
                    <span class="px-2 py-0.5 rounded-full text-xs bg-{conf_color}-50 text-{conf_color}-700">{conf_attr}</span>
                  </td>
                  <td class="px-3 py-2 text-xs text-gray-500">{sector_attr}</td>
                  <td class="px-3 py-2 text-xs text-gray-500">{patterns_attr}</td>
                </tr>'''

            # Build filter dropdowns HTML
            conf_opts_html = "".join(f'<option value="{c}">{c}</option>' for c in confidents)
            sector_opts_html = "".join(f'<option value="{s}">{s}</option>' for s in sectors)
            horizon_opts_html = "".join(f'<option value="{h}">{h}</option>' for h in horizons)
            pattern_opts_html = "".join(f'<option value="{p}">{p}</option>' for p in patterns)
            win_opts_html = "".join(f'<option value="{r[1]}|{r[2]}" {"selected" if r[1] == 55 and r[2] is None else ""}>{r[0]}</option>' for r in win_ranges)
            
            filters_html = f'''
              <div class="grid grid-cols-5 gap-3 mb-4 p-4 bg-gray-50 rounded-lg border border-gray-200">
                <div>
                  <label class="text-xs text-gray-600 font-semibold">Confidence</label>
                  <select id="filter-conf" class="w-full px-2 py-1.5 text-sm border border-gray-300 rounded-lg">
                    <option value="">All</option>
                    {conf_opts_html}
                  </select>
                </div>
                <div>
                  <label class="text-xs text-gray-600 font-semibold">Win %</label>
                  <select id="filter-wr" class="w-full px-2 py-1.5 text-sm border border-gray-300 rounded-lg">
                    {win_opts_html}
                  </select>
                </div>
                <div>
                  <label class="text-xs text-gray-600 font-semibold">Sector</label>
                  <select id="filter-sector" class="w-full px-2 py-1.5 text-sm border border-gray-300 rounded-lg">
                    <option value="">All</option>
                    {sector_opts_html}
                  </select>
                </div>
                <div>
                  <label class="text-xs text-gray-600 font-semibold">Horizon</label>
                  <select id="filter-horizon" class="w-full px-2 py-1.5 text-sm border border-gray-300 rounded-lg">
                    <option value="">All</option>
                    {horizon_opts_html}
                  </select>
                </div>
                <div>
                  <label class="text-xs text-gray-600 font-semibold">Pattern</label>
                  <select id="filter-pattern" class="w-full px-2 py-1.5 text-sm border border-gray-300 rounded-lg">
                    <option value="">All</option>
                    {pattern_opts_html}
                  </select>
                </div>
              </div>
            '''

            signals_table = f'''
            <div class="glass rounded-xl p-5 mb-6 border-purple-200 fade-in">
              <div class="flex items-center justify-between mb-4">
                <div class="flex items-center gap-3">
                  <h3 class="text-lg font-semibold text-gray-800">Qualifying Signals ({qualifying})</h3>
                  <label class="flex items-center gap-2 text-sm text-gray-600 cursor-pointer">
                    <input type="checkbox" id="select-all" checked
                      class="w-4 h-4 rounded border-gray-300 text-purple-600 focus:ring-purple-500">
                    Select All
                  </label>
                </div>
                <div class="flex gap-2">
                  <button onclick="approveSelected()" class="px-4 py-2 rounded-lg bg-emerald-600 text-white text-sm font-medium hover:bg-emerald-700 transition">
                    Approve Selected
                  </button>
                  <button onclick="approveAll()" class="px-4 py-2 rounded-lg bg-purple-600 text-white text-sm font-medium hover:bg-purple-700 transition">
                    Approve All
                  </button>
                  <form method="POST" action="/engine?action=discard" style="display:inline">
                    <button type="submit" class="px-4 py-2 rounded-lg bg-red-100 text-red-700 text-sm font-medium hover:bg-red-200 transition">
                      Discard All
                    </button>
                  </form>
                </div>
              </div>
              {filters_html}
              
              <div class="overflow-x-auto scrollbar-thin">
                <table class="w-full text-sm">
                  <thead>
                    <tr class="border-b border-gray-200">
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase w-10"></th>
                      <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Ticker</th>
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Direction</th>
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Horizon</th>
                      <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Entry</th>
                      <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">CMP</th>
                      <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Target</th>
                      <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">SL</th>
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Win%</th>
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">R:R</th>
                      <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Conf</th>
                      <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Sector</th>
                      <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Patterns</th>
                    </tr>
                  </thead>
                  <tbody id="signals-tbody">{sig_rows}</tbody>
                </table>
              </div>
            </div>'''
        else:
            signals_table = ""

        # Skipped signals (collapsible)
        skipped_html = ""
        if skipped:
            skip_rows = ""
            for sig in skipped:
                dir_color = "emerald"
                dir_icon = "&#9650;"
                reasons = "; ".join(sig.get("skip_reasons", []))
                skip_rows += f'''
                <tr class="hover:bg-red-50/30 border-b border-gray-100 text-gray-400">
                  <td class="px-3 py-1.5">{_e(sig.get("ticker",""))}</td>
                  <td class="px-3 py-1.5 text-center">
                    <span class="text-{dir_color}-400">{dir_icon}</span>
                  </td>
                  <td class="px-3 py-1.5 text-center text-xs">{_e(sig.get("horizon_label",""))}</td>
                  <td class="px-3 py-1.5 text-right">{sig.get("entry_price",0):.2f}</td>
                  <td class="px-3 py-1.5 text-xs text-red-500">{_e(reasons)}</td>
                </tr>'''

            skipped_html = f'''
            <div class="glass rounded-xl p-5 mb-6 border-red-100 fade-in">
              <details>
                <summary class="cursor-pointer text-sm font-semibold text-gray-600 hover:text-gray-800 transition">
                  Filtered Out Signals ({filtered_out}) — click to expand
                </summary>
                <div class="overflow-x-auto scrollbar-thin mt-3">
                  <table class="w-full text-sm">
                    <thead>
                      <tr class="border-b border-gray-200">
                        <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Ticker</th>
                        <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Dir</th>
                        <th class="px-3 py-2 text-center text-xs text-gray-500 uppercase">Horizon</th>
                        <th class="px-3 py-2 text-right text-xs text-gray-500 uppercase">Entry</th>
                        <th class="px-3 py-2 text-left text-xs text-gray-500 uppercase">Skip Reason(s)</th>
                      </tr>
                    </thead>
                    <tbody>{skip_rows}</tbody>
                  </table>
                </div>
              </details>
            </div>'''

        # JS for select all & approve actions + filtering
        review_js = '''
        <script>
        // Filter functionality
        function applyFilters() {
          const confFilter = document.getElementById('filter-conf').value;
          const wrFilter = document.getElementById('filter-wr').value;
          const sectorFilter = document.getElementById('filter-sector').value;
          const horizonFilter = document.getElementById('filter-horizon').value;
          const patternFilter = document.getElementById('filter-pattern').value;
          
          const rows = document.querySelectorAll('.signal-row');
          let visibleCount = 0;
          
          rows.forEach(row => {
            let show = true;
            
            // Confidence filter
            if (confFilter && row.dataset.conf !== confFilter) {
              show = false;
            }
            
            // Win rate filter (range)
            if (show && wrFilter) {
              const [minWr, maxWr] = wrFilter.split('|');
              const rowWr = parseInt(row.dataset.winRate);
              const min = minWr ? parseInt(minWr) : 0;
              const max = maxWr ? parseInt(maxWr) : 100;
              if (rowWr < min || rowWr > max) {
                show = false;
              }
            }
            
            // Sector filter
            if (show && sectorFilter && row.dataset.sector !== sectorFilter) {
              show = false;
            }
            
            // Horizon filter
            if (show && horizonFilter && row.dataset.horizon !== horizonFilter) {
              show = false;
            }
            
            // Pattern filter (check if pattern exists in row)
            if (show && patternFilter) {
              const patterns = row.dataset.patterns;
              if (!patterns.includes(patternFilter)) {
                show = false;
              }
            }
            
            row.style.display = show ? '' : 'none';
            if (show) visibleCount++;
          });
          
          // Update "Select All" state
          updateSelectAll();
        }
        
        // Apply default Win%>=55 filter on load, then wire up change listeners
        applyFilters();
        ['filter-conf', 'filter-wr', 'filter-sector', 'filter-horizon', 'filter-pattern'].forEach(id => {
          const elem = document.getElementById(id);
          if (elem) {
            elem.addEventListener('change', applyFilters);
          }
        });

        function updateSelectAll() {
          const all = document.querySelectorAll('.signal-row:not([style*="display: none"]) .sig-checkbox');
          const checked = document.querySelectorAll('.signal-row:not([style*="display: none"]) .sig-checkbox:checked');
          const selectAllCheckbox = document.getElementById('select-all');
          if (selectAllCheckbox) {
            selectAllCheckbox.checked = all.length > 0 && all.length === checked.length;
          }
        }

        document.getElementById('select-all').addEventListener('change', function() {
          document.querySelectorAll('.signal-row:not([style*="display: none"]) .sig-checkbox').forEach(cb => cb.checked = this.checked);
        });
        document.querySelectorAll('.sig-checkbox').forEach(cb => {
          cb.addEventListener('change', function() {
            updateSelectAll();
          });
        });

        function approveSelected() {
          const checked = document.querySelectorAll('.sig-checkbox:checked');
          if (checked.length === 0) { alert('No signals selected'); return; }
          const indices = Array.from(checked).map(cb => cb.value).join(',');
          const form = document.createElement('form');
          form.method = 'POST';
          form.action = '/engine?action=approve&indices=' + indices;
          document.body.appendChild(form);
          form.submit();
        }

        function approveAll() {
          if (!confirm('Approve all qualifying signals?')) return;
          const form = document.createElement('form');
          form.method = 'POST';
          form.action = '/engine?action=approve';
          document.body.appendChild(form);
          form.submit();
        }
        </script>'''

        review_html = summary_bar + signals_table + skipped_html + review_js

    body = f'''
    <h2 class="text-2xl font-bold text-gray-800 mb-2">Engine Control</h2>
    <p class="text-sm text-gray-500 mb-6">Run the paper trading engine manually or view logs</p>
    {bearish_widget}
    {buttons}
    {live_html}
    {review_html}
    {scan_html}
    {log_html}'''
    return page_shell("Engine Control", "engine", body)


def _read_paper_trader_config():
    """Read current filter values from paper_trader.py. Non-blocking, cached read."""
    import re
    pt_path = os.path.join(SCRIPT_DIR, "paper_trader.py")
    config = {
        "MIN_WIN_RATE": 35.0,
        "MIN_CONFIDENCE": "MEDIUM",
        "MIN_RR_RATIO": 1.5,
        "MIN_MATCHES": 5,
        "MARKET_DECLINE_THRESHOLD_PCT": -1.0,
        "MARKET_DECLINE_BULLISH_MULTIPLIER": 0.7,
        "META_CLASSIFIER_PROBABILITY_THRESHOLD": 0.55,
    }
    try:
        with open(pt_path, "r", encoding="utf-8") as f:
            content = f.read()
        # Parse MIN_WIN_RATE
        m = re.search(r'MIN_WIN_RATE\s*=\s*([\d.]+)', content)
        if m:
            config["MIN_WIN_RATE"] = float(m.group(1))
        # Parse MIN_CONFIDENCE
        m = re.search(r'MIN_CONFIDENCE\s*=\s*["\']([^"\']+)["\']', content)
        if m:
            config["MIN_CONFIDENCE"] = m.group(1)
        # Parse MIN_RR_RATIO
        m = re.search(r'MIN_RR_RATIO\s*=\s*([\d.]+)', content)
        if m:
            config["MIN_RR_RATIO"] = float(m.group(1))
    except Exception as e:
        logger.warning(f"Failed to read paper_trader.py config: {e}")
    return config


def render_filters():
    """Render the Filters page with all trading parameter controls."""
    try:
        import trading_config as tc
    except Exception:
        tc = None
    
    stats = q_stats()
    
    # Read current filter values from PRODUCTION_FILTERS dict
    pf = getattr(tc, 'PRODUCTION_FILTERS', {}) if tc else {}
    min_wr = pf.get('min_win_rate', 55.0)
    min_conf = pf.get('min_confidence', 'MEDIUM')
    min_rr = pf.get('min_rr_ratio', 1.5)
    min_edge = pf.get('min_edge_pct', 8.5)
    
    # Get horizon edge thresholds
    horizon_thresholds = getattr(tc, 'HORIZON_EDGE_THRESHOLDS', {}) if tc else {}
    horizon_html = ""
    for h in sorted(horizon_thresholds.keys()):
        thresh = horizon_thresholds[h]
        h_name = {1: 'BTST', 3: 'Swing-3D', 5: 'Swing-5D', 10: 'Swing-10D', 25: 'Swing-25D'}.get(h, f"H{h}")
        horizon_html += f'''
    <div class="border border-gray-200 rounded-lg p-3">
      <p class="text-sm font-semibold text-gray-800">{h_name}</p>
      <div class="flex gap-4 mt-1 text-xs text-gray-600">
        <span>Neutral Zone: <strong>{thresh['neutral_zone']:.1f}%</strong></span>
        <span>Min Edge: <strong>{thresh['prod_min_edge']:.1f}%</strong></span>
      </div>
    </div>'''
    
    body = f'''
<div class="mb-8">
  <h1 class="text-3xl font-bold text-gray-900 mb-2">Trading Filters</h1>
  <p class="text-base text-gray-600">All signal entry gates. Changes apply on next engine run.</p>
</div>

<!-- Current Status Card -->
<div class="glass rounded-xl p-6 mb-8 bg-gradient-to-r from-blue-50 to-indigo-50 border-blue-200">
  <div class="grid grid-cols-4 gap-4">
    <div>
      <p class="text-xs font-semibold text-blue-600 uppercase tracking-wide">Win Rate</p>
      <p class="text-2xl font-bold text-gray-900 mt-1">{stats['win_rate']:.1f}%</p>
    </div>
    <div>
      <p class="text-xs font-semibold text-emerald-600 uppercase tracking-wide">Open Trades</p>
      <p class="text-2xl font-bold text-gray-900 mt-1">{stats['open_trades']}</p>
    </div>
    <div>
      <p class="text-xs font-semibold text-amber-600 uppercase tracking-wide">Profit Factor</p>
      <p class="text-2xl font-bold text-gray-900 mt-1">{stats['profit_factor']:.2f}x</p>
    </div>
    <div>
      <p class="text-xs font-semibold text-purple-600 uppercase tracking-wide">Total Trades</p>
      <p class="text-2xl font-bold text-gray-900 mt-1">{stats['total_trades']}</p>
    </div>
  </div>
</div>

<!-- SECTION 1: Primary Entry Filters -->
<div class="glass rounded-xl p-6 mb-6 border-emerald-200">
  <h2 class="text-xl font-bold text-gray-800 mb-4 flex items-center gap-2">
    <svg class="w-6 h-6 text-emerald-600" fill="none" stroke="currentColor" viewBox="0 0 24 24">
      <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 4a1 1 0 011-1h16a1 1 0 011 1v2.586a1 1 0 01-.293.707l-6.414 6.414a1 1 0 00-.293.707V17l-4 4v-6.586a1 1 0 00-.293-.707L3.293 7.293A1 1 0 013 6.586V4z"/>
    </svg>
    Primary Entry Filters
  </h2>
  <p class="text-sm text-gray-600 mb-6">ALL signals must pass these gates before being staged for trading.</p>
  
  <div class="space-y-4">
    <!-- Minimum Win Rate -->
    <div class="border border-gray-200 rounded-lg p-4 hover:shadow-md transition-shadow">
      <div class="flex justify-between items-center mb-3">
        <label class="text-sm font-semibold text-gray-800">Minimum Win Rate</label>
        <span class="text-lg font-mono font-bold text-emerald-600 bg-emerald-50 px-3 py-1 rounded">{min_wr:.0f}%</span>
      </div>
      <input type="range" min="30" max="80" step="5" value="{min_wr}" disabled class="w-full h-2 bg-gray-300 rounded-lg">
      <p class="mt-2 text-xs text-gray-600">Current backtest: {stats['win_rate']:.1f}% | Lower threshold = more signals</p>
    </div>

    <!-- Minimum Confidence -->
    <div class="border border-gray-200 rounded-lg p-4 hover:shadow-md transition-shadow">
      <label class="text-sm font-semibold text-gray-800 block mb-3">Minimum Confidence Level</label>
      <div class="flex gap-2 text-xs">
        <span class="px-4 py-2 bg-red-50 border border-red-200 rounded text-red-700">LOW</span>
        <span class="px-4 py-2 bg-amber-100 border border-amber-300 rounded text-amber-900 font-bold">MEDIUM (current)</span>
        <span class="px-4 py-2 bg-green-50 border border-green-200 rounded text-green-700">HIGH</span>
      </div>
    </div>

    <!-- Minimum Risk:Reward -->
    <div class="border border-gray-200 rounded-lg p-4 hover:shadow-md transition-shadow">
      <div class="flex justify-between items-center mb-3">
        <label class="text-sm font-semibold text-gray-800">Minimum Risk:Reward Ratio</label>
        <span class="text-lg font-mono font-bold text-amber-600 bg-amber-50 px-3 py-1 rounded">{min_rr:.1f}x</span>
      </div>
      <input type="range" min="0.5" max="3.0" step="0.1" value="{min_rr}" disabled class="w-full h-2 bg-gray-300 rounded-lg">
      <p class="mt-2 text-xs text-gray-600">Typical: 1.5x–2.0x | Conservative: 2.0x–3.0x</p>
    </div>

    <!-- Minimum Edge -->
    <div class="border border-gray-200 rounded-lg p-4 hover:shadow-md transition-shadow">
      <div class="flex justify-between items-center mb-3">
        <label class="text-sm font-semibold text-gray-800">Minimum Absolute Edge</label>
        <span class="text-lg font-mono font-bold text-indigo-600 bg-indigo-50 px-3 py-1 rounded">{min_edge:.1f}%</span>
      </div>
      <input type="range" min="2" max="15" step="0.5" value="{min_edge}" disabled class="w-full h-2 bg-gray-300 rounded-lg">
      <p class="mt-2 text-xs text-gray-600">Minimum mathematical edge in signal probability vs breakeven</p>
    </div>
  </div>
</div>

<!-- SECTION 2: Per-Horizon Edge Requirements -->
<div class="glass rounded-xl p-6 border-blue-200">
  <h2 class="text-xl font-bold text-gray-800 mb-4 flex items-center gap-2">
    <svg class="w-6 h-6 text-blue-600" fill="none" stroke="currentColor" viewBox="0 0 24 24">
      <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M13 7h8m0 0v8m0-8l-8 8-4-4-6 6"/>
    </svg>
    Per-Horizon Edge Thresholds
  </h2>
  <p class="text-sm text-gray-600 mb-4">Different horizon (timeframe) requires different edge minimums due to noise.</p>
  
  <div class="grid grid-cols-2 md:grid-cols-3 gap-3">{horizon_html}
  </div>
  
  <p class="mt-4 text-xs text-gray-600 bg-blue-50 border border-blue-200 rounded p-3">
    <strong>How it works:</strong> BTST (1-day) needs only 6% edge because of lower noise. 25-day swings need 12% because they're exposed to more macro risk.
  </p>
</div>

<div class="mt-8 p-4 bg-yellow-50 border border-yellow-300 rounded-lg">
  <p class="text-sm text-yellow-900">
    <strong>⚙️ To Edit Filters:</strong><br>
    1. Open <code class="bg-white px-2 py-1 rounded text-yellow-700 font-mono">trading_config.py</code><br>
    2. Modify <code class="bg-white px-2 py-1 rounded text-yellow-700 font-mono">PRODUCTION_FILTERS</code> or <code class="bg-white px-2 py-1 rounded text-yellow-700 font-mono">HORIZON_EDGE_THRESHOLDS</code><br>
    3. Restart paper trader for changes to apply
  </p>
</div>'''

    return page_shell("Filters", "filters", body)


# ============================================================
# HTTP REQUEST HANDLER
# ============================================================
class DashboardHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path.strip("/")

        routes = {
            "": ("dashboard", render_dashboard),
            "dashboard": ("dashboard", render_dashboard),
            "signals": ("signals", render_signals),
            "positions": ("positions", render_positions),
            "history": ("history", render_history),
            "market": ("market", render_market_indices),
            "performance": ("performance", render_performance),
            "engine": ("engine", lambda: render_engine()),
            "feedback": ("feedback", render_feedback),
            "filters": ("filters", render_filters),
        }

        if path == "api/bearish-score":
            # Live overnight bearish score JSON endpoint
            try:
                from global_sentiment import get_overnight_bearish_score
                score = get_overnight_bearish_score()
                status = "red" if score >= 70 else "yellow" if score >= 40 else "green"
                label  = "RED ALERT" if score >= 70 else "CAUTION" if score >= 40 else "SAFE"
                payload = json.dumps({
                    "score": score,
                    "status": status,
                    "label": label,
                    "btst_trim_active": score >= 70,
                    "intraday_trim_threshold": 25,
                    "timestamp": datetime.now().isoformat(),
                })
            except Exception as e:
                payload = json.dumps({"score": 30, "error": str(e)})
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/pending-trims":
            # Pending trim decisions JSON endpoint
            try:
                from startup_checkpoint import StartupCheckpoint
                cp = StartupCheckpoint()
                cp.ensure_trim_table()
                pending = cp.get_pending_trims()
                payload = json.dumps({
                    "count": len(pending),
                    "pending_trims": pending,
                    "timestamp": datetime.now().isoformat(),
                }, default=str)
            except Exception as e:
                payload = json.dumps({"count": 0, "error": str(e)})
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/early-exits":
            # Pending early-exit decisions JSON endpoint
            try:
                from startup_checkpoint import StartupCheckpoint
                import sqlite3
                cp = StartupCheckpoint()
                cp.ensure_trim_table()
                # Query early-exit decisions (trim_reason contains "Early-exit")
                conn = sqlite3.connect(cp.DB_PATH)
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("""
                    SELECT * FROM bearish_trim_decisions
                    WHERE execution_status = 'PENDING'
                      AND trim_reason LIKE '%Early-exit%'
                    ORDER BY decision_timestamp DESC
                """)
                early_exits = [dict(row) for row in cur.fetchall()]
                conn.close()
                payload = json.dumps({
                    "count": len(early_exits),
                    "pending_early_exits": early_exits,
                    "timestamp": datetime.now().isoformat(),
                }, default=str)
            except Exception as e:
                payload = json.dumps({"count": 0, "error": str(e)})
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/stats-by-date":
            # Date-range performance analytics endpoint
            try:
                qs = urllib.parse.parse_qs(parsed.query)
                from_date = qs.get("from", [None])[0]
                to_date = qs.get("to", [None])[0]
                if not from_date or not to_date:
                    raise ValueError("Both 'from' and 'to' query parameters are required")
                # Validate date format
                datetime.strptime(from_date, "%Y-%m-%d")
                datetime.strptime(to_date, "%Y-%m-%d")
                result = q_stats_for_range(from_date, to_date)
                payload = json.dumps(result, default=str)
                self.send_response(200)
            except ValueError as ve:
                payload = json.dumps({"error": str(ve)})
                self.send_response(400)
            except Exception as e:
                payload = json.dumps({"error": str(e)})
                self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/analytics":
            # A2 + A3: Benchmark alpha vs Nifty50 and Bootstrap CI on live WR/PF
            try:
                from paper_trader import PaperTrader as _PT, PaperTradeDB, DB_PATH
                pt = _PT.__new__(_PT)
                pt.db = PaperTradeDB(DB_PATH)
                alpha = pt.compute_benchmark_alpha()
                ci = pt.compute_bootstrap_ci()
                payload = json.dumps({"benchmark_alpha": alpha, "bootstrap_ci": ci}, default=str)
                self.send_response(200)
            except Exception as e:
                payload = json.dumps({"error": str(e)})
                self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/short-trades":
            try:
                import sqlite3
                db_path = "paper_trades/paper_trades.db"
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("""
                    SELECT id, ticker, entry_price, target_price, sl_price,
                           target_pct, sl_pct, rr_ratio, patterns, entry_date,
                           status, actual_return_pct, exit_price, exit_reason
                    FROM trades
                    WHERE direction = 'BEARISH' AND horizon_label = 'SHORT_1d'
                    ORDER BY entry_date DESC LIMIT 50
                """)
                short_trades = [dict(row) for row in cur.fetchall()]
                # Stats
                open_shorts = [t for t in short_trades if t['status'] == 'OPEN']
                closed = [t for t in short_trades if t['status'] not in ('OPEN',)]
                wins = [t for t in closed if (t.get('actual_return_pct') or 0) > 0]
                losses = [t for t in closed if (t.get('actual_return_pct') or 0) <= 0]
                conn.close()
                payload = json.dumps({
                    "active_count": len(open_shorts),
                    "total_closed": len(closed),
                    "wins": len(wins),
                    "losses": len(losses),
                    "win_rate": round(100 * len(wins) / len(closed), 1) if closed else 0,
                    "active_shorts": open_shorts,
                    "recent_closed": closed[:10],
                    "timestamp": datetime.now().isoformat(),
                }, default=str)
            except Exception as e:
                payload = json.dumps({"active_count": 0, "error": str(e)})
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "api/short-force-closes":
            # Pending SHORT_1d force-close decisions (retroactive execution queue)
            try:
                from startup_checkpoint import StartupCheckpoint
                import sqlite3
                cp = StartupCheckpoint()
                cp.ensure_trim_table()
                conn = sqlite3.connect(cp.DB_PATH)
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("""
                    SELECT * FROM bearish_trim_decisions
                    WHERE execution_status = 'PENDING'
                      AND trim_reason LIKE '%Short-force-close%'
                    ORDER BY decision_timestamp DESC
                """)
                pending = [dict(row) for row in cur.fetchall()]
                conn.close()
                payload = json.dumps({
                    "count": len(pending),
                    "pending_force_closes": pending,
                    "timestamp": datetime.now().isoformat(),
                }, default=str)
            except Exception as e:
                payload = json.dumps({"count": 0, "error": str(e)})
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(payload.encode("utf-8"))
            return

        if path == "history/export":
            try:
                xlsx = _history_xlsx_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", "attachment; filename=traqo_closed_trades.xlsx")
                self.send_header("Cache-Control", "no-cache")
                self.end_headers()
                self.wfile.write(xlsx)
            except ImportError as e:
                error_msg = f"Excel export requires 'openpyxl'. Install with: pip install openpyxl"
                logger.error(f"Excel export failed: {error_msg}")
                self.send_response(500)
                self.send_header("Content-Type", "text/plain; charset=utf-8")
                self.end_headers()
                self.wfile.write(error_msg.encode("utf-8"))
            except Exception as e:
                error_msg = f"Error generating Excel export: {str(e)}"
                logger.error(error_msg)
                self.send_response(500)
                self.send_header("Content-Type", "text/plain; charset=utf-8")
                self.end_headers()
                self.wfile.write(error_msg.encode("utf-8"))
        elif path == "feedback/download":
            # CSV download of feedback log
            csv_bytes = _feedback_csv_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", "attachment; filename=traqo_feedback_log.csv")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(csv_bytes)
        elif path == "engine/stream":
            # JSON endpoint for live polling
            status = get_engine_status()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(status).encode("utf-8"))
        elif path == "engine/pending":
            # JSON endpoint: check if pending signals file exists
            pending = get_pending_signals()
            result = {"has_pending": pending is not None}
            if pending:
                result["qualifying"] = pending.get("qualifying", 0)
                result["filtered_out"] = pending.get("filtered_out", 0)
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(result).encode("utf-8"))
        elif path in routes:
            _, renderer = routes[path]
            try:
                html = renderer()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html.encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(f"<h1>Error</h1><pre>{_e(str(e))}</pre>".encode("utf-8"))
        elif path == "favicon.ico":
            self.send_response(204)
            self.end_headers()
        else:
            self.send_response(302)
            self.send_header("Location", "/dashboard")
            self.end_headers()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path.strip("/")
        params = urllib.parse.parse_qs(parsed.query)

        if path == "feedback/delete":
            # Delete selected feedback entries
            try:
                content_len = int(self.headers.get('Content-Length', 0))
                post_data = self.rfile.read(content_len).decode('utf-8')
                parsed_params = urllib.parse.parse_qs(post_data)
                indices_str = parsed_params.get('indices', [''])[0]
                indices = [int(x.strip()) for x in indices_str.split(',') if x.strip().isdigit()]
                
                if not indices:
                    result = {"status": "error", "message": "No valid indices provided"}
                elif len(indices) > 10:
                    result = {"status": "error", "message": "Maximum 10 entries can be deleted at once"}
                else:
                    result = _delete_feedback_entries(indices)
            except Exception as e:
                result = {"status": "error", "message": str(e)}
            
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(result).encode("utf-8"))
            return

        if path == "settings/save":
            # Save filter settings to paper_trader.py
            try:
                content_len = int(self.headers.get('Content-Length', 0))
                post_data = self.rfile.read(content_len).decode('utf-8')
                form_data = urllib.parse.parse_qs(post_data)
                
                # Extract values (safely with defaults)
                min_wr = float(form_data.get('min_win_rate', ['35.0'])[0])
                min_conf = form_data.get('min_confidence', ['MEDIUM'])[0]
                min_rr = float(form_data.get('min_rr_ratio', ['1.5'])[0])
                fb_pen_thresh = int(form_data.get('feedback_penalty_threshold', ['30'])[0])
                hz_pen_thresh = int(form_data.get('horizon_penalty_threshold', ['25'])[0])
                
                # Validate ranges
                if not (15 <= min_wr <= 70):
                    raise ValueError("MIN_WIN_RATE must be 15-70")
                if min_conf not in ("LOW", "MEDIUM", "HIGH"):
                    raise ValueError("MIN_CONFIDENCE must be LOW, MEDIUM, or HIGH")
                if not (0.5 <= min_rr <= 3.0):
                    raise ValueError("MIN_RR_RATIO must be 0.5-3.0")
                if not (15 <= fb_pen_thresh <= 50):
                    raise ValueError("Feedback penalty threshold must be 15-50")
                if not (10 <= hz_pen_thresh <= 40):
                    raise ValueError("Horizon penalty threshold must be 10-40")
                
                # Read paper_trader.py
                pt_path = os.path.join(SCRIPT_DIR, "paper_trader.py")
                with open(pt_path, "r", encoding="utf-8") as f:
                    content = f.read()
                
                # Replace values using regex
                import re
                content = re.sub(r'MIN_WIN_RATE\s*=\s*[\d.]+', f'MIN_WIN_RATE = {min_wr}', content)
                content = re.sub(r'MIN_CONFIDENCE\s*=\s*["\']([^"\']+)["\']', f'MIN_CONFIDENCE = "{min_conf}"', content)
                content = re.sub(r'MIN_RR_RATIO\s*=\s*[\d.]+', f'MIN_RR_RATIO = {min_rr}', content)
                
                # Update feedback/horizon thresholds
                content = re.sub(r'if wr < \d+:', f'if wr < {fb_pen_thresh}:', content)
                content = re.sub(r'if wr < \d+:', f'if wr < {hz_pen_thresh}:', content) # Will need 2 replacements
                
                # Write back (SAFE: backup concept)
                with open(pt_path, "w", encoding="utf-8") as f:
                    f.write(content)
                
                result = {"status": "success", "message": "Settings saved"}
                logger.info(f"✓ Settings saved: MIN_WR={min_wr}%, CONF={min_conf}, R:R={min_rr}")
            except Exception as e:
                result = {"status": "error", "error": str(e)}
                logger.error(f"Failed to save settings: {e}")
            
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(result).encode("utf-8"))
            return

        if path == "engine":
            action = params.get("action", ["run"])[0]
            if action in ("run", "scan", "monitor", "scan_preview"):
                start_engine(action)
                # Redirect to GET /engine so user sees live output
                self.send_response(302)
                self.send_header("Location", "/engine")
                self.end_headers()
                return
            elif action == "approve":
                indices_str = params.get("indices", [None])[0]
                extra = [indices_str] if indices_str else None
                start_engine("approve", extra_args=extra)
                self.send_response(302)
                self.send_header("Location", "/engine")
                self.end_headers()
                return
            elif action == "discard":
                # Quick operation — just delete staging file, no engine needed
                if os.path.exists(PENDING_SIGNALS_FILE):
                    os.remove(PENDING_SIGNALS_FILE)
                self.send_response(302)
                self.send_header("Location", "/engine")
                self.end_headers()
                return

        elif path == "health":
            # Health check endpoint for debugging
            health_status = {
                "yfinance_available": _HAS_YF,
                "yfinance_version": yf.__version__ if _HAS_YF else "Not available",
                "open_trades_count": len(q_open_trades()),
                "timestamp": datetime.now().isoformat()
            }
            
            if _HAS_YF:
                # Test a quick price fetch
                try:
                    test_data = yf.download("SBIN.NS", period="1d", progress=False, multi_level_index=False)
                    health_status["price_test"] = "✅ SUCCESS" if not test_data.empty else "❌ Empty data"
                except Exception as e:
                    health_status["price_test"] = f"❌ ERROR: {e}"
            else:
                health_status["price_test"] = "❌ yfinance not available"
                
            response_html = f"""
            <!DOCTYPE html>
            <html><head><title>Traqo Health Check</title>
            <style>body{{font-family:monospace; padding:20px; background:#f5f5f5;}} 
            .status{{padding:10px; margin:5px; border-radius:5px; background:white;}}</style>
            </head><body>
            <h1>🏥 Traqo Health Check</h1>
            <div class="status"><strong>yfinance Available:</strong> {health_status['yfinance_available']}</div>
            <div class="status"><strong>yfinance Version:</strong> {health_status['yfinance_version']}</div>
            <div class="status"><strong>Price Test:</strong> {health_status['price_test']}</div>
            <div class="status"><strong>Open Trades:</strong> {health_status['open_trades_count']}</div>
            <div class="status"><strong>Timestamp:</strong> {health_status['timestamp']}</div>
            <p><a href="/">← Back to Dashboard</a></p>
            </body></html>
            """
            self.send_response(200)
            self.send_header("Content-type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(response_html.encode("utf-8"))
            return

        elif path == "trade/cancel":
            try:
                trade_id = int(params.get("id", [0])[0])
                if trade_id > 0:
                    cancel_trade(trade_id)
            except Exception:
                pass
            self.send_response(302)
            self.send_header("Location", "/positions")
            self.end_headers()
            return

        elif path == "trade/cancel-bulk":
            try:
                ids_str = params.get("ids", [""])[0]
                ids = [int(x.strip()) for x in ids_str.split(",") if x.strip().isdigit()]
                if ids:
                    cancel_trades_bulk(ids)
            except Exception:
                pass
            self.send_response(302)
            self.send_header("Location", "/positions")
            self.end_headers()
            return

        elif path == "trade/purge":
            # Directly delete closed trades — no external imports, no hanging
            deleted = 0
            fb_removed = 0
            try:
                content_len = int(self.headers.get('Content-Length', 0))
                post_data = self.rfile.read(content_len).decode('utf-8')
                parsed_params = urllib.parse.parse_qs(post_data)
                ids_str = parsed_params.get('ids', [''])[0]
                ids = [int(x.strip()) for x in ids_str.split(",") if x.strip().isdigit()]

                if ids:
                    db_full = os.path.join(SCRIPT_DIR, DB_PATH)
                    placeholders = ",".join("?" * len(ids))

                    # Step 1: Delete from SQLite — trades + position_monitoring
                    conn = sqlite3.connect(db_full)
                    conn.execute(f"DELETE FROM position_monitoring WHERE trade_id IN ({placeholders})", ids)
                    conn.execute(f"DELETE FROM trades WHERE id IN ({placeholders})", ids)
                    conn.commit()
                    conn.close()
                    deleted = len(ids)
                    logger.info(f"✓ Deleted {deleted} trade(s) from database: {ids}")

                    # Step 2: Remove from feedback_log.json
                    fb_path = os.path.join(SCRIPT_DIR, "feedback", "feedback_log.json")
                    if os.path.exists(fb_path):
                        with open(fb_path, "r", encoding="utf-8") as f:
                            feedback = json.load(f)
                        pids = {f"paper_{tid}" for tid in ids}
                        before = len(feedback)
                        feedback = [e for e in feedback if e.get("trade_id") not in pids]
                        fb_removed = before - len(feedback)
                        with open(fb_path, "w", encoding="utf-8") as f:
                            json.dump(feedback, f, indent=2, default=str)
                        logger.info(f"✓ Removed {fb_removed} feedback entries")

            except Exception as e:
                logger.error(f"trade/purge error: {e}", exc_info=True)

            self.send_response(302)
            self.send_header("Location", "/history")
            self.end_headers()
            return

        elif path == "trade/purge-date":
            # Purge all non-OPEN trades closed between from_date and to_date
            try:
                from_dt = params.get("from", [""])[0]
                to_dt = params.get("to", [from_dt])[0]
                if from_dt:
                    purge_trades_by_date(from_dt, to_dt)
            except Exception:
                pass
            self.send_response(302)
            self.send_header("Location", "/trades")
            self.end_headers()
            return

        self.send_response(302)
        self.send_header("Location", "/dashboard")
        self.end_headers()

    def log_message(self, format, *args):
        # Suppress default logging noise
        pass


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    port = 8521
    print(f"\n  Traqo — RAG Powered Quantitative Candlestick Intelligence")
    print(f"  http://localhost:{port}")
    
    # Display yfinance status prominently
    if _HAS_YF:
        print(f"  ✅ Live prices: ENABLED (yfinance v{yf.__version__})")
    else:
        print(f"  ❌ Live prices: DISABLED (yfinance not found)")
        print(f"  💡 Fix: Activate virtual environment and run 'pip install yfinance'")
    
    print(f"  🏥 Health check: http://localhost:{port}/health\n")

    server = ThreadingHTTPServer(("0.0.0.0", port), DashboardHandler)

    try:
        webbrowser.open(f"http://localhost:{port}")
    except Exception:
        pass

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.server_close()
